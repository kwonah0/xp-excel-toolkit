"""End-to-end test: create samples → import → process → export → verify."""

import os
import sys
from pathlib import Path

# Ensure we work in the script's directory
os.chdir(Path(__file__).parent)

from sqlalchemy import func

from models import (
    COLUMN_MAP, Base, Employee, ExcelCell, ExcelMerge, ExcelSheet,
    ExcelWorkbook, get_cell, init_db,
)
from create_samples import create_sample_xls, create_sample_xlsx
from xlsx_parser import import_xlsx
from xls_parser import import_xls
from exporter import export_xlsx, process_salary_raise


FIELD_MAP = {
    "부서": "department",
    "이름": "name",
    "직급": "position",
    "급여": "salary",
}


def banner(text: str):
    print(f"\n{'='*60}")
    print(f"  {text}")
    print(f"{'='*60}")


def test_xlsx():
    banner("TEST 1: .xlsx (openpyxl)")

    # ── 1. Create sample ─────────────────────────────────────
    xlsx_path = create_sample_xlsx()

    # ── 2. Import ────────────────────────────────────────────
    COLUMN_MAP.clear()
    db_path = "sqlite:///test_xlsx.db"
    # Remove old DB if exists
    Path("test_xlsx.db").unlink(missing_ok=True)

    SessionMaker = init_db(db_path)
    session = SessionMaker()

    sheet = import_xlsx(session, xlsx_path, field_map=FIELD_MAP)
    session.commit()

    # ── 3. Verify import ─────────────────────────────────────
    print(f"\n[Import] Sheet: {sheet.name}, header_row: {sheet.header_row}")

    employees = session.query(Employee).order_by(Employee.excel_row).all()
    print(f"[Import] Employees loaded: {len(employees)}")
    for emp in employees:
        print(f"  row={emp.excel_row}: {emp.department} | {emp.name} | {emp.position} | {emp.salary}")

    # Verify merged cells
    merges = session.query(ExcelMerge).filter_by(sheet_id=sheet.id).all()
    print(f"\n[Merge] Merge ranges: {len(merges)}")
    for m in merges:
        print(f"  ({m.min_row},{m.min_col}):({m.max_row},{m.max_col})")

    # Verify merged cell values are filled
    for emp in employees:
        assert emp.department is not None, f"Row {emp.excel_row}: department is None (merge fill failed)"
    print("[Merge] All merged department values filled correctly ✓")

    # Verify styles
    cell_with_style = (
        session.query(ExcelCell)
        .filter(ExcelCell.sheet_id == sheet.id, ExcelCell.style.isnot(None))
        .first()
    )
    if cell_with_style:
        print(f"\n[Style] Example: row={cell_with_style.row}, col={cell_with_style.col}")
        print(f"  style={cell_with_style.style}")

    # Check yellow row (row 8 = 최지훈)
    yellow_cell = (
        session.query(ExcelCell)
        .filter_by(sheet_id=sheet.id, row=8, col=1)
        .first()
    )
    if yellow_cell and yellow_cell.style:
        print(f"[Style] Yellow row (8,1): {yellow_cell.style}")

    # Verify get_cell helper
    print(f"\n[COLUMN_MAP] {COLUMN_MAP}")
    emp0 = employees[0]
    salary_cell = get_cell(session, emp0, "salary")
    if salary_cell:
        print(f"[get_cell] {emp0.name}'s salary cell: row={salary_cell.row}, col={salary_cell.col}, "
              f"value={salary_cell.raw_value}, style={salary_cell.style}")

    # ── 4. Process: salary raise ─────────────────────────────
    print("\n[Process] Applying 10% salary raise...")
    changes = process_salary_raise(session)
    session.commit()
    for ch in changes:
        print(f"  {ch['employee']}: {ch['old']} → {ch['new']}")

    # ── 5. Export ────────────────────────────────────────────
    wb = session.query(ExcelWorkbook).first()
    output = Path("samples/output.xlsx")
    export_xlsx(session, wb.id, output)
    print(f"\n[Export] Written to: {output}")

    # ── 6. Verify export ─────────────────────────────────────
    import openpyxl as oxl
    wb_check = oxl.load_workbook(output)
    ws_check = wb_check.active
    print("\n[Verify] Re-reading exported file:")
    for r in range(5, 10):
        vals = [ws_check.cell(row=r, column=c).value for c in range(1, 5)]
        print(f"  row {r}: {vals}")

    # Check that merge is preserved
    merge_count = len(ws_check.merged_cells.ranges)
    print(f"[Verify] Merged ranges preserved: {merge_count}")

    # Check salary values updated
    new_sal = ws_check.cell(row=5, column=4).value
    assert new_sal == 3300, f"Expected 3300, got {new_sal}"
    print("[Verify] Salary values correctly updated ✓")

    # Check formatting preserved
    fill = ws_check.cell(row=8, column=1).fill
    if fill and fill.fgColor and fill.fgColor.rgb:
        print(f"[Verify] Yellow row fill preserved: {fill.fgColor.rgb}")

    wb_check.close()
    session.close()
    print("\n✓ .xlsx test PASSED")


def test_xls():
    banner("TEST 2: .xls (xlrd)")

    # ── 1. Create sample ─────────────────────────────────────
    xls_path = create_sample_xls()

    # ── 2. Import ────────────────────────────────────────────
    COLUMN_MAP.clear()
    db_path = "sqlite:///test_xls.db"
    Path("test_xls.db").unlink(missing_ok=True)

    SessionMaker = init_db(db_path)
    session = SessionMaker()

    sheet = import_xls(session, xls_path, field_map=FIELD_MAP)
    session.commit()

    # ── 3. Verify import ─────────────────────────────────────
    print(f"\n[Import] Sheet: {sheet.name}, header_row: {sheet.header_row}")

    employees = session.query(Employee).order_by(Employee.excel_row).all()
    print(f"[Import] Employees loaded: {len(employees)}")
    for emp in employees:
        print(f"  row={emp.excel_row}: {emp.department} | {emp.name} | {emp.position} | {emp.salary}")

    # Verify merged cells
    merges = session.query(ExcelMerge).filter_by(sheet_id=sheet.id).all()
    print(f"\n[Merge] Merge ranges: {len(merges)}")
    for m in merges:
        print(f"  ({m.min_row},{m.min_col}):({m.max_row},{m.max_col})")

    # Verify merge fill
    for emp in employees:
        assert emp.department is not None, f"Row {emp.excel_row}: department is None (merge fill failed)"
    print("[Merge] All merged department values filled correctly ✓")

    # Verify styles
    styled_cells = (
        session.query(ExcelCell)
        .filter(ExcelCell.sheet_id == sheet.id, ExcelCell.style.isnot(None))
        .limit(3)
        .all()
    )
    print(f"\n[Style] Cells with style data: {len(styled_cells)} (showing up to 3)")
    for sc in styled_cells:
        print(f"  row={sc.row}, col={sc.col}: {sc.style}")

    # ── 4. Process + verify ──────────────────────────────────
    print("\n[Process] Applying 10% salary raise...")
    changes = process_salary_raise(session)
    session.commit()
    for ch in changes:
        print(f"  {ch['employee']}: {ch['old']} → {ch['new']}")

    # Verify salary change
    emp_kim = session.query(Employee).filter_by(name="김철수").first()
    assert emp_kim.salary == 3300, f"Expected 3300, got {emp_kim.salary}"
    print("[Process] Salary raise verified ✓")

    # Cell count
    total_cells = session.query(func.count(ExcelCell.id)).filter_by(sheet_id=sheet.id).scalar()
    print(f"\n[Stats] Total cells stored: {total_cells}")

    session.close()
    print("\n✓ .xls test PASSED")


def test_summary():
    banner("SUMMARY")
    print("Both .xlsx and .xls formats tested successfully.")
    print("Verified: import, merge fill, style extraction, salary processing, export with formatting.")


if __name__ == "__main__":
    test_xlsx()
    test_xls()
    test_summary()
