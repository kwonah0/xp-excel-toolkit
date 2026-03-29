"""Export modified data back to Excel, preserving original formatting."""

import io
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from sqlalchemy.orm import Session

from models import Employee, ExcelCell, ExcelSheet, ExcelWorkbook, COLUMN_MAP


def process_salary_raise(session: Session, raise_pct: float = 0.10) -> list[dict]:
    """Apply salary raise and return change log."""
    changes: list[dict] = []
    employees = session.query(Employee).all()

    for emp in employees:
        if emp.salary:
            old_salary = emp.salary
            emp.salary = round(emp.salary * (1 + raise_pct))
            changes.append({
                "employee": emp.name,
                "field": "salary",
                "old": old_salary,
                "new": emp.salary,
                "excel_row": emp.excel_row,
            })

    session.flush()
    return changes


def export_xlsx(session: Session, workbook_id: int, output_path: str | Path) -> Path:
    """
    Write modified data back to .xlsx, restoring original formatting from BLOB.

    Strategy: Load original BLOB → overwrite only changed cell values → save.
    This preserves 100% of the original formatting, merges, charts, etc.
    """
    output_path = Path(output_path)

    wb_obj = session.query(ExcelWorkbook).get(workbook_id)
    if not wb_obj or not wb_obj.blob:
        raise ValueError(f"Workbook {workbook_id} not found or has no BLOB")

    # Load original workbook from stored BLOB
    wb = openpyxl.load_workbook(io.BytesIO(wb_obj.blob))

    for sheet_obj in wb_obj.sheets:
        ws = wb[sheet_obj.name]

        if not sheet_obj.header_row:
            continue

        # Get all employees for this sheet
        employees = (
            session.query(Employee)
            .filter_by(sheet_id=sheet_obj.id)
            .all()
        )

        # Build set of merge-origin cells for quick lookup
        merge_origins: set[tuple[int, int]] = set()
        for mr in ws.merged_cells.ranges:
            merge_origins.add((mr.min_row, mr.min_col))

        for emp in employees:
            for field_name, col_idx in COLUMN_MAP.items():
                val = getattr(emp, field_name, None)
                if val is None:
                    continue
                cell = ws.cell(row=emp.excel_row, column=col_idx)
                # Skip non-origin merged cells (read-only in openpyxl)
                if isinstance(cell, MergedCell):
                    # Write to the merge origin instead
                    for mr in ws.merged_cells.ranges:
                        if (mr.min_row <= emp.excel_row <= mr.max_row
                                and mr.min_col <= col_idx <= mr.max_col):
                            ws.cell(row=mr.min_row, column=mr.min_col).value = val
                            break
                else:
                    cell.value = val

    wb.save(output_path)
    wb.close()
    return output_path
