"""Generate a tiny synthetic xlsx so the demo is self-contained."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side


def make_sample(path: Path) -> Path:
    """Write a small Pinmap-style xlsx with merges, styles, and a comment.

    Two sheets:
      - ``Pinmap_A``: title row (horizontally merged) + header + 4 data rows,
        one vertical merge in the Dir column, yellow header fill, a comment
        on the VDD name cell.
      - ``Notes``: plain free-form text (exercises non-domain passthrough).
    """
    wb = openpyxl.Workbook()

    # ── Pinmap sheet ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Pinmap_A"

    # Title row (will be horizontally merged across 3 cols)
    ws.cell(row=1, column=1, value="Pinmap for Chip A").font = Font(bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    # Header row at row 2 — yellow fill + bold
    for col, h in enumerate(["Pin", "Name", "Dir"], start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Data rows
    for i, (pn, nm, di) in enumerate(
        [
            ("A1", "VDD",  "PWR"),
            ("A2", "GND",  "GND"),
            ("A3", "SDA",  "I/O"),
            ("A4", "SCL",  "I/O"),
        ],
        start=3,
    ):
        ws.cell(row=i, column=1, value=pn)
        ws.cell(row=i, column=2, value=nm)
        ws.cell(row=i, column=3, value=di)

    # Vertical merge on Dir column for I/O group (rows 5-6)
    ws.merge_cells(start_row=5, start_column=3, end_row=6, end_column=3)

    # VDD name cell: top border + comment
    vdd = ws.cell(row=3, column=2)
    vdd.border = Border(top=Side(style="thin"))
    vdd.comment = Comment("Supply pin", "demo")

    # ── Notes sheet ───────────────────────────────────────────────
    notes = wb.create_sheet("Notes")
    notes.cell(row=1, column=1, value="Free-form notes")
    notes.cell(row=2, column=1, value="line 1")
    notes.cell(row=3, column=1, value="line 2")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
    return path


if __name__ == "__main__":
    out = Path(__file__).with_name("sample.xlsx")
    make_sample(out)
    print(f"wrote {out}")
