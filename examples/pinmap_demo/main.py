"""End-to-end demo: build sample xlsx → init_db → import via pinmap helpers
→ inspect/modify domain rows → round-trip export → verify.

Run from anywhere::

    uv run python examples/pinmap_demo/main.py

Note: this file imports **only from pinmap.api** — never from
excel_toolkit directly. That's the facade pattern in action: downstream
applications shouldn't have to know which library actually carries the
infra.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl

from make_sample import make_sample
from pinmap.api import (
    ChangeLog,
    ExcelCell,
    PinEntry,
    export_pinmap,
    import_pinmap,
    init_db,
)


def main() -> None:
    with tempfile.TemporaryDirectory() as td:
        workdir = Path(td)
        sample = make_sample(workdir / "sample.xlsx")
        modified = workdir / "modified.xlsx"
        db_url = f"sqlite:///{workdir / 'demo.db'}"

        # ── 1. Spin up the DB through the host facade.
        #       Behind the scenes pinmap.api.init_db is excel_toolkit's
        #       init_db; pinmap.api.Base shares its MetaData with
        #       excel_toolkit.Base, so the infra tables AND pin_entry
        #       are created in one create_all() pass.
        Session = init_db(db_url)

        with Session() as session:
            # ── 2. Import via the thin host wrapper (which already
            #       knows the SheetConfig for the "Pinmap_*" pattern).
            import_pinmap(session, sample)
            session.commit()

            # ── 3. Domain rows are real ORM objects with excel-row tracking.
            print("── Imported PinEntry rows ──")
            for row in session.query(PinEntry).order_by(PinEntry.excel_row):
                print(f"  excel_row={row.excel_row}  "
                      f"{row.pin_no:>3}  {row.name:>5}  {row.direction}")

            # Merge fill: A4/SCL inherits "I/O" from the vertical-merge origin.
            scl = session.query(PinEntry).filter_by(pin_no="A4").one()
            assert scl.direction == "I/O", "merge fill failed"

            # Cell-level access (style/comment/merge) — also re-exported
            # through pinmap.api so the host owns the namespace.
            vdd_cell = (
                session.query(ExcelCell)
                .filter_by(row=3, col=2)
                .one()
            )
            print("── VDD cell metadata ──")
            print(f"  comment: {vdd_cell.comment!r}")
            print(f"  style:   {vdd_cell.style}")

            # ── 4. Mutate a couple of domain rows. The SQLite trigger
            #       registered for pin_entry will record the changes in
            #       change_log. We pick A1/VDD's direction (outside the
            #       vertical-merge range) to keep the example clean —
            #       see the note in the README about merge ranges.
            vdd = session.query(PinEntry).filter_by(pin_no="A1").one()
            vdd.direction = "CORE"
            sda = session.query(PinEntry).filter_by(pin_no="A3").one()
            sda.name = "SDA_NEW"
            session.commit()

            print("── change_log entries ──")
            for log in session.query(ChangeLog).order_by(ChangeLog.id):
                print(f"  {log.operation}  pin_entry.{log.column_name}: "
                      f"{log.old_value!r} → {log.new_value!r}")

            # ── 5. Round-trip export through the host facade.
            export_pinmap(session, modified)

        # ── 6. Re-open the exported file and confirm the changes.
        wb = openpyxl.load_workbook(modified)
        ws = wb["Pinmap_A"]
        print("── Round-tripped xlsx ──")
        print(f"  sheets:        {wb.sheetnames}")
        print(f"  VDD dir  cell: {ws.cell(row=3, column=3).value!r}")
        print(f"  SDA name cell: {ws.cell(row=5, column=2).value!r}")
        print(f"  merge count:   {len(ws.merged_cells.ranges)} (title + Dir merge)")
        print(f"  header fill:   {ws.cell(row=2, column=1).fill.fgColor.rgb}")
        notes = wb["Notes"]
        print(f"  notes A1:      {notes.cell(row=1, column=1).value!r}")
        wb.close()

        assert ws.cell(row=3, column=3).value == "CORE"
        assert ws.cell(row=5, column=2).value == "SDA_NEW"
        assert len(ws.merged_cells.ranges) == 2
        print("\nOK")


if __name__ == "__main__":
    main()
