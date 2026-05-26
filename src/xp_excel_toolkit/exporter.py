"""Generic exporters from ExcelCell/ExcelMerge rows back to xlsx."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from sqlalchemy.orm import Session

from xp_excel_toolkit.models import ExcelCell, ExcelMerge, ExcelSheet


def apply_style(cell: Cell, style: dict | None) -> None:
    """Apply a style dict (from ExcelCell.style) to an openpyxl cell."""
    if not style:
        return

    bg = style.get("bg_color")
    if bg:
        color = bg.lstrip("#")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    font_kwargs = {}
    if style.get("font_bold"):
        font_kwargs["bold"] = True
    fc = style.get("font_color")
    if fc:
        font_kwargs["color"] = fc.lstrip("#")
    if font_kwargs:
        cell.font = Font(**font_kwargs)

    nf = style.get("number_format")
    if nf:
        cell.number_format = nf

    border_sides = {}
    for side_name in ("left", "right", "top", "bottom"):
        bs = style.get(f"border_{side_name}")
        if bs:
            border_sides[side_name] = Side(style=bs)
    if border_sides:
        cell.border = Border(**border_sides)


def export_from_cells(
    session: Session,
    sheet_id: int,
    output_path: str | Path,
) -> Path:
    """Build an .xlsx from ExcelCell + ExcelMerge records (no BLOB needed).

    Creates a fresh xlsx entirely from the DB cell data, suitable for
    partial exports like split-by-IP.
    """
    output_path = Path(output_path)

    sheet_obj = session.get(ExcelSheet, sheet_id)
    if not sheet_obj:
        raise ValueError(f"Sheet {sheet_id} not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_obj.name

    cells = (
        session.query(ExcelCell)
        .filter(ExcelCell.sheet_id == sheet_id)
        .all()
    )
    for cell_rec in cells:
        if cell_rec.formula_type == "array" and cell_rec.formula_ref:
            value = ArrayFormula(ref=cell_rec.formula_ref, text=cell_rec.raw_value)
        elif cell_rec.formula_type == "dataTable" and cell_rec.formula_ref:
            value = DataTableFormula(ref=cell_rec.formula_ref)
        else:
            value = cell_rec.raw_value
        c = ws.cell(row=cell_rec.row, column=cell_rec.col, value=value)
        apply_style(c, cell_rec.style)
        if cell_rec.comment:
            c.comment = Comment(cell_rec.comment, "")

    merges = (
        session.query(ExcelMerge)
        .filter(ExcelMerge.sheet_id == sheet_id)
        .all()
    )
    for m in merges:
        ws.merge_cells(
            start_row=m.min_row, start_column=m.min_col,
            end_row=m.max_row, end_column=m.max_col,
        )

    if sheet_obj.header_row:
        ws.freeze_panes = f"A{sheet_obj.header_row + 1}"

    wb.save(output_path)
    wb.close()
    return output_path
