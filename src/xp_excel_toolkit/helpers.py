"""DB-side helpers + merge-write conflict handling.

These read from the ExcelCell / ExcelSheet tables and return plain
Python structures — they never insert into domain tables.

Merge-write conflict handling solves the silent-loss case where multiple
cells in a single merged range receive disagreeing intended values during
export. xp_excel_toolkit provides the *mechanism* (detect + policy resolvers); the
caller picks the *policy*.
"""

from __future__ import annotations

from collections.abc import Iterable, Iterator
from dataclasses import dataclass
from typing import Any, Literal, TYPE_CHECKING

from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from xp_excel_toolkit.models import ExcelCell, ExcelMerge

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


# ── Header detection ────────────────────────────────────────────────

HeaderMatch = Literal["any", "all"]


def find_header_row_db(
    session: Session,
    sheet_id: int,
    expected_headers: Iterable[str],
    *,
    match: HeaderMatch = "any",
    max_scan: int = 30,
) -> int | None:
    """Auto-detect the header row from ExcelCell rows in the DB.

    Args:
        expected_headers: Header strings the row is expected to contain.
            Any iterable of strings (list, tuple, set, ``dict.keys()``).
            At least one non-empty string is required.
        match: ``"any"`` (default) returns the first row containing at
            least one of ``expected_headers``; ``"all"`` requires every
            expected header to be present.
        max_scan: Stop scanning after this row number.

    Returns:
        1-based row index, or ``None`` if no match is found within
        ``max_scan``.

    Raises:
        ValueError: if ``expected_headers`` is empty or contains only
            blank strings.
    """
    expected = {s.strip() for s in expected_headers if s and s.strip()}
    if not expected:
        raise ValueError(
            "expected_headers must contain at least one non-empty string"
        )

    rows = session.execute(
        select(ExcelCell.row, ExcelCell.raw_value)
        .where(ExcelCell.sheet_id == sheet_id, ExcelCell.row <= max_scan)
        .order_by(ExcelCell.row, ExcelCell.col)
    ).all()

    by_row: dict[int, set[str]] = {}
    for r, v in rows:
        if v and isinstance(v, str) and v.strip():
            by_row.setdefault(r, set()).add(v.strip())

    for r in sorted(by_row):
        values = by_row[r]
        if match == "all":
            if expected.issubset(values):
                return r
        else:  # "any"
            if values & expected:
                return r
    return None


def iter_rows_by_header(
    session: Session,
    sheet_id: int,
    header_row: int,
    *,
    headers: Iterable[str] | None = None,
) -> Iterator[tuple[int, dict[str, Any]]]:
    """Yield ``(row_idx, {header_str: value})`` for every non-empty data row.

    xp_excel_toolkit reads ``header_row`` internally to build a ``col_idx → header_str``
    mapping, then projects each subsequent row through it. The caller never
    needs to know column indices — only header strings.

    Layout assumption: single-row header, flat columns. Multi-row /
    compound / pivot layouts must use raw ``ExcelCell`` SELECTs.

    Args:
        session: SQLAlchemy session.
        sheet_id: ``ExcelSheet`` id.
        header_row: 1-based index of the header row.
        headers: Optional whitelist of header strings. If given, only
            these columns appear in the output dicts (and the SQL fetch
            is narrowed accordingly). If ``None``, every column whose
            header cell holds a non-empty string is included.

    Yields:
        ``(row_idx, cells)`` per data row. ``cells`` is a dict keyed by
        the header string. Rows where every mapped value is ``None`` are
        skipped.
    """
    header_cells = session.execute(
        select(ExcelCell.col, ExcelCell.raw_value)
        .where(ExcelCell.sheet_id == sheet_id, ExcelCell.row == header_row)
    ).all()

    col_to_header: dict[int, str] = {}
    for col, v in header_cells:
        if v and isinstance(v, str) and v.strip():
            col_to_header[col] = v.strip()

    if headers is not None:
        wanted = {s.strip() for s in headers if s and s.strip()}
        col_to_header = {c: h for c, h in col_to_header.items() if h in wanted}

    if not col_to_header:
        return

    data_cells = session.execute(
        select(ExcelCell.row, ExcelCell.col, ExcelCell.raw_value)
        .where(
            ExcelCell.sheet_id == sheet_id,
            ExcelCell.row > header_row,
            ExcelCell.col.in_(col_to_header.keys()),
        )
        .order_by(ExcelCell.row, ExcelCell.col)
    ).all()

    rows: dict[int, dict[str, Any]] = {}
    for r, c, v in data_cells:
        rows.setdefault(r, {})[col_to_header[c]] = v

    for row_idx in sorted(rows):
        cells = rows[row_idx]
        if any(v is not None for v in cells.values()):
            yield row_idx, cells


# ── Merge-write conflict handling ───────────────────────────────────

MergePolicy = Literal["error", "propagate", "split", "unmerge", "origin_wins"]

MERGE_POLICIES: tuple[MergePolicy, ...] = (
    "error", "propagate", "split", "unmerge", "origin_wins",
)


@dataclass(frozen=True)
class MergeWriteConflict:
    """A merge range where intended cell-writes disagree."""

    sheet_id: int
    merge_id: int
    range: tuple[int, int, int, int]   # (min_row, min_col, max_row, max_col)
    writes: dict[tuple[int, int], Any]   # (row, col) → intended value


def detect_merge_conflicts(
    session: Session,
    sheet_id: int,
    writes: dict[tuple[int, int], Any],
) -> list[MergeWriteConflict]:
    """Group intended writes by merge_id; emit a conflict per group
    whose values disagree."""
    if not writes:
        return []

    # Load merge memberships for every cell in this sheet that has one.
    merged_cells = session.execute(
        select(ExcelCell.row, ExcelCell.col, ExcelCell.merge_id)
        .where(
            ExcelCell.sheet_id == sheet_id,
            ExcelCell.merge_id.isnot(None),
        )
    ).all()

    by_merge: dict[int, dict[tuple[int, int], Any]] = {}
    for r, c, mid in merged_cells:
        if (r, c) in writes:
            by_merge.setdefault(mid, {})[(r, c)] = writes[(r, c)]

    conflicts: list[MergeWriteConflict] = []
    for mid, group in by_merge.items():
        if len(set(group.values())) > 1:
            m = session.get(ExcelMerge, mid)
            conflicts.append(MergeWriteConflict(
                sheet_id=sheet_id,
                merge_id=mid,
                range=(m.min_row, m.min_col, m.max_row, m.max_col),
                writes=group,
            ))
    return conflicts


def resolve_conflicts(
    policy: MergePolicy,
    conflicts: list[MergeWriteConflict],
    writes: dict[tuple[int, int], Any],
    ws: Worksheet,
    session: Session,
) -> dict[tuple[int, int], Any]:
    """Apply the chosen policy to ``conflicts``. May mutate ``writes`` and ``ws``.

    Returns the (possibly modified) writes dict ready for cell-level writes.
    """
    if not conflicts:
        return writes
    if policy not in MERGE_POLICIES:
        raise ValueError(
            f"Unknown merge_policy: {policy!r}. "
            f"Valid: {', '.join(MERGE_POLICIES)}"
        )

    if policy == "error":
        _raise_conflict_error(conflicts)
    elif policy == "propagate":
        return _resolve_propagate(conflicts, writes, session)
    elif policy == "split":
        return _resolve_split(conflicts, writes, ws)
    elif policy == "unmerge":
        return _resolve_unmerge(conflicts, writes, ws)
    elif policy == "origin_wins":
        return _resolve_origin_wins(conflicts, writes)
    return writes


def _raise_conflict_error(conflicts: list[MergeWriteConflict]) -> None:
    lines = []
    for c in conflicts:
        r1, c1, r2, c2 = c.range
        vals = {f"({r},{col})": v for (r, col), v in c.writes.items()}
        lines.append(f"  merge {r1}:{c1}-{r2}:{c2}: {vals}")
    raise ValueError(
        f"Merge write conflicts ({len(conflicts)}):\n"
        + "\n".join(lines)
        + "\n  Pass merge_policy='propagate'/'split'/'unmerge'/'origin_wins' "
        "to choose a resolution."
    )


def _resolve_propagate(
    conflicts: list[MergeWriteConflict],
    writes: dict[tuple[int, int], Any],
    session: Session,
) -> dict[tuple[int, int], Any]:
    """If each conflict has exactly 1 distinct *changed* value, propagate it.

    "Changed" means: intended write differs from ExcelCell.raw_value (the
    imported original). If multiple distinct changes exist within a single
    merge, raise — the user must pick another policy.
    """
    # Load original values for every coord involved in any conflict
    all_coords = {coord for c in conflicts for coord in c.writes}
    if not all_coords:
        return writes

    sheet_id = conflicts[0].sheet_id
    originals: dict[tuple[int, int], Any] = {}
    rows_in_play = {r for r, _ in all_coords}
    cols_in_play = {c for _, c in all_coords}
    cells = session.execute(
        select(ExcelCell.row, ExcelCell.col, ExcelCell.raw_value)
        .where(
            ExcelCell.sheet_id == sheet_id,
            ExcelCell.row.in_(rows_in_play),
            ExcelCell.col.in_(cols_in_play),
        )
    ).all()
    for r, col, raw in cells:
        if (r, col) in all_coords:
            originals[(r, col)] = raw

    for c in conflicts:
        changes = {
            coord: val for coord, val in c.writes.items()
            if val != originals.get(coord)
        }
        distinct = set(changes.values())
        if len(distinct) > 1:
            r1, c1, r2, c2 = c.range
            raise ValueError(
                f"propagate: merge {r1}:{c1}-{r2}:{c2} has multiple distinct "
                f"changed values {distinct}. Use 'split' to keep them separate "
                f"or 'unmerge' to break the merge fully."
            )
        if distinct:
            new_val = next(iter(distinct))
            for coord in c.writes:
                writes[coord] = new_val
    return writes


def _resolve_split(
    conflicts: list[MergeWriteConflict],
    writes: dict[tuple[int, int], Any],
    ws: Worksheet,
) -> dict[tuple[int, int], Any]:
    """Smart split — break each conflicting merge into the minimal number of
    sub-ranges so that same-value contiguous runs stay merged. 2D rectangular
    merges fall back to full unmerge."""
    for c in conflicts:
        r1, c1, r2, c2 = c.range
        is_vertical = (c1 == c2 and r1 < r2)
        is_horizontal = (r1 == r2 and c1 < c2)

        if not (is_vertical or is_horizontal):
            # 2D rectangle — minimal-cover is non-trivial, full unmerge.
            ws.unmerge_cells(
                f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
            )
            continue

        if is_vertical:
            axis = [(r, writes.get((r, c1))) for r in range(r1, r2 + 1)]
        else:
            axis = [(col, writes.get((r1, col))) for col in range(c1, c2 + 1)]

        # Run-length groups along the merge axis
        groups: list[tuple[int, int, Any]] = []
        s, cur, prev = axis[0][0], axis[0][1], axis[0][0]
        for pos, val in axis[1:]:
            if val != cur:
                groups.append((s, prev, cur))
                s, cur = pos, val
            prev = pos
        groups.append((s, prev, cur))

        # Unmerge original range, then re-merge runs of length ≥ 2
        ws.unmerge_cells(
            f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
        )
        for start, end, val in groups:
            if is_vertical:
                if end > start:
                    ws.merge_cells(
                        start_row=start, start_column=c1,
                        end_row=end, end_column=c1,
                    )
                for r in range(start + 1, end + 1):
                    writes.pop((r, c1), None)
                writes[(start, c1)] = val
            else:
                if end > start:
                    ws.merge_cells(
                        start_row=r1, start_column=start,
                        end_row=r1, end_column=end,
                    )
                for col in range(start + 1, end + 1):
                    writes.pop((r1, col), None)
                writes[(r1, start)] = val
    return writes


def _resolve_unmerge(
    conflicts: list[MergeWriteConflict],
    writes: dict[tuple[int, int], Any],
    ws: Worksheet,
) -> dict[tuple[int, int], Any]:
    """Fully unmerge every conflicting range — each cell becomes independent."""
    for c in conflicts:
        r1, c1, r2, c2 = c.range
        ws.unmerge_cells(
            f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
        )
    return writes


def _resolve_origin_wins(
    conflicts: list[MergeWriteConflict],
    writes: dict[tuple[int, int], Any],
) -> dict[tuple[int, int], Any]:
    """Keep only the origin row's intended value; drop member writes."""
    for c in conflicts:
        r1, c1, _, _ = c.range
        origin = (r1, c1)
        origin_val = c.writes.get(origin)
        for coord in c.writes:
            if coord != origin:
                writes.pop(coord, None)
            else:
                writes[coord] = origin_val
    return writes
