"""Domain round-trip export — overwrite domain cells on the original BLOB.

Unlike :func:`xp_excel_toolkit.export.writer.export_from_cells` (which rebuilds
a sheet from scratch out of ``excel_cell`` records), this path loads the
original workbook BLOB stored in ``excel_workbook`` and overwrites only the
cells belonging to your domain rows. Original formatting, non-domain sheets,
merges, styles, charts, and formulas are preserved — the file still looks like
the source. Use it when editing a structured spec (register-map / pinmap / …)
via ORM and re-emitting an xlsx that mirrors the original.
"""
from __future__ import annotations

import fnmatch
import io
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from xp_excel_toolkit.models import ExcelSheet, ExcelWorkbook


def write_cell(ws: Worksheet, row: int, col: int, val) -> None:
    """Write a value to a cell, redirecting writes inside merge ranges to the
    origin cell so openpyxl's MergedCell guard doesn't raise."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if (mr.min_row <= row <= mr.max_row
                    and mr.min_col <= col <= mr.max_col):
                ws.cell(row=mr.min_row, column=mr.min_col).value = val
                break
    else:
        cell.value = val


def build_column_map(
    ws: Worksheet,
    header_row: int,
    field_map: dict[str, str],
) -> dict[str, int]:
    """Build {domain_field_name: column_index} from a header row.

    ``field_map`` is the same {excel_header: domain_field} dict you'd use for a
    flat-table sheet config.
    """
    col_map: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value and isinstance(cell.value, str):
            header = cell.value.strip()
            if header in field_map:
                col_map[field_map[header]] = cell.column
    return col_map


# Signature: ``exporter_func(session, ws, sheet_obj) -> int`` (rows written).
# Use this when your domain layout is too irregular for the flat ``field_map``
# path — e.g. key-value sections written into specific column-A markers.
ExporterFunc = Callable[[Session, Worksheet, ExcelSheet], int]


@dataclass
class ExportHandler:
    """Per-sheet-pattern export configuration.

    Provide ``field_map`` + ``domain_cls`` for flat-table sheets, or
    ``exporter_func`` for custom layouts.
    """

    pattern: str
    """Sheet-name fnmatch pattern (case-insensitive)."""

    field_map: dict[str, str] | None = None
    """{excel_header: domain_field}. Required when using domain_cls."""

    domain_cls: type | None = None
    """ORM class to query (filter_by sheet_id). Each row's ``excel_row`` must
    be set so the writer knows which row to overwrite."""

    exporter_func: ExporterFunc | None = None
    """Custom writer. Takes precedence over field_map/domain_cls when set."""


def _export_flat_handler(
    session: Session,
    ws: Worksheet,
    sheet_obj: ExcelSheet,
    domain_cls: type,
    field_map: dict[str, str],
) -> int:
    """Default flat-table export: ORM rows → cells via field_map."""
    if not sheet_obj.header_row:
        return 0

    col_map = build_column_map(ws, sheet_obj.header_row, field_map)
    if not col_map:
        return 0

    rows = session.query(domain_cls).filter_by(sheet_id=sheet_obj.id).all()
    for row_obj in rows:
        excel_row = getattr(row_obj, "excel_row", None)
        if excel_row is None:
            continue
        for field_name, col_idx in col_map.items():
            val = getattr(row_obj, field_name, None)
            write_cell(ws, excel_row, col_idx, val)
    return len(rows)


def export_domain_xlsx(
    session: Session,
    output_path: str | Path,
    handlers: list[ExportHandler],
    *,
    workbook_id: int | None = None,
    on_progress: Callable[[str], None] | None = None,
) -> Path:
    """Round-trip export: load the original BLOB and overwrite domain cells.

    All formatting, merges, styles, and non-domain sheets are preserved.

    Args:
        session: SQLAlchemy session.
        output_path: Destination xlsx path.
        handlers: List of :class:`ExportHandler` — first matching pattern per
            sheet wins.
        workbook_id: Explicit workbook to export. Defaults to the first
            workbook in the DB.
        on_progress: Optional progress callback.

    Raises:
        ValueError: If no workbook is found or its blob is missing.
    """
    output_path = Path(output_path)

    q = session.query(ExcelWorkbook)
    if workbook_id is not None:
        q = q.filter(ExcelWorkbook.id == workbook_id)
    wb_obj = q.first()
    if not wb_obj or not wb_obj.blob:
        raise ValueError("No workbook found in DB or BLOB is missing")

    wb = openpyxl.load_workbook(io.BytesIO(wb_obj.blob))

    total = 0
    for sheet_obj in wb_obj.sheets:
        if sheet_obj.name not in wb.sheetnames:
            continue
        ws = wb[sheet_obj.name]

        for handler in handlers:
            if not fnmatch.fnmatch(sheet_obj.name.lower(), handler.pattern.lower()):
                continue

            if handler.exporter_func is not None:
                count = handler.exporter_func(session, ws, sheet_obj)
            elif handler.domain_cls is not None and handler.field_map is not None:
                count = _export_flat_handler(
                    session, ws, sheet_obj, handler.domain_cls, handler.field_map,
                )
            else:
                count = 0

            if count and on_progress:
                on_progress(f"  {sheet_obj.name}: {count} rows")
            total += count
            break  # first match wins

    wb.save(output_path)
    wb.close()

    if on_progress:
        on_progress(f"Exported {total} domain rows to {output_path}")

    return output_path
