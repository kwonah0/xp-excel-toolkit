"""구조 편집 — 행/열 insert·delete (셀 shift + merge 경계 조정 + 서식 클론).

openpyxl 의 ``ws.insert_rows``/``insert_cols`` 는 셀 값·스타일은 옮기지만 **병합 범위를
자동 조정하지 않는다**(알려진 함정). 이 모듈이 unmerge→shift→re-merge 로 병합을 정확히
보정하고(범위를 가로지르면 확장, 우측이면 이동), 새 행/열에 인접 라인의 서식을 클론한다.

export 핸들러가 받는 살아있는 ``ws`` 위에서 호출한다(blob round-trip 과 동일 통로).
"""
from __future__ import annotations

import copy

from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet


def _capture_unmerge(ws: Worksheet) -> list[CellRange]:
    """현재 병합 범위를 떠서 모두 해제 — openpyxl insert/delete *전에* 호출해야 한다
    (insert 후엔 셀이 밀려 unmerge 가 깨진다)."""
    ranges = [CellRange(str(m)) for m in list(ws.merged_cells.ranges)]
    for r in ranges:
        ws.unmerge_cells(str(r))
    return ranges


def _shift(b: int, idx: int, amount: int, *, delete: bool, is_min: bool) -> int:
    """경계 좌표 하나를 엑셀과 동일하게 조정.

    insert: idx 이상이면 +amount(가로지르면 max만 밀려 확장).
    delete: 삭제 범위 [idx, idx+amount) 안 경계는 **clamp**(min→idx, max→idx-1),
            그 우측은 -amount, 좌측은 그대로 — 엑셀 동작과 일치.
    """
    if not delete:
        return b + amount if b >= idx else b
    if b < idx:
        return b
    if b < idx + amount:
        return idx if is_min else idx - 1
    return b - amount


def _remerge(ws: Worksheet, ranges: list[CellRange], *, row_idx: int | None = None,
             col_idx: int | None = None, amount: int, delete: bool) -> None:
    """op *후에* 호출 — 경계를 엑셀과 동일하게 조정해 다시 병합."""
    for r in ranges:
        mr, xr, mc, xc = r.min_row, r.max_row, r.min_col, r.max_col
        if row_idx is not None:
            mr = _shift(mr, row_idx, amount, delete=delete, is_min=True)
            xr = _shift(xr, row_idx, amount, delete=delete, is_min=False)
        if col_idx is not None:
            mc = _shift(mc, col_idx, amount, delete=delete, is_min=True)
            xc = _shift(xc, col_idx, amount, delete=delete, is_min=False)
        if mr > xr or mc > xc or (mr == xr and mc == xc):
            continue                       # 삭제로 사라졌거나 1×1
        ws.merge_cells(start_row=mr, start_column=mc, end_row=xr, end_column=xc)


def _row_styles(ws: Worksheet, row: int) -> list:
    return [copy.copy(ws.cell(row, c)._style) for c in range(1, ws.max_column + 1)]


def _col_styles(ws: Worksheet, col: int) -> list:
    return [copy.copy(ws.cell(r, col)._style) for r in range(1, ws.max_row + 1)]


def insert_rows(ws: Worksheet, idx: int, amount: int = 1, *,
                copy_style_from: int | None = None) -> None:
    """idx 앞에 amount 개 행 삽입. 병합 보정 + (copy_style_from 행의)서식 클론."""
    styles = _row_styles(ws, copy_style_from) if copy_style_from else None
    ranges = _capture_unmerge(ws)
    ws.insert_rows(idx, amount)
    _remerge(ws, ranges, row_idx=idx, amount=amount, delete=False)
    if styles:
        for i in range(amount):
            for c, st in enumerate(styles, start=1):
                ws.cell(idx + i, c)._style = copy.copy(st)


def insert_cols(ws: Worksheet, idx: int, amount: int = 1, *,
                copy_style_from: int | None = None) -> None:
    """idx 앞에 amount 개 열 삽입. 병합 보정(헤더 그룹 확장 포함) + 서식 클론."""
    styles = _col_styles(ws, copy_style_from) if copy_style_from else None
    ranges = _capture_unmerge(ws)
    ws.insert_cols(idx, amount)
    _remerge(ws, ranges, col_idx=idx, amount=amount, delete=False)
    if styles:
        for i in range(amount):
            for r, st in enumerate(styles, start=1):
                ws.cell(r, idx + i)._style = copy.copy(st)


def delete_rows(ws: Worksheet, idx: int, amount: int = 1) -> None:
    """idx 부터 amount 개 행 삭제(위로 당김) + 병합 보정."""
    ranges = _capture_unmerge(ws)
    ws.delete_rows(idx, amount)
    _remerge(ws, ranges, row_idx=idx, amount=amount, delete=True)


def delete_cols(ws: Worksheet, idx: int, amount: int = 1) -> None:
    """idx 부터 amount 개 열 삭제(왼쪽으로 당김) + 병합 보정."""
    ranges = _capture_unmerge(ws)
    ws.delete_cols(idx, amount)
    _remerge(ws, ranges, col_idx=idx, amount=amount, delete=True)
