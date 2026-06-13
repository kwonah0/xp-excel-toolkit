"""openpyxl-based parser for .xlsx files.

Cell-only: populates ``ExcelWorkbook`` / ``ExcelSheet`` / ``ExcelMerge``
/ ``ExcelCell`` and detects the header row. Domain interpretation lives
in caller-provided builders that read the resulting tables.
"""

from __future__ import annotations

from collections.abc import Callable, Iterable
from pathlib import Path
from typing import Any, Literal

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import bindparam, insert, update
from sqlalchemy.orm import Session

from xp_excel_toolkit.ingest.convert import validate_xlsx_format
from xp_excel_toolkit.merge import MergeResolver
from xp_excel_toolkit.models import ExcelCell, ExcelMerge, ExcelSheet, ExcelWorkbook

# Max dicts per Core bulk insert execute() call (avoids SQLite variable limit)
BULK_CHUNK = 500


def extract_cell_value(value: Any) -> tuple[str | None, str | None, str | None]:
    """Extract (raw_value, formula_type, formula_ref) from an openpyxl cell value."""
    if value is None:
        return None, None, None
    if isinstance(value, ArrayFormula):
        text = value.text if value.text else None
        return text, "array", value.ref
    if isinstance(value, DataTableFormula):
        return None, "dataTable", value.ref
    return str(value), None, None


def extract_style(cell: Cell) -> dict[str, Any] | None:
    """Extract visual style info from an openpyxl cell."""
    style: dict[str, Any] = {}

    # Background color
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
        rgb = fill.fgColor.rgb
        if isinstance(rgb, str) and len(rgb) == 8:
            style["bg_color"] = f"#{rgb[2:]}"  # strip alpha

    # Font
    font = cell.font
    if font:
        if font.bold:
            style["font_bold"] = True
        if font.strike:
            style["font_strike"] = True
        if font.color and font.color.rgb and font.color.rgb != "00000000":
            rgb = font.color.rgb
            if isinstance(rgb, str) and len(rgb) == 8:
                style["font_color"] = f"#{rgb[2:]}"

    # Number format
    if cell.number_format and cell.number_format != "General":
        style["number_format"] = cell.number_format

    # Borders
    border = cell.border
    if border:
        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(border, side_name)
            if side and side.style:
                style[f"border_{side_name}"] = side.style

    return style if style else None


def find_header_row(
    ws: Worksheet,
    expected_headers: Iterable[str],
    *,
    match: Literal["any", "all"] = "any",
    max_scan: int = 30,
) -> int | None:
    """Auto-detect the header row in an openpyxl worksheet.

    Args:
        expected_headers: Header strings the row is expected to contain.
            Any iterable of strings (list, tuple, set, ``dict.keys()``).
            At least one non-empty string is required.
        match: ``"any"`` (default) returns the first row containing at
            least one of ``expected_headers``; ``"all"`` requires every
            expected header to be present.
        max_scan: Stop scanning after this row number.

    Raises:
        ValueError: if ``expected_headers`` is empty or contains only
            blank strings.
    """
    expected = {s.strip() for s in expected_headers if s and s.strip()}
    if not expected:
        raise ValueError(
            "expected_headers must contain at least one non-empty string"
        )

    max_row = min(max_scan, ws.max_row or max_scan)
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        values = {
            str(cell.value).strip()
            for cell in row
            if cell.value is not None and isinstance(cell.value, str) and cell.value.strip()
        }
        if match == "all":
            if expected.issubset(values):
                return row[0].row
        else:  # "any"
            if values & expected:
                return row[0].row
    return None


def _import_ws(
    session: Session,
    ws: Worksheet,
    wb_id: int,
    *,
    header_row: int | None = None,
) -> ExcelSheet:
    """Import one openpyxl worksheet into the DB (cells + merges).

    ``header_row`` is only stored if the caller passes it; xp_excel_toolkit itself
    does not try to guess. Domain builders detect the header via
    :func:`xp_excel_toolkit.query.find_header_row_db` with their own header
    keyword list and update ``ExcelSheet.header_row`` themselves.
    """
    sheet_obj = ExcelSheet(workbook_id=wb_id, name=ws.title)
    session.add(sheet_obj)
    session.flush()

    sid = sheet_obj.id

    # -- Merge ranges (Core bulk insert) ------------------------------------
    merger = MergeResolver.from_worksheet(ws)

    merge_dicts = [
        {
            "sheet_id": sid,
            "min_row": mr.min_row,
            "min_col": mr.min_col,
            "max_row": mr.max_row,
            "max_col": mr.max_col,
        }
        for mr in merger.ranges
    ]

    merge_id_map: dict[str, int] = {}  # "min_row:min_col" -> merge.id
    if merge_dicts:
        for i in range(0, len(merge_dicts), BULK_CHUNK):
            session.execute(insert(ExcelMerge), merge_dicts[i:i + BULK_CHUNK])
        session.flush()

        for mid, mrow, mcol in (
            session.query(ExcelMerge.id, ExcelMerge.min_row, ExcelMerge.min_col)
            .filter(ExcelMerge.sheet_id == sid)
            .all()
        ):
            merge_id_map[f"{mrow}:{mcol}"] = mid

    # -- Cells (Core bulk insert) -------------------------------------------
    cell_dicts: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            r, c = cell.row, cell.column

            if merger.is_merged(r, c):
                raw_val = merger.get_value(r, c)
                f_type, f_ref = None, None
            else:
                raw_val, f_type, f_ref = extract_cell_value(cell.value)

            merge_key = None
            is_origin = merger.is_origin(r, c)
            origin = merger.get_origin(r, c)
            if origin:
                merge_key = merge_id_map.get(f"{origin[0]}:{origin[1]}")

            style = extract_style(cell) if cell.value is not None or is_origin else None
            comment_text = cell.comment.text if cell.comment else None

            cell_dicts.append({
                "sheet_id": sid,
                "row": r,
                "col": c,
                "raw_value": raw_val,
                "style": style,
                "comment": comment_text,
                "formula_type": f_type,
                "formula_ref": f_ref,
                "merge_id": merge_key,
                "is_merge_origin": is_origin,
            })

    if cell_dicts:
        for i in range(0, len(cell_dicts), BULK_CHUNK):
            session.execute(insert(ExcelCell), cell_dicts[i:i + BULK_CHUNK])
        session.flush()

    # Header row is left to the domain builder (it knows the expected names).
    if header_row is not None:
        sheet_obj.header_row = header_row

    session.flush()
    return sheet_obj


def import_sheet(
    session: Session,
    path: str | Path,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> ExcelSheet:
    """Import a single sheet from a .xlsx file (cells + merges)."""
    path = Path(path)
    blob = path.read_bytes()

    wb_xl = openpyxl.load_workbook(path, data_only=False)
    ws = wb_xl[sheet_name] if sheet_name else wb_xl.active

    wb_obj = ExcelWorkbook(filename=path.name, blob=blob)
    session.add(wb_obj)
    session.flush()

    sheet_obj = _import_ws(session, ws, wb_obj.id, header_row=header_row)

    wb_xl.close()
    return sheet_obj


def import_xlsx(
    session: Session,
    path: str | Path,
    *,
    on_progress: Callable[[str], None] | None = None,
    with_formulas: bool = False,
) -> list[ExcelSheet]:
    """Import all sheets from a .xlsx file. Cells/merges only.

    Args:
        session: SQLAlchemy session.
        path: Path to .xlsx file.
        on_progress: Optional progress callback (sheet-level messages).
        with_formulas: If True, load formulas first then overlay cached
            values (loads the workbook twice).

    Returns:
        List of created ExcelSheet rows in source order.
    """
    path = Path(path)

    if path.suffix.lower() == ".xlsx":
        validate_xlsx_format(path)

    blob = path.read_bytes()

    if with_formulas:
        if on_progress:
            on_progress(f"Loading workbook (formulas): {path.name}")
        wb_xl = openpyxl.load_workbook(path, data_only=False)
    else:
        if on_progress:
            on_progress(f"Loading workbook: {path.name}")
        wb_xl = openpyxl.load_workbook(path, data_only=True)

    wb_obj = ExcelWorkbook(filename=path.name, blob=blob)
    session.add(wb_obj)
    session.flush()

    total = len(wb_xl.sheetnames)
    sheets: list[ExcelSheet] = []
    for idx, name in enumerate(wb_xl.sheetnames, 1):
        if on_progress:
            on_progress(f"  [{idx}/{total}] Importing sheet: {name}")
        ws = wb_xl[name]
        sheet_obj = _import_ws(session, ws, wb_obj.id)
        sheets.append(sheet_obj)

    wb_xl.close()

    if with_formulas:
        if on_progress:
            on_progress("Loading cached formula values (data_only=True)...")
        wb_val = openpyxl.load_workbook(path, data_only=True)
        overlay = (
            update(ExcelCell)
            .where(
                ExcelCell.sheet_id == bindparam("b_sheet_id"),
                ExcelCell.row == bindparam("b_row"),
                ExcelCell.col == bindparam("b_col"),
                ExcelCell.raw_value.like("=%"),
            )
            .values(cached_value=bindparam("b_val"))
        )
        for sheet_obj in sheets:
            ws_val = wb_val[sheet_obj.name]
            params: list[dict[str, Any]] = []
            for row in ws_val.iter_rows(
                min_row=1,
                max_row=ws_val.max_row,
                max_col=ws_val.max_column,
            ):
                for cell in row:
                    if cell.value is not None:
                        params.append({
                            "b_sheet_id": sheet_obj.id,
                            "b_row": cell.row,
                            "b_col": cell.column,
                            "b_val": str(cell.value),
                        })
            if params:
                conn = session.connection()
                for i in range(0, len(params), BULK_CHUNK):
                    conn.execute(overlay, params[i:i + BULK_CHUNK])
                session.flush()
        wb_val.close()
        if on_progress:
            on_progress("Cached formula values saved.")

    return sheets
