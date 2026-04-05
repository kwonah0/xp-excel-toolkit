"""Split a register map by sheet, with each IP as a sub-sheet."""

from __future__ import annotations

from collections.abc import Callable
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell
from sqlalchemy import distinct, insert
from sqlalchemy.orm import Session

from dsm.merge import MergeResolver
from dsm.models import ExcelCell, ExcelMerge, ExcelSheet, ExcelWorkbook, SheetConfigEntry
from dsm.domain_models import REGMAP_FIELD_MAP, Register
from dsm.xlsx_parser import _import_ws, _BULK_CHUNK


def _find_register_sheets(
    session: Session,
    workbook_id: int,
) -> list[ExcelSheet]:
    """Find sheets that match register-type configs from DB.

    Reads SheetConfigEntry rows where domain_type='register' and uses their
    patterns to filter ExcelSheet records via fnmatch.
    Falls back to 'level2_%' LIKE filter if no configs in DB.
    """
    import fnmatch as _fnmatch

    configs = (
        session.query(SheetConfigEntry)
        .filter(SheetConfigEntry.domain_type == "register")
        .all()
    )

    all_sheets = (
        session.query(ExcelSheet)
        .filter(ExcelSheet.workbook_id == workbook_id)
        .all()
    )

    if not configs:
        # Fallback: legacy hardcoded pattern
        return [s for s in all_sheets if s.name.startswith("level2_")]

    patterns = [c.pattern for c in configs]
    matched: list[ExcelSheet] = []
    for sheet in all_sheets:
        for pat in patterns:
            if _fnmatch.fnmatch(sheet.name, pat):
                matched.append(sheet)
                break
    return matched


def _build_ip_sheet(
    session: Session,
    src_sheet: ExcelSheet,
    merger: MergeResolver,
    ip_name: str,
    registers: list[Register],
    dst_wb_id: int,
) -> ExcelSheet:
    """Create a new ExcelSheet in DB for one IP, with remapped rows."""
    src_rows = sorted({reg.excel_row for reg in registers})
    row_map = {src_r: dst_i for dst_i, src_r in enumerate(src_rows, start=2)}

    ip_sheet = ExcelSheet(
        workbook_id=dst_wb_id,
        name=ip_name,
        header_row=1,
    )
    session.add(ip_sheet)
    session.flush()

    sid = ip_sheet.id

    # Copy header row cells (bulk insert)
    header_cells = (
        session.query(ExcelCell)
        .filter(ExcelCell.sheet_id == src_sheet.id, ExcelCell.row == src_sheet.header_row)
        .all()
    )
    if header_cells:
        hdr_dicts = [
            {
                "sheet_id": sid, "row": 1, "col": cell.col,
                "raw_value": cell.raw_value, "style": cell.style,
                "comment": cell.comment,
                "merge_id": None, "is_merge_origin": False,
            }
            for cell in header_cells
        ]
        session.execute(insert(ExcelCell), hdr_dicts)

    # Copy merges with row remapping (bulk insert + query back)
    src_merges = (
        session.query(ExcelMerge)
        .filter(ExcelMerge.sheet_id == src_sheet.id)
        .all()
    )

    # Build merge dicts and track which src merge.id they came from
    merge_dicts: list[dict] = []
    src_merge_ids: list[int] = []  # parallel list: src merge.id per dict
    for merge in src_merges:
        covered = [r for r in src_rows if merge.min_row <= r <= merge.max_row]

        if merge.min_row == merge.max_row and merge.min_row in row_map:
            merge_dicts.append({
                "sheet_id": sid,
                "min_row": row_map[merge.min_row],
                "min_col": merge.min_col,
                "max_row": row_map[merge.min_row],
                "max_col": merge.max_col,
            })
            src_merge_ids.append(merge.id)
        elif len(covered) > 1:
            merge_dicts.append({
                "sheet_id": sid,
                "min_row": row_map[covered[0]],
                "min_col": merge.min_col,
                "max_row": row_map[covered[-1]],
                "max_col": merge.max_col,
            })
            src_merge_ids.append(merge.id)

    merge_id_map: dict[int, int] = {}  # src_merge.id -> new_merge.id
    if merge_dicts:
        for i in range(0, len(merge_dicts), _BULK_CHUNK):
            session.execute(insert(ExcelMerge), merge_dicts[i:i + _BULK_CHUNK])
        session.flush()

        # Query back new merge IDs, match by (min_row, min_col)
        new_merges = (
            session.query(ExcelMerge.id, ExcelMerge.min_row, ExcelMerge.min_col)
            .filter(ExcelMerge.sheet_id == sid)
            .all()
        )
        # Build reverse lookup: (min_row, min_col) -> new_id
        new_merge_lookup = {(mrow, mcol): mid for mid, mrow, mcol in new_merges}
        for md, src_id in zip(merge_dicts, src_merge_ids):
            new_id = new_merge_lookup.get((md["min_row"], md["min_col"]))
            if new_id:
                merge_id_map[src_id] = new_id

    # Copy data cells (bulk insert)
    data_cells = (
        session.query(ExcelCell)
        .filter(
            ExcelCell.sheet_id == src_sheet.id,
            ExcelCell.row.in_(src_rows),
        )
        .all()
    )
    cell_dicts: list[dict] = []
    for cell in data_cells:
        raw_value = cell.raw_value
        if raw_value is None and merger.is_merged(cell.row, cell.col):
            raw_value = merger.get_value(cell.row, cell.col)

        new_merge_id = merge_id_map.get(cell.merge_id) if cell.merge_id else None
        is_origin = merger.is_origin(cell.row, cell.col) if cell.merge_id else False

        cell_dicts.append({
            "sheet_id": sid,
            "row": row_map[cell.row],
            "col": cell.col,
            "raw_value": raw_value,
            "style": cell.style,
            "comment": cell.comment,
            "merge_id": new_merge_id,
            "is_merge_origin": is_origin,
        })

    if cell_dicts:
        for i in range(0, len(cell_dicts), _BULK_CHUNK):
            session.execute(insert(ExcelCell), cell_dicts[i:i + _BULK_CHUNK])
        session.flush()

    return ip_sheet


def split_regmap(
    session: Session,
    path: str | Path,
    output_dir: str | Path,
) -> dict[str, Path]:
    """Split a register map xlsx: one output file per source sheet,
    with each IP as a separate sheet inside.

    e.g. regmap_sample.xlsx (sheets: level2_common, level2_buscon)
         → level2_common.xlsx  (sheets: SENSOR_A, AMPLIFIER, ...)
         → level2_buscon.xlsx  (sheets: GPIO_PORT, TIMER_A, ...)

    Args:
        session: SQLAlchemy session.
        path: Path to the source regmap xlsx.
        output_dir: Directory to write output files.

    Returns:
        Dict mapping source sheet name to output file path.
    """
    path = Path(path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load workbook once for all sheets
    blob = path.read_bytes()
    wb_xl = openpyxl.load_workbook(path, data_only=False)

    wb_obj = ExcelWorkbook(filename=path.name, blob=blob)
    session.add(wb_obj)
    session.flush()

    results: dict[str, Path] = {}

    # Only process level2_* sheets (register bit-field specs)
    level2_sheets = [sn for sn in wb_xl.sheetnames if sn.startswith("level2_")]

    for sn in level2_sheets:
        ws = wb_xl[sn]
        src_sheet = _import_ws(
            session, ws, wb_obj.id,
            field_map=REGMAP_FIELD_MAP,
            domain_cls=Register,
        )
        merger = MergeResolver.from_db(session, src_sheet.id)

        ip_names = (
            session.query(distinct(Register.name))
            .filter(Register.sheet_id == src_sheet.id, Register.name.isnot(None))
            .all()
        )
        if not ip_names:
            continue

        out_wb = ExcelWorkbook(filename=f"{sn}.xlsx")
        session.add(out_wb)
        session.flush()

        ip_sheets: list[int] = []

        for (ip_name,) in ip_names:
            ip_name = ip_name.strip()
            if not ip_name:
                continue

            registers = (
                session.query(Register)
                .filter(Register.sheet_id == src_sheet.id, Register.name == ip_name)
                .order_by(Register.excel_row)
                .all()
            )
            if not registers:
                continue

            ip_sheet = _build_ip_sheet(
                session, src_sheet, merger, ip_name, registers, out_wb.id,
            )
            ip_sheets.append(ip_sheet.id)

        out_path = output_dir / f"{sn}.xlsx"
        _export_multi_sheet(session, ip_sheets, out_path)
        results[sn] = out_path

    wb_xl.close()
    return results


def _export_multi_sheet(
    session: Session,
    sheet_ids: list[int],
    output_path: Path,
) -> None:
    """Export multiple ExcelSheet records into one xlsx file, each as a sheet."""
    from openpyxl.comments import Comment
    from dsm.exporter import _apply_style

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for sheet_id in sheet_ids:
        sheet_obj = session.get(ExcelSheet, sheet_id)
        ws = wb.create_sheet(title=sheet_obj.name)

        # Write cells
        cells = (
            session.query(ExcelCell)
            .filter(ExcelCell.sheet_id == sheet_id)
            .all()
        )
        for cell_rec in cells:
            c = ws.cell(row=cell_rec.row, column=cell_rec.col, value=cell_rec.raw_value)
            _apply_style(c, cell_rec.style)
            if cell_rec.comment:
                c.comment = Comment(cell_rec.comment, "")

        # Apply merges
        merges = (
            session.query(ExcelMerge)
            .filter(ExcelMerge.sheet_id == sheet_id)
            .all()
        )
        for m in merges:
            ws.merge_cells(
                start_row=m.min_row, start_column=m.min_col,
                end_row=m.max_row, end_column=m.max_col,
            )

        if sheet_obj.header_row:
            ws.freeze_panes = f"A{sheet_obj.header_row + 1}"

    wb.save(output_path)
    wb.close()


def split_regmap_from_db(
    session: Session,
    workbook_filename: str,
    output_dir: str | Path,
    on_progress: Callable[[str], None] | None = None,
) -> dict[str, Path]:
    """Split regmap using already-imported DB data (no xlsx reload needed).

    Args:
        session: SQLAlchemy session with previously imported data.
        workbook_filename: Filename to match in ExcelWorkbook table.
        output_dir: Directory to write output files.

    Returns:
        Dict mapping source sheet name to output file path.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    wb_obj = (
        session.query(ExcelWorkbook)
        .filter_by(filename=workbook_filename)
        .first()
    )
    if not wb_obj:
        raise ValueError(
            f"Workbook '{workbook_filename}' not found in DB. Run 'dsm import' first."
        )

    level2_sheets = _find_register_sheets(session, wb_obj.id)

    results: dict[str, Path] = {}
    total = len(level2_sheets)

    for si, src_sheet in enumerate(level2_sheets, 1):
        if on_progress:
            on_progress(f"  [{si}/{total}] Processing sheet: {src_sheet.name}")
        merger = MergeResolver.from_db(session, src_sheet.id)

        ip_names = (
            session.query(distinct(Register.name))
            .filter(Register.sheet_id == src_sheet.id, Register.name.isnot(None))
            .all()
        )
        if not ip_names:
            continue

        out_wb = ExcelWorkbook(filename=f"{src_sheet.name}.xlsx")
        session.add(out_wb)
        session.flush()

        ip_sheets: list[int] = []

        for (ip_name,) in ip_names:
            ip_name = ip_name.strip()
            if not ip_name:
                continue

            registers = (
                session.query(Register)
                .filter(Register.sheet_id == src_sheet.id, Register.name == ip_name)
                .order_by(Register.excel_row)
                .all()
            )
            if not registers:
                continue

            ip_sheet = _build_ip_sheet(
                session, src_sheet, merger, ip_name, registers, out_wb.id,
            )
            ip_sheets.append(ip_sheet.id)

        out_path = output_dir / f"{src_sheet.name}.xlsx"
        if on_progress:
            on_progress(f"  [{si}/{total}] Writing {src_sheet.name}.xlsx ({len(ip_sheets)} IPs)")
        _export_multi_sheet(session, ip_sheets, out_path)
        results[src_sheet.name] = out_path

    return results


# -- merge split files back ---------------------------------------------------


def _copy_cell_style(src: Cell, dst: Cell) -> None:
    """Copy openpyxl cell style attributes from src to dst."""
    from copy import copy

    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)


def merge_split_files(
    input_dir: Path,
    output_path: Path,
    on_progress: Callable[[str], None] | None = None,
) -> dict[str, list[str]]:
    """Merge split xlsx files back into a single xlsx.

    Each split file's stem becomes the sheet name (e.g. level2_common.xlsx → level2_common).
    IP tabs within each file are stacked vertically into one sheet.

    Args:
        input_dir: Directory containing split xlsx files.
        output_path: Path for the merged output xlsx.

    Returns:
        Dict mapping sheet name to list of IP names merged.
    """
    from copy import copy
    from openpyxl.comments import Comment

    input_dir = Path(input_dir)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    split_files = sorted(input_dir.glob("*.xlsx"))
    if not split_files:
        raise ValueError(f"No .xlsx files found in {input_dir}")

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    result: dict[str, list[str]] = {}
    total = len(split_files)

    for fi, split_file in enumerate(split_files, 1):
        source_sheet_name = split_file.stem
        if on_progress:
            on_progress(f"  [{fi}/{total}] Merging: {split_file.name}")
        ws_out = wb_out.create_sheet(title=source_sheet_name)

        split_wb = openpyxl.load_workbook(split_file)
        ip_names: list[str] = []

        # Copy header from first IP sheet (row 1)
        first_sheet = split_wb.worksheets[0]
        max_col = first_sheet.max_column or 1
        for col in range(1, max_col + 1):
            src = first_sheet.cell(row=1, column=col)
            dst = ws_out.cell(row=1, column=col, value=src.value)
            _copy_cell_style(src, dst)

        current_row = 2
        for ip_sheet in split_wb.worksheets:
            ip_names.append(ip_sheet.title)
            ip_start_row = current_row  # track where this IP starts
            max_r = ip_sheet.max_row or 1
            max_c = ip_sheet.max_column or 1

            # Copy data rows (skip header row 1)
            for r in range(2, max_r + 1):
                for c in range(1, max_c + 1):
                    src = ip_sheet.cell(row=r, column=c)
                    dst = ws_out.cell(row=current_row, column=c, value=src.value)
                    _copy_cell_style(src, dst)
                    if src.comment:
                        dst.comment = Comment(src.comment.text, src.comment.author)
                current_row += 1

            # Copy merge ranges with row offset
            row_offset = ip_start_row - 2  # ip row 2 → ws_out ip_start_row
            for merge_range in ip_sheet.merged_cells.ranges:
                # Skip header merges (row 1)
                if merge_range.min_row < 2:
                    continue
                ws_out.merge_cells(
                    start_row=merge_range.min_row + row_offset,
                    start_column=merge_range.min_col,
                    end_row=merge_range.max_row + row_offset,
                    end_column=merge_range.max_col,
                )

        ws_out.freeze_panes = "A2"
        split_wb.close()
        result[source_sheet_name] = ip_names

    wb_out.save(output_path)
    wb_out.close()
    return result
