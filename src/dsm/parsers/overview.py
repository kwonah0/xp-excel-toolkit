"""Parser for Overview-style sheets (dict-like key-value with categories).

Layout:
    Col A (1): key or ``#CategoryName``
    Col B (2): value
    Col C (3): comment

Rules:
    - ``#Text`` in col A **without** value in col B → category header row.
    - ``#Text`` in col A **with** value in col B → commented-out entry.
    - Normal text in col A → regular key-value entry.
    - The current category applies to all subsequent rows until the next
      category marker.
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import insert
from sqlalchemy.orm import Session

from dsm.domain_models import OverviewEntry
from dsm.xlsx_parser import _BULK_CHUNK, _extract_cell_value


def parse_overview_entries(
    session: Session,
    ws: Worksheet,
    sheet_id: int,
) -> None:
    """Parse an Overview sheet and insert OverviewEntry records.

    Args:
        session: SQLAlchemy session.
        ws: openpyxl Worksheet object.
        sheet_id: ID of the ExcelSheet record (already created by _import_ws).
    """
    current_category: str | None = None
    bulk_rows: list[dict] = []

    for row_idx in range(1, (ws.max_row or 0) + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        cell_b = ws.cell(row=row_idx, column=2)
        cell_c = ws.cell(row=row_idx, column=3)

        raw_a, _, _ = _extract_cell_value(cell_a.value)
        raw_b, _, _ = _extract_cell_value(cell_b.value)
        raw_c, _, _ = _extract_cell_value(cell_c.value)

        # Skip completely empty rows
        if raw_a is None and raw_b is None and raw_c is None:
            continue

        key_str = raw_a.strip() if raw_a else None

        # Detect # prefix
        if key_str and key_str.startswith("#"):
            stripped_key = key_str[1:].strip()

            if raw_b is None or (raw_b and raw_b.strip() == ""):
                # Category header: #CategoryName without value
                current_category = stripped_key
                bulk_rows.append({
                    "sheet_id": sheet_id,
                    "excel_row": row_idx,
                    "category": current_category,
                    "key": None,
                    "value": None,
                    "comment": raw_c.strip() if raw_c else None,
                    "is_commented": False,
                    "is_category": True,
                })
                continue
            else:
                # Commented-out entry: #key with value
                bulk_rows.append({
                    "sheet_id": sheet_id,
                    "excel_row": row_idx,
                    "category": current_category,
                    "key": stripped_key,
                    "value": raw_b.strip() if raw_b else None,
                    "comment": raw_c.strip() if raw_c else None,
                    "is_commented": True,
                    "is_category": False,
                })
                continue

        # Regular key-value entry (or comment-only row)
        bulk_rows.append({
            "sheet_id": sheet_id,
            "excel_row": row_idx,
            "category": current_category,
            "key": key_str,
            "value": raw_b.strip() if raw_b else None,
            "comment": raw_c.strip() if raw_c else None,
            "is_commented": False,
            "is_category": False,
        })

    # Bulk insert
    if bulk_rows:
        for i in range(0, len(bulk_rows), _BULK_CHUNK):
            session.execute(insert(OverviewEntry), bulk_rows[i:i + _BULK_CHUNK])
        session.flush()
