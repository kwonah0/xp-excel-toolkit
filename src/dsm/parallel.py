"""Parallel import and split using multiprocessing with per-sheet temp SQLite DBs."""

from __future__ import annotations

import multiprocessing as mp
import sys
import tempfile
from pathlib import Path
from typing import Any

from sqlalchemy import create_engine, insert
from sqlalchemy.orm import Session, sessionmaker

from dsm.models import Base, ExcelCell, ExcelMerge, ExcelSheet, ExcelWorkbook


# ---------------------------------------------------------------------------
# Parallel import
# ---------------------------------------------------------------------------

def _import_sheet_worker(args: tuple) -> dict[str, Any]:
    """Worker: import one sheet into a temp SQLite DB.

    Each worker independently loads the xlsx and processes a single sheet.
    """
    xlsx_path, sheet_name, config_dict, tmp_dir = args

    import openpyxl
    from dsm.xlsx_parser import _import_ws

    # Resolve config from serializable names
    domain_cls = None
    field_map = None
    if config_dict.get("domain_cls_name"):
        from dsm.domain_models import Register, MemoryMapEntry, REGMAP_FIELD_MAP, MEMMAP_FIELD_MAP
        cls_map = {"Register": Register, "MemoryMapEntry": MemoryMapEntry}
        fmap_map = {"REGMAP_FIELD_MAP": REGMAP_FIELD_MAP, "MEMMAP_FIELD_MAP": MEMMAP_FIELD_MAP}
        domain_cls = cls_map.get(config_dict["domain_cls_name"])
        field_map = fmap_map.get(config_dict.get("field_map_name"))

    tmp_db_path = Path(tmp_dir) / f"dsm_{sheet_name}.db"

    try:
        wb_xl = openpyxl.load_workbook(xlsx_path, data_only=False)
        ws = wb_xl[sheet_name]

        engine = create_engine(f"sqlite:///{tmp_db_path}", echo=False)
        Base.metadata.create_all(engine)
        TmpSession = sessionmaker(bind=engine)

        with TmpSession() as session:
            wb_obj = ExcelWorkbook(filename=Path(xlsx_path).name, blob=None)
            session.add(wb_obj)
            session.flush()

            _import_ws(
                session, ws, wb_obj.id,
                header_row=config_dict.get("header_row"),
                field_map=field_map,
                domain_cls=domain_cls,
            )
            session.commit()

        wb_xl.close()
        engine.dispose()

        return {"tmp_db_path": str(tmp_db_path), "sheet_name": sheet_name,
                "success": True, "error": None}
    except Exception as e:
        return {"tmp_db_path": str(tmp_db_path), "sheet_name": sheet_name,
                "success": False, "error": str(e)}


def _merge_temp_db_into_main(
    main_session: Session,
    main_wb_id: int,
    tmp_db_path: str,
) -> ExcelSheet | None:
    """Merge one temp DB into the main DB.

    Reads from temp DB via separate sqlite3 connection,
    writes to main DB via the ORM session.
    Handles ID remapping for sheet_id and merge_id.
    """
    import json
    import sqlite3

    tmp_conn = sqlite3.connect(tmp_db_path)
    tmp_cur = tmp_conn.cursor()

    try:
        # 1. Read temp sheet info
        tmp_cur.execute("SELECT id, name, header_row FROM excel_sheet LIMIT 1")
        row = tmp_cur.fetchone()
        if not row:
            return None
        tmp_sheet_id, tmp_name, tmp_header_row = row

        # 2. Create sheet in main DB via ORM
        sheet_obj = ExcelSheet(workbook_id=main_wb_id, name=tmp_name, header_row=tmp_header_row)
        main_session.add(sheet_obj)
        main_session.flush()
        new_sheet_id = sheet_obj.id

        # 3. Read merges from temp and insert into main
        tmp_cur.execute(
            "SELECT id, min_row, min_col, max_row, max_col "
            "FROM excel_merge WHERE sheet_id = ?", (tmp_sheet_id,)
        )
        tmp_merges = tmp_cur.fetchall()

        merge_remap: dict[int, int] = {}
        if tmp_merges:
            merge_dicts = [
                {"sheet_id": new_sheet_id, "min_row": r[1], "min_col": r[2],
                 "max_row": r[3], "max_col": r[4]}
                for r in tmp_merges
            ]
            for i in range(0, len(merge_dicts), 500):
                main_session.execute(insert(ExcelMerge), merge_dicts[i:i + 500])
            main_session.flush()

            # Build remap: (min_row, min_col) -> new_id
            new_merges = (
                main_session.query(ExcelMerge.id, ExcelMerge.min_row, ExcelMerge.min_col)
                .filter(ExcelMerge.sheet_id == new_sheet_id)
                .all()
            )
            new_lookup = {(mr, mc): mid for mid, mr, mc in new_merges}
            for tmp_id, min_r, min_c, _, _ in tmp_merges:
                new_id = new_lookup.get((min_r, min_c))
                if new_id:
                    merge_remap[tmp_id] = new_id

        # 4. Copy cells with merge_id remap
        tmp_cur.execute(
            "SELECT row, col, raw_value, style, comment, merge_id, is_merge_origin "
            "FROM excel_cell WHERE sheet_id = ?", (tmp_sheet_id,)
        )
        cell_rows = tmp_cur.fetchall()
        if cell_rows:
            cell_dicts = [
                {"sheet_id": new_sheet_id, "row": r, "col": c,
                 "raw_value": raw_val,
                 "style": json.loads(style) if isinstance(style, str) else style,
                 "comment": comment_text,
                 "merge_id": merge_remap.get(old_mid) if old_mid else None,
                 "is_merge_origin": bool(is_origin)}
                for r, c, raw_val, style, comment_text, old_mid, is_origin in cell_rows
            ]
            for i in range(0, len(cell_dicts), 500):
                main_session.execute(insert(ExcelCell), cell_dicts[i:i + 500])

        # 5. Copy Register rows
        tmp_cur.execute(
            "SELECT type, indx, page, para, name, "
            "d7, d6, d5, d4, d3, d2, d1, d0, init, excel_row "
            "FROM register WHERE sheet_id = ?", (tmp_sheet_id,)
        )
        reg_rows = tmp_cur.fetchall()
        if reg_rows:
            from dsm.domain_models import Register
            reg_dicts = [
                {"sheet_id": new_sheet_id, "type": r[0], "indx": r[1],
                 "page": r[2], "para": r[3], "name": r[4],
                 "d7": r[5], "d6": r[6], "d5": r[7], "d4": r[8],
                 "d3": r[9], "d2": r[10], "d1": r[11], "d0": r[12],
                 "init": r[13], "excel_row": r[14]}
                for r in reg_rows
            ]
            for i in range(0, len(reg_dicts), 500):
                main_session.execute(insert(Register), reg_dicts[i:i + 500])

        # 6. Copy MemoryMapEntry rows
        try:
            tmp_cur.execute(
                "SELECT baseaddr, [group], midgroup, comment, special, excel_row "
                "FROM memorymap_entry WHERE sheet_id = ?", (tmp_sheet_id,)
            )
            mm_rows = tmp_cur.fetchall()
            if mm_rows:
                from dsm.domain_models import MemoryMapEntry
                mm_dicts = [
                    {"sheet_id": new_sheet_id, "baseaddr": r[0], "group": r[1],
                     "midgroup": r[2], "comment": r[3], "special": r[4],
                     "excel_row": r[5]}
                    for r in mm_rows
                ]
                for i in range(0, len(mm_dicts), 500):
                    main_session.execute(insert(MemoryMapEntry), mm_dicts[i:i + 500])
        except Exception:
            pass  # table might not exist if no memorymap data

        main_session.flush()
        return sheet_obj

    finally:
        tmp_conn.close()


def parallel_import_xlsx(
    session: Session,
    path: str | Path,
    *,
    sheet_configs: dict | None = None,
    workers: int | None = None,
) -> list[ExcelSheet]:
    """Import all sheets from xlsx in parallel.

    Each sheet is processed by a separate worker into a temp SQLite DB,
    then merged into the main DB.
    """
    path = Path(path)
    blob = path.read_bytes()

    if sheet_configs is None:
        from dsm.domain_models import _default_sheet_configs
        sheet_configs = _default_sheet_configs()

    import openpyxl
    wb_xl = openpyxl.load_workbook(path, data_only=False)
    sheet_names = wb_xl.sheetnames
    wb_xl.close()

    # Create main workbook record
    wb_obj = ExcelWorkbook(filename=path.name, blob=blob)
    session.add(wb_obj)
    session.flush()

    # Prepare serializable config per sheet
    from dsm.xlsx_parser import _match_config
    from dsm.domain_models import Register, MemoryMapEntry, REGMAP_FIELD_MAP, MEMMAP_FIELD_MAP

    worker_args = []
    with tempfile.TemporaryDirectory(prefix="dsm_import_") as tmp_dir:
        for name in sheet_names:
            config = _match_config(name, sheet_configs)
            config_dict: dict[str, Any] = {
                "field_map_name": None,
                "domain_cls_name": None,
                "header_row": None,
            }
            if config:
                if config.field_map is REGMAP_FIELD_MAP:
                    config_dict["field_map_name"] = "REGMAP_FIELD_MAP"
                elif config.field_map is MEMMAP_FIELD_MAP:
                    config_dict["field_map_name"] = "MEMMAP_FIELD_MAP"
                if config.domain_cls is Register:
                    config_dict["domain_cls_name"] = "Register"
                elif config.domain_cls is MemoryMapEntry:
                    config_dict["domain_cls_name"] = "MemoryMapEntry"
                config_dict["header_row"] = config.header_row

            worker_args.append((str(path), name, config_dict, tmp_dir))

        num_workers = workers or min(mp.cpu_count(), len(sheet_names))
        with mp.Pool(processes=num_workers) as pool:
            results = pool.map(_import_sheet_worker, worker_args)

        # Merge temp DBs into main DB (sequential)
        sheets: list[ExcelSheet] = []
        for result in results:
            if result["success"]:
                sheet_obj = _merge_temp_db_into_main(
                    session, wb_obj.id, result["tmp_db_path"],
                )
                if sheet_obj:
                    sheets.append(sheet_obj)
            else:
                print(
                    f"Warning: Failed to import sheet {result['sheet_name']}: "
                    f"{result['error']}",
                    file=sys.stderr,
                )

    session.flush()
    return sheets


# ---------------------------------------------------------------------------
# Parallel split
# ---------------------------------------------------------------------------

def _split_sheet_worker(args: tuple) -> dict[str, Any]:
    """Worker: split one level2 sheet into per-IP xlsx.

    Reads from main DB (read-only), builds workspace in :memory: DB,
    writes output xlsx file.
    """
    main_db_path, sheet_id, sheet_name, output_path = args

    from sqlalchemy import create_engine, distinct, insert
    from sqlalchemy.orm import sessionmaker
    from dsm.models import Base, ExcelWorkbook, ExcelSheet, ExcelCell, ExcelMerge
    from dsm.domain_models import Register
    from dsm.merge import MergeResolver
    from dsm.splitter import _build_ip_sheet, _export_multi_sheet

    _CHUNK = 500

    try:
        # Read source data from main DB
        main_engine = create_engine(f"sqlite:///{main_db_path}", echo=False)
        MainSession = sessionmaker(bind=main_engine)

        # Create in-memory workspace
        mem_engine = create_engine("sqlite:///:memory:", echo=False)
        Base.metadata.create_all(mem_engine)
        MemSession = sessionmaker(bind=mem_engine)

        with MainSession() as main_s, MemSession() as mem_s:
            # Create workbook + sheet in memory
            mem_wb = ExcelWorkbook(filename=f"{sheet_name}.xlsx")
            mem_s.add(mem_wb)
            mem_s.flush()

            main_sheet = main_s.get(ExcelSheet, sheet_id)
            mem_sheet = ExcelSheet(
                workbook_id=mem_wb.id,
                name=main_sheet.name,
                header_row=main_sheet.header_row,
            )
            mem_s.add(mem_sheet)
            mem_s.flush()

            # Copy merges
            merges = main_s.query(ExcelMerge).filter_by(sheet_id=sheet_id).all()
            merge_remap: dict[int, int] = {}
            if merges:
                m_dicts = [
                    {"sheet_id": mem_sheet.id, "min_row": m.min_row,
                     "min_col": m.min_col, "max_row": m.max_row, "max_col": m.max_col}
                    for m in merges
                ]
                for i in range(0, len(m_dicts), _CHUNK):
                    mem_s.execute(insert(ExcelMerge), m_dicts[i:i + _CHUNK])
                mem_s.flush()

                # Build remap by (min_row, min_col)
                new_merges = (
                    mem_s.query(ExcelMerge.id, ExcelMerge.min_row, ExcelMerge.min_col)
                    .filter(ExcelMerge.sheet_id == mem_sheet.id)
                    .all()
                )
                new_lookup = {(r, c): mid for mid, r, c in new_merges}
                for orig in merges:
                    new_id = new_lookup.get((orig.min_row, orig.min_col))
                    if new_id:
                        merge_remap[orig.id] = new_id

            # Copy cells
            cells = main_s.query(ExcelCell).filter_by(sheet_id=sheet_id).all()
            if cells:
                c_dicts = [
                    {"sheet_id": mem_sheet.id, "row": c.row, "col": c.col,
                     "raw_value": c.raw_value, "style": c.style,
                     "comment": c.comment,
                     "merge_id": merge_remap.get(c.merge_id) if c.merge_id else None,
                     "is_merge_origin": c.is_merge_origin}
                    for c in cells
                ]
                for i in range(0, len(c_dicts), _CHUNK):
                    mem_s.execute(insert(ExcelCell), c_dicts[i:i + _CHUNK])
                mem_s.flush()

            # Copy registers
            regs = main_s.query(Register).filter_by(sheet_id=sheet_id).all()
            if regs:
                r_dicts = [
                    {"sheet_id": mem_sheet.id, "excel_row": r.excel_row,
                     "type": r.type, "indx": r.indx, "page": r.page,
                     "para": r.para, "name": r.name,
                     "d7": r.d7, "d6": r.d6, "d5": r.d5, "d4": r.d4,
                     "d3": r.d3, "d2": r.d2, "d1": r.d1, "d0": r.d0,
                     "init": r.init}
                    for r in regs
                ]
                for i in range(0, len(r_dicts), _CHUNK):
                    mem_s.execute(insert(Register), r_dicts[i:i + _CHUNK])
                mem_s.flush()

            # Run split logic
            merger = MergeResolver.from_db(mem_s, mem_sheet.id)

            ip_names = (
                mem_s.query(distinct(Register.name))
                .filter(Register.sheet_id == mem_sheet.id, Register.name.isnot(None))
                .all()
            )

            out_wb = ExcelWorkbook(filename=f"{sheet_name}.xlsx")
            mem_s.add(out_wb)
            mem_s.flush()

            ip_sheets: list[int] = []
            for (ip_name,) in ip_names:
                ip_name = ip_name.strip()
                if not ip_name:
                    continue
                ip_regs = (
                    mem_s.query(Register)
                    .filter(Register.sheet_id == mem_sheet.id, Register.name == ip_name)
                    .order_by(Register.excel_row)
                    .all()
                )
                if not ip_regs:
                    continue
                ip_sheet = _build_ip_sheet(
                    mem_s, mem_sheet, merger, ip_name, ip_regs, out_wb.id,
                )
                ip_sheets.append(ip_sheet.id)

            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            _export_multi_sheet(mem_s, ip_sheets, Path(output_path))

        main_engine.dispose()
        mem_engine.dispose()

        return {"sheet_name": sheet_name, "output_path": output_path,
                "success": True, "error": None}
    except Exception as e:
        return {"sheet_name": sheet_name, "output_path": output_path,
                "success": False, "error": str(e)}


def parallel_split_regmap(
    session: Session,
    xlsx_path: str | Path,
    output_dir: str | Path,
    *,
    workers: int | None = None,
) -> dict[str, Path]:
    """Split regmap in parallel: one worker per level2 sheet.

    Requires data to already be imported into the DB.
    """
    xlsx_path = Path(xlsx_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Get main DB file path from session
    engine = session.get_bind()
    db_url = str(engine.url)
    main_db_path = db_url.replace("sqlite:///", "")

    wb_obj = (
        session.query(ExcelWorkbook)
        .filter_by(filename=xlsx_path.name)
        .first()
    )
    if not wb_obj:
        raise ValueError("Workbook not found in DB. Run 'dsm import' first.")

    level2_sheets = (
        session.query(ExcelSheet)
        .filter(ExcelSheet.workbook_id == wb_obj.id, ExcelSheet.name.like("level2_%"))
        .all()
    )

    worker_args = [
        (main_db_path, s.id, s.name, str(output_dir / f"{s.name}.xlsx"))
        for s in level2_sheets
    ]

    num_workers = workers or min(mp.cpu_count(), len(worker_args))
    with mp.Pool(processes=num_workers) as pool:
        results = pool.map(_split_sheet_worker, worker_args)

    output_map: dict[str, Path] = {}
    for result in results:
        if result["success"]:
            output_map[result["sheet_name"]] = Path(result["output_path"])
        else:
            print(
                f"Warning: Failed to split {result['sheet_name']}: {result['error']}",
                file=sys.stderr,
            )

    return output_map
