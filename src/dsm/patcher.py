"""Patch merge: apply split-file edits back onto the original xlsx."""

from __future__ import annotations

import io
import multiprocessing as mp
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment

from dsm.domain_models import REGMAP_FIELD_MAP, Register
from dsm.models import ExcelSheet, ExcelWorkbook, init_db


_REG_FIELDS = list(REGMAP_FIELD_MAP.values())


@dataclass
class CellChange:
    """One cell that was patched."""

    sheet_name: str
    ip_name: str
    row: int
    col: int
    field: str
    old_value: str | None
    new_value: str | None


@dataclass
class PatchResult:
    """Summary of a patch merge operation."""

    changes: list[CellChange] = field(default_factory=list)
    skipped_keys: list[tuple] = field(default_factory=list)
    patched_path: Path | None = None


def _norm(val) -> str | None:
    """Normalise a cell value to str for comparison."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _resolve_merged_value(ws, row: int, col: int):
    """Read value from a possibly-merged cell in an openpyxl worksheet."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if (mr.min_row <= row <= mr.max_row
                    and mr.min_col <= col <= mr.max_col):
                return ws.cell(row=mr.min_row, column=mr.min_col).value
        return None
    return cell.value


def _parse_split_registers(split_path: Path) -> list[dict]:
    """Parse registers from a split xlsx (no DB needed).

    Each worksheet tab in the split file is one IP.
    Row 1 = header, rows 2+ = data.

    Returns list of dicts with REGMAP fields + ``_ip_tab`` (tab name).
    """
    wb = openpyxl.load_workbook(split_path, data_only=False)
    results: list[dict] = []

    for ws in wb.worksheets:
        ip_name = ws.title

        # Build field map from header
        col_to_field: dict[int, str] = {}
        for cell in ws[1]:
            if cell.value and isinstance(cell.value, str):
                header = cell.value.strip()
                if header in REGMAP_FIELD_MAP:
                    col_to_field[cell.column] = REGMAP_FIELD_MAP[header]

        max_row = ws.max_row or 1
        for row_idx in range(2, max_row + 1):
            row_data: dict[str, str | None] = {"_ip_tab": ip_name}
            has_data = False
            for col_idx, field_name in col_to_field.items():
                val = _resolve_merged_value(ws, row_idx, col_idx)
                normed = _norm(val)
                row_data[field_name] = normed
                if normed is not None:
                    has_data = True

            # Capture comments keyed by field
            comments: dict[str, str] = {}
            for col_idx, field_name in col_to_field.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell) and cell.comment:
                    comments[field_name] = cell.comment.text
            if comments:
                row_data["_comments"] = comments

            if has_data:
                results.append(row_data)

    wb.close()
    return results


def _parse_split_worker(path_str: str) -> dict[str, Any]:
    """Multiprocessing worker: parse one split xlsx file."""
    try:
        path = Path(path_str)
        regs = _parse_split_registers(path)
        return {"stem": path.stem, "regs": regs, "success": True, "error": None}
    except Exception as e:
        return {"stem": Path(path_str).stem, "regs": [], "success": False, "error": str(e)}


def patch_merge(
    db_path: Path,
    split_dir: Path,
    output_path: Path,
    on_progress: Callable[[str], None] | None = None,
) -> PatchResult:
    """Apply edits from split xlsx files back onto the original xlsx.

    The original workbook is loaded from the DB blob so all formatting,
    non-level2 sheets, and extra columns are preserved.  Only cells whose
    register-field values actually changed are overwritten.
    """
    split_dir = Path(split_dir)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    Session = init_db(f"sqlite:///{db_path}")
    result = PatchResult(patched_path=output_path)

    with Session() as session:
        # --- Load original workbook blob ---
        if on_progress:
            on_progress("Loading original workbook from DB...")
        wb_obj = session.query(ExcelWorkbook).first()
        if not wb_obj or not wb_obj.blob:
            raise ValueError("No workbook blob found in DB")

        wb = openpyxl.load_workbook(io.BytesIO(wb_obj.blob))

        # --- Build original register index: key → Register ---
        if on_progress:
            on_progress("Building register index from DB...")
        original_regs: dict[tuple, Register] = {}
        sheet_names: dict[int, str] = {}
        level2_sheets: list[ExcelSheet] = (
            session.query(ExcelSheet)
            .filter(
                ExcelSheet.workbook_id == wb_obj.id,
                ExcelSheet.name.like("level2_%"),
            )
            .all()
        )
        for sheet_obj in level2_sheets:
            sheet_names[sheet_obj.id] = sheet_obj.name
            regs = (
                session.query(Register)
                .filter_by(sheet_id=sheet_obj.id)
                .all()
            )
            for reg in regs:
                key = (sheet_obj.name, reg.name, reg.indx, reg.page, reg.para)
                original_regs[key] = reg

        # --- Build column_map per sheet from workbook header ---
        column_maps: dict[str, dict[str, int]] = {}
        for sheet_obj in level2_sheets:
            if not sheet_obj.header_row:
                continue
            ws = wb[sheet_obj.name]
            col_map: dict[str, int] = {}
            for cell in ws[sheet_obj.header_row]:
                if cell.value and isinstance(cell.value, str):
                    header = cell.value.strip()
                    if header in REGMAP_FIELD_MAP:
                        col_map[REGMAP_FIELD_MAP[header]] = cell.column
            column_maps[sheet_obj.name] = col_map

        # --- Parse all split files in parallel ---
        split_files = sorted(split_dir.glob("*.xlsx"))
        relevant_files = [f for f in split_files if f.stem in column_maps]
        if on_progress:
            on_progress(f"Parsing {len(relevant_files)} split files...")

        parsed_map: dict[str, list[dict]] = {}
        if relevant_files:
            num_workers = min(mp.cpu_count(), len(relevant_files))
            with mp.Pool(processes=num_workers) as pool:
                results_list = pool.map(
                    _parse_split_worker,
                    [str(f) for f in relevant_files],
                )
            for r in results_list:
                if r["success"]:
                    parsed_map[r["stem"]] = r["regs"]

        # --- Apply parsed data to workbook ---
        if on_progress:
            on_progress("Applying patches to workbook...")
        for split_file in relevant_files:
            source_sheet = split_file.stem
            split_regs = parsed_map.get(source_sheet, [])
            if not split_regs:
                continue

            col_map = column_maps[source_sheet]
            ws = wb[source_sheet]

            for reg_data in split_regs:
                key = (
                    source_sheet,
                    reg_data.get("name"),
                    reg_data.get("indx"),
                    reg_data.get("page"),
                    reg_data.get("para"),
                )
                orig = original_regs.get(key)
                if orig is None:
                    result.skipped_keys.append(key)
                    continue

                target_row = orig.excel_row
                ip_name = reg_data.get("_ip_tab", "")

                # Compare and patch each field
                for field_name in _REG_FIELDS:
                    split_val = reg_data.get(field_name)
                    orig_val = _norm(getattr(orig, field_name, None))

                    if split_val == orig_val:
                        continue

                    col_idx = col_map.get(field_name)
                    if col_idx is None:
                        continue

                    # Write to original workbook
                    cell = ws.cell(row=target_row, column=col_idx)
                    if isinstance(cell, MergedCell):
                        for mr in ws.merged_cells.ranges:
                            if (mr.min_row <= target_row <= mr.max_row
                                    and mr.min_col <= col_idx <= mr.max_col):
                                ws.cell(
                                    row=mr.min_row, column=mr.min_col,
                                ).value = split_val
                                break
                    else:
                        cell.value = split_val

                    result.changes.append(CellChange(
                        sheet_name=source_sheet,
                        ip_name=ip_name,
                        row=target_row,
                        col=col_idx,
                        field=field_name,
                        old_value=orig_val,
                        new_value=split_val,
                    ))

                # Patch comments
                for field_name, comment_text in reg_data.get("_comments", {}).items():
                    col_idx = col_map.get(field_name)
                    if col_idx is None:
                        continue
                    cell = ws.cell(row=target_row, column=col_idx)
                    if isinstance(cell, MergedCell):
                        continue
                    existing = cell.comment.text if cell.comment else None
                    if comment_text != existing:
                        cell.comment = Comment(comment_text, "")

    if on_progress:
        on_progress(f"Saving patched workbook: {output_path.name}")
    wb.save(output_path)
    wb.close()
    return result
