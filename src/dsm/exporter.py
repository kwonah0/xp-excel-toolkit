"""Export Excel data from the ORM model back to .xlsx files."""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from sqlalchemy.orm import Session

from dsm.models import ExcelCell, ExcelMerge, ExcelSheet, ExcelWorkbook
from dsm.domain_models import REGMAP_FIELD_MAP, Register


# Reverse map: field_name -> header_text
_FIELD_TO_HEADER = {v: k for k, v in REGMAP_FIELD_MAP.items()}


def export_regmap_xlsx(
    session: Session,
    workbook_id: int,
    output_path: str | Path,
    column_map: dict[str, int] | None = None,
) -> Path:
    """
    Write modified Register data back to .xlsx, restoring original formatting.

    Strategy: Load original BLOB -> overwrite only changed cell values -> save.
    This preserves 100% of the original formatting, merges, colors, etc.
    """
    output_path = Path(output_path)

    wb_obj = session.get(ExcelWorkbook, workbook_id)
    if not wb_obj or not wb_obj.blob:
        raise ValueError(f"Workbook {workbook_id} not found or has no BLOB")

    # Load original workbook from stored BLOB
    wb = openpyxl.load_workbook(io.BytesIO(wb_obj.blob))

    for sheet_obj in wb_obj.sheets:
        ws = wb[sheet_obj.name]

        if not sheet_obj.header_row:
            continue

        # Get all registers for this sheet
        registers = (
            session.query(Register)
            .filter_by(sheet_id=sheet_obj.id)
            .all()
        )

        if not column_map:
            # Build column_map from header row
            column_map = {}
            for cell in ws[sheet_obj.header_row]:
                if cell.value and isinstance(cell.value, str):
                    header = cell.value.strip()
                    if header in REGMAP_FIELD_MAP:
                        column_map[REGMAP_FIELD_MAP[header]] = cell.column

        for reg in registers:
            for field_name, col_idx in column_map.items():
                val = getattr(reg, field_name, None)
                if val is None:
                    continue
                cell = ws.cell(row=reg.excel_row, column=col_idx)
                if isinstance(cell, MergedCell):
                    # Write to the merge origin instead
                    for mr in ws.merged_cells.ranges:
                        if (mr.min_row <= reg.excel_row <= mr.max_row
                                and mr.min_col <= col_idx <= mr.max_col):
                            ws.cell(row=mr.min_row, column=mr.min_col).value = val
                            break
                else:
                    cell.value = val

    wb.save(output_path)
    wb.close()
    return output_path


def _apply_style(cell: Cell, style: dict | None) -> None:
    """Apply style dict from ExcelCell to an openpyxl cell."""
    if not style:
        return

    # Background color
    bg = style.get("bg_color")
    if bg:
        color = bg.lstrip("#")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Font
    font_kwargs = {}
    if style.get("font_bold"):
        font_kwargs["bold"] = True
    fc = style.get("font_color")
    if fc:
        font_kwargs["color"] = fc.lstrip("#")
    if font_kwargs:
        cell.font = Font(**font_kwargs)

    # Number format
    nf = style.get("number_format")
    if nf:
        cell.number_format = nf

    # Borders
    border_sides = {}
    for side_name in ("left", "right", "top", "bottom"):
        bs = style.get(f"border_{side_name}")
        if bs:
            border_sides[side_name] = Side(style=bs)
    if border_sides:
        cell.border = Border(**border_sides)


def export_from_cells(
    session: Session,
    sheet_id: int,
    output_path: str | Path,
) -> Path:
    """Build an .xlsx from ExcelCell and ExcelMerge records (no BLOB needed).

    This creates a fresh xlsx entirely from the DB cell data, suitable for
    partial exports like split-by-IP.
    """
    output_path = Path(output_path)

    sheet_obj = session.get(ExcelSheet, sheet_id)
    if not sheet_obj:
        raise ValueError(f"Sheet {sheet_id} not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_obj.name

    # Write cells
    cells = (
        session.query(ExcelCell)
        .filter(ExcelCell.sheet_id == sheet_id)
        .all()
    )
    for cell_rec in cells:
        # Restore formula objects if applicable
        if cell_rec.formula_type == "array" and cell_rec.formula_ref:
            value = ArrayFormula(ref=cell_rec.formula_ref, text=cell_rec.raw_value)
        elif cell_rec.formula_type == "dataTable" and cell_rec.formula_ref:
            value = DataTableFormula(ref=cell_rec.formula_ref)
        else:
            value = cell_rec.raw_value
        c = ws.cell(row=cell_rec.row, column=cell_rec.col, value=value)
        _apply_style(c, cell_rec.style)
        if cell_rec.comment:
            c.comment = Comment(cell_rec.comment, "")

    # Apply merges
    merges = (
        session.query(ExcelMerge)
        .filter(ExcelMerge.sheet_id == sheet_id)
        .all()
    )
    for m in merges:
        ws.merge_cells(
            start_row=m.min_row, start_column=m.min_col,
            end_row=m.max_row, end_column=m.max_col,
        )

    # Freeze header if set
    if sheet_obj.header_row:
        ws.freeze_panes = f"A{sheet_obj.header_row + 1}"

    wb.save(output_path)
    wb.close()
    return output_path
