"""openpyxl-based parser for .xlsx files."""

from __future__ import annotations

import fnmatch
from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import insert
from sqlalchemy.orm import Session

from dsm.merge import MergeResolver
from dsm.models import ExcelCell, ExcelMerge, ExcelSheet, ExcelWorkbook

# Max dicts per Core bulk insert execute() call (avoids SQLite variable limit)
_BULK_CHUNK = 500


@dataclass
class SheetConfig:
    """Per-sheet import configuration: which domain model and field mapping to use.

    Example usage::

        sheet_configs = {
            "level2_*": SheetConfig(
                field_map=REGMAP_FIELD_MAP,
                domain_cls=Register,
            ),
            "memorymap": SheetConfig(
                field_map=MEMMAP_FIELD_MAP,
                domain_cls=MemoryMapEntry,
            ),
        }
        sheets = import_xlsx(session, path, sheet_configs=sheet_configs)
    """
    field_map: dict[str, str] | None = None
    domain_cls: type | None = None
    header_row: int | None = None


def extract_style(cell: Cell) -> dict | None:
    """Extract visual style info from an openpyxl cell."""
    style: dict = {}

    # Background color
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
        rgb = fill.fgColor.rgb
        if isinstance(rgb, str) and len(rgb) == 8:
            style["bg_color"] = f"#{rgb[2:]}"  # strip alpha

    # Font
    font = cell.font
    if font:
        if font.bold:
            style["font_bold"] = True
        if font.color and font.color.rgb and font.color.rgb != "00000000":
            rgb = font.color.rgb
            if isinstance(rgb, str) and len(rgb) == 8:
                style["font_color"] = f"#{rgb[2:]}"

    # Number format
    if cell.number_format and cell.number_format != "General":
        style["number_format"] = cell.number_format

    # Borders
    border = cell.border
    if border:
        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(border, side_name)
            if side and side.style:
                style[f"border_{side_name}"] = side.style

    return style if style else None



def find_header_row(
    ws: Worksheet,
    field_map: dict[str, str] | None = None,
    min_cols: int = 3,
) -> int | None:
    """Auto-detect header row by matching field_map keys against cell values.

    If field_map is provided, finds the first row containing at least one
    of the expected header names. Otherwise falls back to the first row
    with >= min_cols non-empty string cells.
    """
    expected = {k.strip() for k in field_map} if field_map else None

    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row or 30)):
        values = {
            str(cell.value).strip()
            for cell in row
            if cell.value is not None and isinstance(cell.value, str) and cell.value.strip()
        }
        if expected:
            if values & expected:
                return row[0].row
        elif len(values) >= min_cols:
            return row[0].row
    return None


def _match_config(
    sheet_name: str,
    configs: dict[str, SheetConfig] | None,
) -> SheetConfig | None:
    """Find a SheetConfig matching the given sheet name.

    Checks exact match first, then falls back to fnmatch patterns.
    """
    if not configs:
        return None
    if sheet_name in configs:
        return configs[sheet_name]
    for pattern, config in configs.items():
        if fnmatch.fnmatch(sheet_name, pattern):
            return config
    return None


def _import_ws(
    session: Session,
    ws: Worksheet,
    wb_id: int,
    *,
    header_row: int | None = None,
    field_map: dict[str, str] | None = None,
    domain_cls: type | None = None,
    column_map: dict[str, int] | None = None,
) -> ExcelSheet:
    """Core: import a single openpyxl worksheet into the DB.

    This is the shared implementation used by both import_sheet and import_xlsx.
    """
    sheet_obj = ExcelSheet(workbook_id=wb_id, name=ws.title)
    session.add(sheet_obj)
    session.flush()

    sid = sheet_obj.id

    # -- Merge ranges (Core bulk insert) ------------------------------------
    merger = MergeResolver(ws)

    merge_dicts = [
        {
            "sheet_id": sid,
            "min_row": mr.min_row,
            "min_col": mr.min_col,
            "max_row": mr.max_row,
            "max_col": mr.max_col,
        }
        for mr in merger.ranges
    ]

    merge_id_map: dict[str, int] = {}  # "min_row:min_col" -> merge.id
    if merge_dicts:
        for i in range(0, len(merge_dicts), _BULK_CHUNK):
            session.execute(insert(ExcelMerge), merge_dicts[i:i + _BULK_CHUNK])
        session.flush()

        # Query back IDs for cell linkage
        for mid, mrow, mcol in (
            session.query(ExcelMerge.id, ExcelMerge.min_row, ExcelMerge.min_col)
            .filter(ExcelMerge.sheet_id == sid)
            .all()
        ):
            merge_id_map[f"{mrow}:{mcol}"] = mid

    # -- Cells (Core bulk insert) -------------------------------------------
    cell_dicts: list[dict] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            r, c = cell.row, cell.column

            if merger.is_merged(r, c):
                raw_val = merger.get_value(r, c)
            else:
                raw_val = str(cell.value) if cell.value is not None else None

            merge_key = None
            is_origin = merger.is_origin(r, c)
            origin = merger.get_origin(r, c)
            if origin:
                merge_key = merge_id_map.get(f"{origin[0]}:{origin[1]}")

            style = extract_style(cell) if cell.value is not None or is_origin else None
            comment_text = cell.comment.text if cell.comment else None

            cell_dicts.append({
                "sheet_id": sid,
                "row": r,
                "col": c,
                "raw_value": raw_val,
                "style": style,
                "comment": comment_text,
                "merge_id": merge_key,
                "is_merge_origin": is_origin,
            })

    if cell_dicts:
        for i in range(0, len(cell_dicts), _BULK_CHUNK):
            session.execute(insert(ExcelCell), cell_dicts[i:i + _BULK_CHUNK])
        session.flush()

    # -- Header detection + domain object creation (Core bulk insert) -------
    if header_row is None:
        header_row = find_header_row(ws, field_map=field_map)
    sheet_obj.header_row = header_row

    if header_row and field_map and domain_cls:
        headers: dict[int, str] = {}
        for cell in ws[header_row]:
            if cell.value and isinstance(cell.value, str):
                headers[cell.column] = cell.value.strip()

        col_to_field: dict[int, str] = {}
        for col_idx, header_text in headers.items():
            if header_text in field_map:
                domain_field = field_map[header_text]
                col_to_field[col_idx] = domain_field
                if column_map is not None:
                    column_map[domain_field] = col_idx

        bulk_rows: list[dict[str, Any]] = []
        for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
            row_data: dict[str, Any] = {}
            has_data = False
            for col_idx, field_name in col_to_field.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is None and merger.is_merged(row_idx, col_idx):
                    val = merger.get_value(row_idx, col_idx)
                if val is not None:
                    has_data = True
                    row_data[field_name] = str(val)
                else:
                    row_data[field_name] = None

            if has_data:
                row_data["sheet_id"] = sid
                row_data["excel_row"] = row_idx
                bulk_rows.append(row_data)

        if bulk_rows:
            for i in range(0, len(bulk_rows), _BULK_CHUNK):
                session.execute(insert(domain_cls), bulk_rows[i:i + _BULK_CHUNK])
            session.flush()

    session.flush()
    return sheet_obj


def import_sheet(
    session: Session,
    path: str | Path,
    sheet_name: str | None = None,
    header_row: int | None = None,
    field_map: dict[str, str] | None = None,
    domain_cls: type | None = None,
    column_map: dict[str, int] | None = None,
) -> ExcelSheet:
    """Import a single sheet from a .xlsx file.

    Opens the file, creates an ExcelWorkbook record, and imports one sheet.
    For loading all sheets at once, use import_xlsx instead.

    Args:
        session: SQLAlchemy session.
        path: Path to .xlsx file.
        sheet_name: Target sheet name (default: active sheet).
        header_row: 1-based header row index (auto-detected if None).
        field_map: Mapping of Excel column header -> domain field name.
        domain_cls: ORM class to instantiate for each data row.
        column_map: Optional dict to populate with {field_name: col_index}.

    Returns:
        The created ExcelSheet ORM object.
    """
    path = Path(path)
    blob = path.read_bytes()

    wb_xl = openpyxl.load_workbook(path, data_only=False)
    ws = wb_xl[sheet_name] if sheet_name else wb_xl.active

    wb_obj = ExcelWorkbook(filename=path.name, blob=blob)
    session.add(wb_obj)
    session.flush()

    sheet_obj = _import_ws(
        session, ws, wb_obj.id,
        header_row=header_row,
        field_map=field_map,
        domain_cls=domain_cls,
        column_map=column_map,
    )

    wb_xl.close()
    return sheet_obj


def import_xlsx(
    session: Session,
    path: str | Path,
    *,
    sheet_configs: dict[str, SheetConfig] | None = None,
    on_progress: Callable[[str], None] | None = None,
) -> list[ExcelSheet]:
    """Import all sheets from a .xlsx file.

    Opens the file once, creates a single ExcelWorkbook record, and imports
    every sheet. Each sheet is matched against sheet_configs by name
    (supports fnmatch patterns like ``"level2_*"``).

    Args:
        session: SQLAlchemy session.
        path: Path to .xlsx file.
        sheet_configs: Mapping of sheet name pattern -> SheetConfig.
            Sheets without a matching config are still imported as raw
            cells (no domain objects).

    Returns:
        List of created ExcelSheet ORM objects, one per sheet.

    Example::

        from dsm import (
            SheetConfig, REGMAP_FIELD_MAP, Register,
            MEMMAP_FIELD_MAP, MemoryMapEntry,
        )

        sheets = import_xlsx(session, "regmap.xlsx", sheet_configs={
            "level2_*": SheetConfig(
                field_map=REGMAP_FIELD_MAP,
                domain_cls=Register,
            ),
            "memorymap": SheetConfig(
                field_map=MEMMAP_FIELD_MAP,
                domain_cls=MemoryMapEntry,
            ),
        })
    """
    path = Path(path)
    blob = path.read_bytes()

    if sheet_configs is None:
        sheet_configs = _load_configs_from_db(session)

    if on_progress:
        on_progress(f"Loading workbook: {path.name}")
    wb_xl = openpyxl.load_workbook(path, data_only=False)

    wb_obj = ExcelWorkbook(filename=path.name, blob=blob)
    session.add(wb_obj)
    session.flush()

    total = len(wb_xl.sheetnames)
    sheets: list[ExcelSheet] = []
    for idx, name in enumerate(wb_xl.sheetnames, 1):
        if on_progress:
            on_progress(f"  [{idx}/{total}] Importing sheet: {name}")
        config = _match_config(name, sheet_configs)
        ws = wb_xl[name]
        sheet_obj = _import_ws(
            session, ws, wb_obj.id,
            header_row=config.header_row if config else None,
            field_map=config.field_map if config else None,
            domain_cls=config.domain_cls if config else None,
        )
        sheets.append(sheet_obj)

    wb_xl.close()
    return sheets


def _load_configs_from_db(session: Session) -> dict[str, SheetConfig]:
    """Load SheetConfigEntry rows from DB and convert to SheetConfig dict.

    If no configs exist in DB, seeds defaults first.
    """
    import json
    from dsm.models import SheetConfigEntry
    from dsm.domain_models import DOMAIN_REGISTRY, FIELD_MAP_REGISTRY, seed_default_configs

    seed_default_configs(session)

    configs: dict[str, SheetConfig] = {}
    for entry in session.query(SheetConfigEntry).all():
        domain_cls = DOMAIN_REGISTRY.get(entry.domain_type) if entry.domain_type else None
        if entry.field_map_json:
            field_map = json.loads(entry.field_map_json)
        elif entry.domain_type:
            field_map = FIELD_MAP_REGISTRY.get(entry.domain_type)
        else:
            field_map = None
        configs[entry.pattern] = SheetConfig(
            field_map=field_map,
            domain_cls=domain_cls,
            header_row=entry.header_row,
        )
    return configs
