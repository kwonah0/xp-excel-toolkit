"""Export Excel data from the ORM model back to .xlsx files.

Two paths:

- :func:`export_domain_xlsx` — round-trip on top of the original BLOB stored
  in ``excel_workbook``. Original formatting, non-domain sheets, merges, and
  styles are preserved; only the cells belonging to your domain rows are
  overwritten. This is the path you want when editing register-map / pinmap /
  any structured spec via ORM models and re-emitting an xlsx that looks like
  the source.

- :func:`export_from_cells` — build a fresh xlsx purely from ``excel_cell``
  and ``excel_merge`` records, with style JSON and comments restored. No
  original BLOB needed. Useful for partial exports (e.g. splitting one sheet
  out of a workbook).
"""

from __future__ import annotations

import fnmatch
import io
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from excel_toolkit.models import ExcelCell, ExcelMerge, ExcelSheet, ExcelWorkbook


# ── Cell writing helpers ─────────────────────────────────────────────

def write_cell(ws: Worksheet, row: int, col: int, val) -> None:
    """Write a value to a cell, redirecting writes inside merge ranges to the
    origin cell so openpyxl's MergedCell guard doesn't raise."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if (mr.min_row <= row <= mr.max_row
                    and mr.min_col <= col <= mr.max_col):
                ws.cell(row=mr.min_row, column=mr.min_col).value = val
                break
    else:
        cell.value = val


def build_column_map(
    ws: Worksheet,
    header_row: int,
    field_map: dict[str, str],
) -> dict[str, int]:
    """Build {domain_field_name: column_index} mapping from a header row.

    ``field_map`` is the same {excel_header: domain_field} dict you pass to
    :class:`~excel_toolkit.SheetConfig`.
    """
    col_map: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value and isinstance(cell.value, str):
            header = cell.value.strip()
            if header in field_map:
                col_map[field_map[header]] = cell.column
    return col_map


def apply_style(cell: Cell, style: dict | None) -> None:
    """Apply a style dict (as produced by ``extract_style``) onto an openpyxl cell."""
    if not style:
        return

    bg = style.get("bg_color")
    if bg:
        color = bg.lstrip("#")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    font_kwargs = {}
    if style.get("font_bold"):
        font_kwargs["bold"] = True
    fc = style.get("font_color")
    if fc:
        font_kwargs["color"] = fc.lstrip("#")
    if font_kwargs:
        cell.font = Font(**font_kwargs)

    nf = style.get("number_format")
    if nf:
        cell.number_format = nf

    border_sides = {}
    for side_name in ("left", "right", "top", "bottom"):
        bs = style.get(f"border_{side_name}")
        if bs:
            border_sides[side_name] = Side(style=bs)
    if border_sides:
        cell.border = Border(**border_sides)


# Backwards-compatible aliases (underscored names were the original public
# surface used internally by sibling modules).
_write_cell = write_cell
_build_column_map = build_column_map
_apply_style = apply_style


# ── Domain round-trip ────────────────────────────────────────────────

# Signature: ``exporter_func(session, ws, sheet_obj) -> int`` (rows written).
# Use this when your domain layout is too irregular for the flat ``field_map``
# path — e.g. key-value sections written into specific column-A markers.
ExporterFunc = Callable[[Session, Worksheet, ExcelSheet], int]


@dataclass
class ExportHandler:
    """Per-sheet-pattern export configuration.

    Provide ``field_map`` + ``domain_cls`` for flat-table sheets (mirrors
    :class:`~excel_toolkit.SheetConfig`), or ``exporter_func`` for custom
    layouts.
    """

    pattern: str
    """Sheet-name fnmatch pattern (case-insensitive)."""

    field_map: dict[str, str] | None = None
    """{excel_header: domain_field}. Required when using domain_cls."""

    domain_cls: type | None = None
    """ORM class to query (filter_by sheet_id). Each row's ``excel_row`` must
    be set so the writer knows which row to overwrite."""

    exporter_func: ExporterFunc | None = None
    """Custom writer. Takes precedence over field_map/domain_cls when set."""


def _export_flat_handler(
    session: Session,
    ws: Worksheet,
    sheet_obj: ExcelSheet,
    domain_cls: type,
    field_map: dict[str, str],
) -> int:
    """Default flat-table export: ORM rows → cells via field_map."""
    if not sheet_obj.header_row:
        return 0

    col_map = build_column_map(ws, sheet_obj.header_row, field_map)
    if not col_map:
        return 0

    rows = (
        session.query(domain_cls)
        .filter_by(sheet_id=sheet_obj.id)
        .all()
    )

    for row_obj in rows:
        excel_row = getattr(row_obj, "excel_row", None)
        if excel_row is None:
            continue
        for field_name, col_idx in col_map.items():
            val = getattr(row_obj, field_name, None)
            write_cell(ws, excel_row, col_idx, val)

    return len(rows)


def export_domain_xlsx(
    session: Session,
    output_path: str | Path,
    handlers: list[ExportHandler],
    *,
    workbook_id: int | None = None,
    on_progress: Callable[[str], None] | None = None,
) -> Path:
    """Round-trip export: load the original BLOB and overwrite domain cells.

    All formatting, merges, styles, and non-domain sheets are preserved.

    Args:
        session: SQLAlchemy session.
        output_path: Destination xlsx path.
        handlers: List of :class:`ExportHandler` — first matching pattern per
            sheet wins.
        workbook_id: Explicit workbook to export. Defaults to the first
            workbook in the DB.
        on_progress: Optional progress callback.

    Raises:
        ValueError: If no workbook is found or its blob is missing.
    """
    output_path = Path(output_path)

    q = session.query(ExcelWorkbook)
    if workbook_id is not None:
        q = q.filter(ExcelWorkbook.id == workbook_id)
    wb_obj = q.first()
    if not wb_obj or not wb_obj.blob:
        raise ValueError("No workbook found in DB or BLOB is missing")

    wb = openpyxl.load_workbook(io.BytesIO(wb_obj.blob))

    total = 0
    for sheet_obj in wb_obj.sheets:
        if sheet_obj.name not in wb.sheetnames:
            continue
        ws = wb[sheet_obj.name]

        for handler in handlers:
            if not fnmatch.fnmatch(sheet_obj.name.lower(), handler.pattern.lower()):
                continue

            if handler.exporter_func is not None:
                count = handler.exporter_func(session, ws, sheet_obj)
            elif handler.domain_cls is not None and handler.field_map is not None:
                count = _export_flat_handler(
                    session, ws, sheet_obj, handler.domain_cls, handler.field_map,
                )
            else:
                count = 0

            if count and on_progress:
                on_progress(f"  {sheet_obj.name}: {count} rows")
            total += count
            break  # first match wins

    wb.save(output_path)
    wb.close()

    if on_progress:
        on_progress(f"Exported {total} domain rows to {output_path}")

    return output_path


# ── Cell-level export (no BLOB) ──────────────────────────────────────

def export_from_cells(
    session: Session,
    sheet_id: int,
    output_path: str | Path,
) -> Path:
    """Build an .xlsx from ExcelCell and ExcelMerge records (no BLOB needed).

    Creates a fresh xlsx entirely from the DB cell data. Suitable for partial
    exports (e.g. extracting a single sheet). Style JSON, comments, formula
    type/ref, and merges are restored.
    """
    output_path = Path(output_path)

    sheet_obj = session.get(ExcelSheet, sheet_id)
    if not sheet_obj:
        raise ValueError(f"Sheet {sheet_id} not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_obj.name

    cells = (
        session.query(ExcelCell)
        .filter(ExcelCell.sheet_id == sheet_id)
        .all()
    )
    for cell_rec in cells:
        if cell_rec.formula_type == "array" and cell_rec.formula_ref:
            value = ArrayFormula(ref=cell_rec.formula_ref, text=cell_rec.raw_value)
        elif cell_rec.formula_type == "dataTable" and cell_rec.formula_ref:
            value = DataTableFormula(ref=cell_rec.formula_ref)
        else:
            value = cell_rec.raw_value
        c = ws.cell(row=cell_rec.row, column=cell_rec.col, value=value)
        apply_style(c, cell_rec.style)
        if cell_rec.comment:
            c.comment = Comment(cell_rec.comment, "")

    merges = (
        session.query(ExcelMerge)
        .filter(ExcelMerge.sheet_id == sheet_id)
        .all()
    )
    for m in merges:
        ws.merge_cells(
            start_row=m.min_row, start_column=m.min_col,
            end_row=m.max_row, end_column=m.max_col,
        )

    if sheet_obj.header_row:
        ws.freeze_panes = f"A{sheet_obj.header_row + 1}"

    wb.save(output_path)
    wb.close()
    return output_path
