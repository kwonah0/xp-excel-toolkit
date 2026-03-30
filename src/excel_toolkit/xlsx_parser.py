"""openpyxl-based parser for .xlsx files."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell
from sqlalchemy.orm import Session

from excel_toolkit.merge import MergeResolver
from excel_toolkit.models import ExcelCell, ExcelMerge, ExcelSheet, ExcelWorkbook


def extract_style(cell: Cell) -> dict | None:
    """Extract visual style info from an openpyxl cell."""
    style: dict = {}

    # Background color
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
        rgb = fill.fgColor.rgb
        if isinstance(rgb, str) and len(rgb) == 8:
            style["bg_color"] = f"#{rgb[2:]}"  # strip alpha

    # Font
    font = cell.font
    if font:
        if font.bold:
            style["font_bold"] = True
        if font.color and font.color.rgb and font.color.rgb != "00000000":
            rgb = font.color.rgb
            if isinstance(rgb, str) and len(rgb) == 8:
                style["font_color"] = f"#{rgb[2:]}"

    # Number format
    if cell.number_format and cell.number_format != "General":
        style["number_format"] = cell.number_format

    # Borders
    border = cell.border
    if border:
        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(border, side_name)
            if side and side.style:
                style[f"border_{side_name}"] = side.style

    return style if style else None



def find_header_row(
    ws,
    field_map: dict[str, str] | None = None,
    min_cols: int = 3,
) -> int | None:
    """Auto-detect header row by matching field_map keys against cell values.

    If field_map is provided, finds the first row containing at least one
    of the expected header names. Otherwise falls back to the first row
    with >= min_cols non-empty string cells.
    """
    expected = {k.strip() for k in field_map} if field_map else None

    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row or 30)):
        values = {
            str(cell.value).strip()
            for cell in row
            if cell.value is not None and isinstance(cell.value, str) and cell.value.strip()
        }
        if expected:
            if values & expected:
                return row[0].row
        elif len(values) >= min_cols:
            return row[0].row
    return None


def import_xlsx(
    session: Session,
    path: str | Path,
    sheet_name: str | None = None,
    header_row: int | None = None,
    field_map: dict[str, str] | None = None,
    domain_cls: type | None = None,
    column_map: dict[str, int] | None = None,
) -> ExcelSheet:
    """
    Import a .xlsx file into the DB.

    Args:
        session: SQLAlchemy session
        path: Path to .xlsx file
        sheet_name: Target sheet name (default: first sheet)
        header_row: 1-based header row index (auto-detected if None)
        field_map: Mapping of Excel column header -> domain field name
                   e.g. {"TYPE": "type", "INDX": "indx", ...}
        domain_cls: ORM class to instantiate for each data row
                    (must have sheet_id, excel_row attributes)
        column_map: Optional dict to populate with {field_name: col_index}

    Returns:
        The created ExcelSheet ORM object.
    """
    path = Path(path)

    # Store original binary for round-trip
    blob = path.read_bytes()

    wb_xl = openpyxl.load_workbook(path, data_only=False)
    ws = wb_xl[sheet_name] if sheet_name else wb_xl.active

    # Create workbook record
    wb_obj = ExcelWorkbook(filename=path.name, blob=blob)
    session.add(wb_obj)
    session.flush()

    # Create sheet record
    sheet_obj = ExcelSheet(workbook_id=wb_obj.id, name=ws.title)
    session.add(sheet_obj)
    session.flush()

    # -- Merge ranges -------------------------------------------------------
    merger = MergeResolver(ws)
    merge_db: dict[str, ExcelMerge] = {}  # "min_row:min_col" -> ExcelMerge

    for mr in merger.ranges:
        m = ExcelMerge(
            sheet_id=sheet_obj.id,
            min_row=mr.min_row,
            min_col=mr.min_col,
            max_row=mr.max_row,
            max_col=mr.max_col,
        )
        session.add(m)
        session.flush()
        merge_db[f"{mr.min_row}:{mr.min_col}"] = m

    # -- Cells --------------------------------------------------------------
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            r, c = cell.row, cell.column

            # Determine value (fill merged cells with origin value)
            if merger.is_merged(r, c):
                raw_val = merger.get_value(r, c)
            else:
                raw_val = str(cell.value) if cell.value is not None else None

            # Merge linkage
            merge_key = None
            is_origin = merger.is_origin(r, c)
            origin = merger.get_origin(r, c)
            if origin:
                key = f"{origin[0]}:{origin[1]}"
                merge_key = merge_db[key].id

            # Style (only from the actual cell object, not merged placeholders)
            style = extract_style(cell) if cell.value is not None or is_origin else None

            cell_obj = ExcelCell(
                sheet_id=sheet_obj.id,
                row=r,
                col=c,
                raw_value=raw_val,
                style=style,
                merge_id=merge_key,
                is_merge_origin=is_origin,
            )
            session.add(cell_obj)

    session.flush()

    # -- Header detection + domain object creation --------------------------
    if header_row is None:
        header_row = find_header_row(ws, field_map=field_map)
    sheet_obj.header_row = header_row

    if header_row and field_map and domain_cls:
        # Build column index mapping from header row
        headers: dict[int, str] = {}
        for cell in ws[header_row]:
            if cell.value and isinstance(cell.value, str):
                headers[cell.column] = cell.value.strip()

        # Map Excel col index -> domain field
        col_to_field: dict[int, str] = {}
        for col_idx, header_text in headers.items():
            if header_text in field_map:
                domain_field = field_map[header_text]
                col_to_field[col_idx] = domain_field
                if column_map is not None:
                    column_map[domain_field] = col_idx

        # Create domain records for data rows
        for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
            row_data: dict[str, Any] = {}
            has_data = False
            for col_idx, field_name in col_to_field.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                # Fill from merge origin if needed
                if val is None and merger.is_merged(row_idx, col_idx):
                    val = merger.get_value(row_idx, col_idx)
                if val is not None:
                    has_data = True
                    row_data[field_name] = str(val)
                else:
                    row_data[field_name] = None

            if has_data:
                obj = domain_cls(
                    sheet_id=sheet_obj.id,
                    excel_row=row_idx,
                    **row_data,
                )
                session.add(obj)
                session.flush()

    session.flush()
    wb_xl.close()
    return sheet_obj
