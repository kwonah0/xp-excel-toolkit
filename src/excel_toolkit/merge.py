"""Merge cell resolver — works from openpyxl worksheet or DB records."""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet
    from sqlalchemy.orm import Session


class MergeResolver:
    """Resolve merged cell values for a sheet.

    Two construction paths:
      - MergeResolver(ws)          — from openpyxl worksheet (import time)
      - MergeResolver.from_db(...)  — from ExcelMerge + ExcelCell records (post-import)
    """

    def __init__(self, ws: Worksheet) -> None:
        self._merge_map: dict[tuple[int, int], tuple[int, int]] = {}
        self._origin_values: dict[tuple[int, int], str | None] = {}
        self.ranges = list(ws.merged_cells.ranges)

        for mr in self.ranges:
            origin = (mr.min_row, mr.min_col)
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    self._merge_map[(r, c)] = origin

            origin_cell = ws.cell(row=mr.min_row, column=mr.min_col)
            val = origin_cell.value
            if val is None:
                self._origin_values[origin] = None
            elif isinstance(val, ArrayFormula):
                self._origin_values[origin] = val.text if val.text else str(val.ref)
            elif isinstance(val, DataTableFormula):
                self._origin_values[origin] = None
            else:
                self._origin_values[origin] = str(val)

    @classmethod
    def from_db(cls, session: Session, sheet_id: int) -> MergeResolver:
        """Build a MergeResolver from DB records (no xlsx file needed)."""
        from excel_toolkit.models import ExcelCell, ExcelMerge

        resolver = cls.__new__(cls)
        resolver._merge_map = {}
        resolver._origin_values = {}
        resolver.ranges = []

        merges = (
            session.query(ExcelMerge)
            .filter(ExcelMerge.sheet_id == sheet_id)
            .all()
        )

        # Collect all origin coordinates to batch-query their values
        origin_coords: list[tuple[int, int]] = []

        for m in merges:
            resolver.ranges.append(m)
            origin = (m.min_row, m.min_col)
            origin_coords.append(origin)
            for r in range(m.min_row, m.max_row + 1):
                for c in range(m.min_col, m.max_col + 1):
                    resolver._merge_map[(r, c)] = origin

        # Batch-query origin cell values
        if origin_coords:
            origin_cells = (
                session.query(ExcelCell.row, ExcelCell.col, ExcelCell.raw_value)
                .filter(
                    ExcelCell.sheet_id == sheet_id,
                    ExcelCell.is_merge_origin.is_(True),
                )
                .all()
            )
            for row, col, raw_value in origin_cells:
                if (row, col) in resolver._merge_map:
                    resolver._origin_values[(row, col)] = raw_value

        return resolver

    def is_merged(self, row: int, col: int) -> bool:
        return (row, col) in self._merge_map

    def is_origin(self, row: int, col: int) -> bool:
        origin = self._merge_map.get((row, col))
        return origin == (row, col) if origin else False

    def get_origin(self, row: int, col: int) -> tuple[int, int] | None:
        return self._merge_map.get((row, col))

    def get_value(self, row: int, col: int) -> str | None:
        """Return the origin value for a merged cell."""
        origin = self._merge_map.get((row, col))
        return self._origin_values.get(origin) if origin else None
