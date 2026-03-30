"""Split a register map by sheet, with each IP as a sub-sheet."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from sqlalchemy import distinct
from sqlalchemy.orm import Session

from excel_toolkit.exporter import export_from_cells
from excel_toolkit.merge import MergeResolver
from excel_toolkit.models import ExcelCell, ExcelMerge, ExcelSheet, ExcelWorkbook
from excel_toolkit.domain_models import REGMAP_FIELD_MAP, Register
from excel_toolkit.xlsx_parser import import_xlsx


def _build_ip_sheet(
    session: Session,
    src_sheet: ExcelSheet,
    merger: MergeResolver,
    ip_name: str,
    registers: list[Register],
    dst_wb_id: int,
) -> ExcelSheet:
    """Create a new ExcelSheet in DB for one IP, with remapped rows."""
    src_rows = sorted({reg.excel_row for reg in registers})
    row_map = {src_r: dst_i for dst_i, src_r in enumerate(src_rows, start=2)}

    ip_sheet = ExcelSheet(
        workbook_id=dst_wb_id,
        name=ip_name,
        header_row=1,
    )
    session.add(ip_sheet)
    session.flush()

    # Copy header row cells
    header_cells = (
        session.query(ExcelCell)
        .filter(ExcelCell.sheet_id == src_sheet.id, ExcelCell.row == src_sheet.header_row)
        .all()
    )
    for cell in header_cells:
        session.add(ExcelCell(
            sheet_id=ip_sheet.id,
            row=1,
            col=cell.col,
            raw_value=cell.raw_value,
            style=cell.style,
        ))

    # Copy merges with row remapping
    merge_id_map: dict[int, int] = {}
    src_merges = (
        session.query(ExcelMerge)
        .filter(ExcelMerge.sheet_id == src_sheet.id)
        .all()
    )

    for merge in src_merges:
        covered = [r for r in src_rows if merge.min_row <= r <= merge.max_row]

        if merge.min_row == merge.max_row and merge.min_row in row_map:
            new_m = ExcelMerge(
                sheet_id=ip_sheet.id,
                min_row=row_map[merge.min_row],
                min_col=merge.min_col,
                max_row=row_map[merge.min_row],
                max_col=merge.max_col,
            )
            session.add(new_m)
            session.flush()
            merge_id_map[merge.id] = new_m.id
        elif len(covered) > 1:
            new_m = ExcelMerge(
                sheet_id=ip_sheet.id,
                min_row=row_map[covered[0]],
                min_col=merge.min_col,
                max_row=row_map[covered[-1]],
                max_col=merge.max_col,
            )
            session.add(new_m)
            session.flush()
            merge_id_map[merge.id] = new_m.id

    # Copy data cells, resolving merge values
    data_cells = (
        session.query(ExcelCell)
        .filter(
            ExcelCell.sheet_id == src_sheet.id,
            ExcelCell.row.in_(src_rows),
        )
        .all()
    )
    for cell in data_cells:
        raw_value = cell.raw_value
        if raw_value is None and merger.is_merged(cell.row, cell.col):
            raw_value = merger.get_value(cell.row, cell.col)

        new_merge_id = merge_id_map.get(cell.merge_id) if cell.merge_id else None
        is_origin = merger.is_origin(cell.row, cell.col) if cell.merge_id else False

        session.add(ExcelCell(
            sheet_id=ip_sheet.id,
            row=row_map[cell.row],
            col=cell.col,
            raw_value=raw_value,
            style=cell.style,
            merge_id=new_merge_id,
            is_merge_origin=is_origin,
        ))

    session.flush()
    return ip_sheet


def split_regmap(
    session: Session,
    path: str | Path,
    output_dir: str | Path,
) -> dict[str, Path]:
    """Split a register map xlsx: one output file per source sheet,
    with each IP as a separate sheet inside.

    e.g. regmap_sample.xlsx (sheets: level2_common, level2_buscon)
         → level2_common.xlsx  (sheets: SENSOR_A, AMPLIFIER, ...)
         → level2_buscon.xlsx  (sheets: GPIO_PORT, TIMER_A, ...)

    Args:
        session: SQLAlchemy session.
        path: Path to the source regmap xlsx.
        output_dir: Directory to write output files.

    Returns:
        Dict mapping source sheet name to output file path.
    """
    path = Path(path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    wb_xl = openpyxl.load_workbook(path, read_only=True)
    sheet_names = wb_xl.sheetnames
    wb_xl.close()

    results: dict[str, Path] = {}

    # Only process level2_* sheets (register bit-field specs)
    level2_sheets = [sn for sn in sheet_names if sn.startswith("level2_")]

    for sn in level2_sheets:
        src_sheet = import_xlsx(
            session, path,
            sheet_name=sn,
            field_map=REGMAP_FIELD_MAP,
            domain_cls=Register,
        )
        merger = MergeResolver.from_db(session, src_sheet.id)

        ip_names = (
            session.query(distinct(Register.name))
            .filter(Register.sheet_id == src_sheet.id, Register.name.isnot(None))
            .all()
        )
        if not ip_names:
            continue

        out_wb = ExcelWorkbook(filename=f"{sn}.xlsx")
        session.add(out_wb)
        session.flush()

        ip_sheets: list[int] = []

        for (ip_name,) in ip_names:
            ip_name = ip_name.strip()
            if not ip_name:
                continue

            registers = (
                session.query(Register)
                .filter(Register.sheet_id == src_sheet.id, Register.name == ip_name)
                .order_by(Register.excel_row)
                .all()
            )
            if not registers:
                continue

            ip_sheet = _build_ip_sheet(
                session, src_sheet, merger, ip_name, registers, out_wb.id,
            )
            ip_sheets.append(ip_sheet.id)

        out_path = output_dir / f"{sn}.xlsx"
        _export_multi_sheet(session, ip_sheets, out_path)
        results[sn] = out_path

    return results


def _export_multi_sheet(
    session: Session,
    sheet_ids: list[int],
    output_path: Path,
) -> None:
    """Export multiple ExcelSheet records into one xlsx file, each as a sheet."""
    from excel_toolkit.exporter import _apply_style

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for sheet_id in sheet_ids:
        sheet_obj = session.get(ExcelSheet, sheet_id)
        ws = wb.create_sheet(title=sheet_obj.name)

        # Write cells
        cells = (
            session.query(ExcelCell)
            .filter(ExcelCell.sheet_id == sheet_id)
            .all()
        )
        for cell_rec in cells:
            c = ws.cell(row=cell_rec.row, column=cell_rec.col, value=cell_rec.raw_value)
            _apply_style(c, cell_rec.style)

        # Apply merges
        merges = (
            session.query(ExcelMerge)
            .filter(ExcelMerge.sheet_id == sheet_id)
            .all()
        )
        for m in merges:
            ws.merge_cells(
                start_row=m.min_row, start_column=m.min_col,
                end_row=m.max_row, end_column=m.max_col,
            )

        if sheet_obj.header_row:
            ws.freeze_panes = f"A{sheet_obj.header_row + 1}"

    wb.save(output_path)
    wb.close()
