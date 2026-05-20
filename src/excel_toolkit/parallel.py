"""Parallel split using multiprocessing — one worker per level2 sheet."""

from __future__ import annotations

import multiprocessing as mp
import sys
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from excel_toolkit.models import ExcelSheet, ExcelWorkbook

def _split_sheet_worker(args: tuple) -> dict[str, Any]:
    """Worker: split one level2 sheet into per-IP xlsx.

    Reads directly from main DB (read-only) → builds xlsx without
    intermediate memory DB copy.
    """
    main_db_path, sheet_id, sheet_name, output_path = args

    import openpyxl
    from openpyxl.comments import Comment
    from sqlalchemy import create_engine, distinct
    from sqlalchemy.orm import sessionmaker
    from excel_toolkit.models import ExcelSheet, ExcelCell, ExcelMerge
    from excel_toolkit.domain_models import Register
    from excel_toolkit.merge import MergeResolver
    from excel_toolkit.exporter import _apply_style

    try:
        main_engine = create_engine(f"sqlite:///{main_db_path}", echo=False)
        MainSession = sessionmaker(bind=main_engine)

        with MainSession() as s:
            src_sheet = s.get(ExcelSheet, sheet_id)
            merger = MergeResolver.from_db(s, sheet_id)

            ip_names = (
                s.query(distinct(Register.name))
                .filter(Register.sheet_id == sheet_id, Register.name.isnot(None))
                .all()
            )
            if not ip_names:
                main_engine.dispose()
                return {"sheet_name": sheet_name, "output_path": output_path,
                        "success": True, "error": None}

            # Pre-load header cells and all source cells once
            header_cells = (
                s.query(ExcelCell)
                .filter(ExcelCell.sheet_id == sheet_id,
                        ExcelCell.row == src_sheet.header_row)
                .all()
            )

            # Pre-load all source merges
            src_merges = s.query(ExcelMerge).filter_by(sheet_id=sheet_id).all()

            # Build xlsx directly
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            for (ip_name,) in ip_names:
                ip_name = ip_name.strip()
                if not ip_name:
                    continue

                ip_regs = (
                    s.query(Register)
                    .filter(Register.sheet_id == sheet_id, Register.name == ip_name)
                    .order_by(Register.excel_row)
                    .all()
                )
                if not ip_regs:
                    continue

                # Build row map
                src_rows = sorted({reg.excel_row for reg in ip_regs})
                row_map = {src_r: dst_i for dst_i, src_r in enumerate(src_rows, start=2)}

                ws = wb.create_sheet(title=ip_name)

                # Write header
                for cell in header_cells:
                    c = ws.cell(row=1, column=cell.col, value=cell.raw_value)
                    _apply_style(c, cell.style)
                    if cell.comment:
                        c.comment = Comment(cell.comment, "")

                # Load data cells for this IP's rows
                data_cells = (
                    s.query(ExcelCell)
                    .filter(ExcelCell.sheet_id == sheet_id,
                            ExcelCell.row.in_(src_rows))
                    .all()
                )

                for cell in data_cells:
                    raw_value = cell.raw_value
                    if raw_value is None and merger.is_merged(cell.row, cell.col):
                        raw_value = merger.get_value(cell.row, cell.col)

                    c = ws.cell(row=row_map[cell.row], column=cell.col, value=raw_value)
                    _apply_style(c, cell.style)
                    if cell.comment:
                        c.comment = Comment(cell.comment, "")

                # Apply remapped merges
                for merge in src_merges:
                    covered = [r for r in src_rows if merge.min_row <= r <= merge.max_row]
                    if merge.min_row == merge.max_row and merge.min_row in row_map:
                        ws.merge_cells(
                            start_row=row_map[merge.min_row], start_column=merge.min_col,
                            end_row=row_map[merge.min_row], end_column=merge.max_col,
                        )
                    elif len(covered) > 1:
                        ws.merge_cells(
                            start_row=row_map[covered[0]], start_column=merge.min_col,
                            end_row=row_map[covered[-1]], end_column=merge.max_col,
                        )

                ws.freeze_panes = "A2"

            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            wb.close()

        main_engine.dispose()

        return {"sheet_name": sheet_name, "output_path": output_path,
                "success": True, "error": None}
    except Exception as e:
        return {"sheet_name": sheet_name, "output_path": output_path,
                "success": False, "error": str(e)}


def parallel_split_regmap(
    session: Session,
    xlsx_path: str | Path,
    output_dir: str | Path,
    *,
    workers: int | None = None,
) -> dict[str, Path]:
    """Split regmap in parallel: one worker per level2 sheet.

    Requires data to already be imported into the DB.
    """
    xlsx_path = Path(xlsx_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Get main DB file path from session
    engine = session.get_bind()
    db_url = str(engine.url)
    main_db_path = db_url.replace("sqlite:///", "")

    wb_obj = (
        session.query(ExcelWorkbook)
        .filter_by(filename=xlsx_path.name)
        .first()
    )
    if not wb_obj:
        raise ValueError("Workbook not found in DB. Run 'dsm import' first.")

    level2_sheets = (
        session.query(ExcelSheet)
        .filter(ExcelSheet.workbook_id == wb_obj.id, ExcelSheet.name.like("level2_%"))
        .all()
    )

    worker_args = [
        (main_db_path, s.id, s.name, str(output_dir / f"{s.name}.xlsx"))
        for s in level2_sheets
    ]

    num_workers = workers or min(mp.cpu_count(), len(worker_args))
    with mp.Pool(processes=num_workers) as pool:
        results = pool.map(_split_sheet_worker, worker_args)

    output_map: dict[str, Path] = {}
    for result in results:
        if result["success"]:
            output_map[result["sheet_name"]] = Path(result["output_path"])
        else:
            print(
                f"Warning: Failed to split {result['sheet_name']}: {result['error']}",
                file=sys.stderr,
            )

    return output_map
