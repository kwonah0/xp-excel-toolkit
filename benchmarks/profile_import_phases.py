"""Profile each sub-phase of sheet import in detail."""
import time
import tempfile
import openpyxl
from pathlib import Path

# --- Generate test xlsx (20 sheets × 700 rows × 15 cols, with merges & comments) ---
print("=== Generating test xlsx ===")
wb = openpyxl.Workbook()
wb.remove(wb.active)
for si in range(20):
    ws = wb.create_sheet(title=f"level2_sheet{si}")
    for c in range(1, 16):
        cell = ws.cell(row=1, column=c, value=f"Header_{c}")
        cell.font = openpyxl.styles.Font(bold=True, color="FF0000")
        cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", fill_type="solid")
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin"),
            bottom=openpyxl.styles.Side(style="medium"),
        )
    for r in range(2, 702):
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c, value=f"V_{si}_{r}_{c}")
            cell.font = openpyxl.styles.Font(bold=(c == 1))
            if r % 10 == 0:
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", fill_type="solid")
            if r % 50 == 0 and c == 1:
                cell.comment = openpyxl.comments.Comment(f"Note for row {r}", "Author")
    # merges
    for r in range(10, 700, 50):
        ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=1)
    ws.merge_cells("B20:C20")

xlsx_path = Path(tempfile.mktemp(suffix=".xlsx"))
wb.save(xlsx_path)
wb.close()
print(f"  {xlsx_path.stat().st_size / 1024 / 1024:.1f} MB")

# ============================================================
# Phase 1: load_workbook (normal vs read_only)
# ============================================================
print("\n=== Phase 1: load_workbook() ===")

t0 = time.perf_counter()
wb_normal = openpyxl.load_workbook(xlsx_path, data_only=False)
t_load_normal = time.perf_counter() - t0
print(f"  Normal:    {t_load_normal:.3f}s")

t0 = time.perf_counter()
wb_ro = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
t_load_ro = time.perf_counter() - t0
print(f"  Read-only: {t_load_ro:.4f}s  ({t_load_normal/t_load_ro:.0f}x faster)")

# ============================================================
# Phase 2: Per-sheet breakdown (normal mode, 1 sheet)
# ============================================================
print("\n=== Phase 2: Per-sheet breakdown (normal mode, 1 sheet = 700 rows × 15 cols) ===")

ws = wb_normal["level2_sheet0"]
from dsm.xlsx_parser import extract_style
from dsm.merge import MergeResolver

# 2a: merged_cells extraction
t0 = time.perf_counter()
merger = MergeResolver(ws)
t_merge = time.perf_counter() - t0
print(f"  2a. MergeResolver(ws):       {t_merge:.4f}s  ({len(merger.ranges)} ranges)")

# 2b: cell iteration only (no style, no comment)
t0 = time.perf_counter()
count = 0
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        _ = cell.value
        _ = cell.row
        _ = cell.column
        count += 1
t_iter_only = time.perf_counter() - t0
print(f"  2b. Cell iteration only:     {t_iter_only:.4f}s  ({count} cells)")

# 2c: cell iteration + value conversion
t0 = time.perf_counter()
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        val = str(cell.value) if cell.value is not None else None
t_iter_val = time.perf_counter() - t0
print(f"  2c. + str(value):            {t_iter_val:.4f}s")

# 2d: cell iteration + extract_style
t0 = time.perf_counter()
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        val = str(cell.value) if cell.value is not None else None
        s = extract_style(cell)
t_iter_style = time.perf_counter() - t0
print(f"  2d. + extract_style():       {t_iter_style:.4f}s  (delta: +{t_iter_style - t_iter_val:.4f}s)")

# 2e: cell iteration + comment
t0 = time.perf_counter()
comment_count = 0
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        val = str(cell.value) if cell.value is not None else None
        if cell.comment:
            _ = cell.comment.text
            comment_count += 1
t_iter_comment = time.perf_counter() - t0
print(f"  2e. + comment check:         {t_iter_comment:.4f}s  ({comment_count} comments, delta: +{t_iter_comment - t_iter_val:.4f}s)")

# 2f: cell iteration + merge resolve
t0 = time.perf_counter()
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        r, c = cell.row, cell.column
        if merger.is_merged(r, c):
            val = merger.get_value(r, c)
        else:
            val = str(cell.value) if cell.value is not None else None
        _ = merger.is_origin(r, c)
        _ = merger.get_origin(r, c)
t_iter_merge = time.perf_counter() - t0
print(f"  2f. + merge resolve:         {t_iter_merge:.4f}s  (delta: +{t_iter_merge - t_iter_val:.4f}s)")

# 2g: FULL pipeline (value + style + comment + merge) = what _import_ws does
t0 = time.perf_counter()
cell_dicts = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        r, c = cell.row, cell.column
        if merger.is_merged(r, c):
            raw_val = merger.get_value(r, c)
        else:
            raw_val = str(cell.value) if cell.value is not None else None
        is_origin = merger.is_origin(r, c)
        style = extract_style(cell) if cell.value is not None or is_origin else None
        comment_text = cell.comment.text if cell.comment else None
        origin = merger.get_origin(r, c)
        cell_dicts.append({
            "row": r, "col": c, "raw_value": raw_val,
            "style": style, "comment": comment_text,
            "is_merge_origin": is_origin,
        })
t_full = time.perf_counter() - t0
print(f"  2g. FULL pipeline:           {t_full:.4f}s  ({len(cell_dicts)} dicts)")

# 2h: DB bulk insert
from sqlalchemy import create_engine, insert
from sqlalchemy.orm import sessionmaker
from dsm.models import Base, ExcelCell, ExcelSheet, ExcelWorkbook, ExcelMerge

tmp_db = Path(tempfile.mktemp(suffix=".db"))
engine = create_engine(f"sqlite:///{tmp_db}", echo=False)
Base.metadata.create_all(engine)
S = sessionmaker(bind=engine)

with S() as session:
    wb_obj = ExcelWorkbook(filename="test.xlsx", blob=None)
    session.add(wb_obj)
    session.flush()
    sheet_obj = ExcelSheet(workbook_id=wb_obj.id, name="test", header_row=1)
    session.add(sheet_obj)
    session.flush()

    for d in cell_dicts:
        d["sheet_id"] = sheet_obj.id
        d["merge_id"] = None

    t0 = time.perf_counter()
    for i in range(0, len(cell_dicts), 500):
        session.execute(insert(ExcelCell), cell_dicts[i:i+500])
    session.commit()
    t_db_insert = time.perf_counter() - t0
print(f"  2h. DB bulk insert:          {t_db_insert:.4f}s  ({len(cell_dicts)} rows)")

engine.dispose()
tmp_db.unlink()

# ============================================================
# Phase 3: Same breakdown with read_only mode
# ============================================================
print("\n=== Phase 3: Per-sheet breakdown (read_only mode, 1 sheet) ===")

ws_ro = wb_ro["level2_sheet0"]

# 3a: cell iteration only
t0 = time.perf_counter()
count = 0
for row in ws_ro.iter_rows(min_row=1, max_row=701, max_col=15):
    for cell in row:
        _ = cell.value
        count += 1
t_ro_iter = time.perf_counter() - t0
print(f"  3a. Cell iteration only:     {t_ro_iter:.4f}s  ({count} cells)")

# 3b: + value conversion
wb_ro.close()
wb_ro = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
ws_ro = wb_ro["level2_sheet0"]
t0 = time.perf_counter()
for row in ws_ro.iter_rows(min_row=1, max_row=701, max_col=15):
    for cell in row:
        val = str(cell.value) if cell.value is not None else None
t_ro_val = time.perf_counter() - t0
print(f"  3b. + str(value):            {t_ro_val:.4f}s")

# Need to re-open because read_only iterators are consumed
wb_ro.close()
wb_ro = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
ws_ro = wb_ro["level2_sheet0"]

# 3c: + extract_style
t0 = time.perf_counter()
for row in ws_ro.iter_rows(min_row=1, max_row=701, max_col=15):
    for cell in row:
        val = str(cell.value) if cell.value is not None else None
        s = extract_style(cell)
t_ro_style = time.perf_counter() - t0
print(f"  3c. + extract_style():       {t_ro_style:.4f}s  (delta: +{t_ro_style - t_ro_val:.4f}s)")

# 3d: comment → NOT AVAILABLE
print(f"  3d. comment:                 NOT AVAILABLE (ReadOnlyCell has no .comment)")

# 3e: merged_cells → NOT AVAILABLE
print(f"  3e. merged_cells:            NOT AVAILABLE (ReadOnlyWorksheet)")

# ============================================================
# Phase 4: XML-based merge & comment extraction
# ============================================================
print("\n=== Phase 4: ZIP XML extraction (merge + comment) ===")
import zipfile
import xml.etree.ElementTree as ET
import re

# 4a: All sheets merge info
t0 = time.perf_counter()
merge_info = {}
with zipfile.ZipFile(xlsx_path) as zf:
    for entry in zf.namelist():
        if entry.startswith("xl/worksheets/sheet") and entry.endswith(".xml"):
            tree = ET.parse(zf.open(entry))
            root = tree.getroot()
            ns = root.tag.split('}')[0] + '}' if '}' in root.tag else ''
            merges_el = root.find(f'{ns}mergeCells')
            if merges_el is not None:
                sheet_merges = [mc.get('ref') for mc in merges_el.findall(f'{ns}mergeCell')]
                merge_info[entry] = sheet_merges
t_xml_merge_all = time.perf_counter() - t0
print(f"  4a. Merge (all 20 sheets):   {t_xml_merge_all:.4f}s")

# 4b: Single sheet merge info (targeted)
t0 = time.perf_counter()
with zipfile.ZipFile(xlsx_path) as zf:
    tree = ET.parse(zf.open("xl/worksheets/sheet1.xml"))
    root = tree.getroot()
    ns = root.tag.split('}')[0] + '}' if '}' in root.tag else ''
    merges_el = root.find(f'{ns}mergeCells')
    if merges_el is not None:
        single_merges = [mc.get('ref') for mc in merges_el.findall(f'{ns}mergeCell')]
t_xml_merge_one = time.perf_counter() - t0
print(f"  4b. Merge (1 sheet):         {t_xml_merge_one:.4f}s  ({len(single_merges)} ranges)")

# 4c: Comment extraction from XML
t0 = time.perf_counter()
comments_info = {}
with zipfile.ZipFile(xlsx_path) as zf:
    comment_files = [f for f in zf.namelist() if f.startswith("xl/comments") and f.endswith(".xml")]
    for cf in comment_files:
        tree = ET.parse(zf.open(cf))
        root = tree.getroot()
        ns = root.tag.split('}')[0] + '}' if '}' in root.tag else ''
        for comment_el in root.findall(f'.//{ns}comment'):
            ref = comment_el.get('ref')
            texts = comment_el.findall(f'.//{ns}t')
            text = ''.join(t.text or '' for t in texts)
            comments_info[f"{cf}:{ref}"] = text
t_xml_comment = time.perf_counter() - t0
print(f"  4c. Comments (all files):    {t_xml_comment:.4f}s  ({len(comments_info)} comments)")

# 4d: Rels to map comment files to sheets
t0 = time.perf_counter()
sheet_comment_map = {}
with zipfile.ZipFile(xlsx_path) as zf:
    for entry in zf.namelist():
        if entry.startswith("xl/worksheets/_rels/sheet") and entry.endswith(".xml.rels"):
            tree = ET.parse(zf.open(entry))
            root = tree.getroot()
            for rel in root:
                target = rel.get('Target', '')
                if 'comments' in target.lower():
                    sheet_num = re.search(r'sheet(\d+)', entry)
                    if sheet_num:
                        sheet_comment_map[f"sheet{sheet_num.group(1)}"] = target
t_xml_rels = time.perf_counter() - t0
print(f"  4d. Rels mapping:            {t_xml_rels:.4f}s  ({len(sheet_comment_map)} mappings)")

# ============================================================
# Summary
# ============================================================
print("\n" + "=" * 60)
print("=== SUMMARY (per-sheet, 700 rows × 15 cols = 10,500 cells) ===")
print("=" * 60)
print(f"""
Normal mode:
  load_workbook (total):     {t_load_normal:.3f}s
  MergeResolver:             {t_merge:.4f}s
  Cell iter only:            {t_iter_only:.4f}s
  + str(value):              {t_iter_val:.4f}s
  + extract_style:           {t_iter_style:.4f}s  (style adds +{t_iter_style - t_iter_val:.4f}s)
  + comment check:           {t_iter_comment:.4f}s  (comment adds +{t_iter_comment - t_iter_val:.4f}s)
  + merge resolve:           {t_iter_merge:.4f}s  (merge adds +{t_iter_merge - t_iter_val:.4f}s)
  FULL pipeline:             {t_full:.4f}s
  DB bulk insert:            {t_db_insert:.4f}s
  TOTAL per sheet:           {t_merge + t_full + t_db_insert:.4f}s

Read-only mode:
  load_workbook (total):     {t_load_ro:.4f}s
  Cell iter only:            {t_ro_iter:.4f}s
  + extract_style:           {t_ro_style:.4f}s
  comment/merge:             NOT AVAILABLE

XML extraction (all 20 sheets):
  Merge ranges:              {t_xml_merge_all:.4f}s
  Comments:                  {t_xml_comment:.4f}s
  Rels mapping:              {t_xml_rels:.4f}s
  TOTAL XML:                 {t_xml_merge_all + t_xml_comment + t_xml_rels:.4f}s

Estimated parallel import with read_only (20 sheets, 4 workers):
  XML parse (main, 1x):     {t_xml_merge_all + t_xml_comment + t_xml_rels:.4f}s
  Worker load+iter (÷4):    ~{t_ro_style * 20 / 4:.4f}s  (20 sheets / 4 workers)
  DB insert (÷4):           ~{t_db_insert * 20 / 4:.4f}s
  Merge into main DB:       ~{t_db_insert * 20:.4f}s  (sequential)
""")

wb_normal.close()
wb_ro.close()
xlsx_path.unlink()
print("Done!")
