"""Quick test: read_only=True vs normal mode for parallel import feasibility."""
import time
import openpyxl
import tempfile
from pathlib import Path

# --- Generate test xlsx ---
print("=== Generating test xlsx (20 sheets × 700 rows × 15 cols) ===")
wb = openpyxl.Workbook()
wb.remove(wb.active)
for si in range(20):
    ws = wb.create_sheet(title=f"level2_sheet{si}")
    # header
    for c in range(1, 16):
        cell = ws.cell(row=1, column=c, value=f"Header_{c}")
        cell.font = openpyxl.styles.Font(bold=True, color="FF0000")
        cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", fill_type="solid")
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin"),
            bottom=openpyxl.styles.Side(style="medium"),
        )
    # data
    for r in range(2, 702):
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c, value=f"V_{si}_{r}_{c}")
            cell.font = openpyxl.styles.Font(bold=(c == 1))
            if r % 10 == 0:
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", fill_type="solid")
    # merges
    ws.merge_cells("A5:A10")
    ws.merge_cells("B20:C20")

xlsx_path = Path(tempfile.mktemp(suffix=".xlsx"))
wb.save(xlsx_path)
wb.close()
print(f"  Saved to {xlsx_path} ({xlsx_path.stat().st_size / 1024 / 1024:.1f} MB)")

# --- Benchmark: normal vs read_only load ---
print("\n=== 1. load_workbook() speed ===")

t0 = time.perf_counter()
wb_normal = openpyxl.load_workbook(xlsx_path, data_only=False)
t_normal_load = time.perf_counter() - t0
print(f"  Normal mode load:    {t_normal_load:.3f}s")

t0 = time.perf_counter()
wb_ro = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
t_ro_load = time.perf_counter() - t0
print(f"  Read-only mode load: {t_ro_load:.3f}s")
print(f"  Speedup: {t_normal_load / t_ro_load:.1f}x")

# --- Check: style info in read_only mode ---
print("\n=== 2. Style availability in read_only mode ===")
ws_ro = wb_ro["level2_sheet0"]
for row in ws_ro.iter_rows(min_row=1, max_row=1, max_col=3):
    for cell in row:
        print(f"  Cell({cell.row},{cell.column}): value={cell.value}")
        print(f"    font.bold={cell.font.bold}, font.color={cell.font.color}")
        print(f"    fill.fgColor={cell.fill.fgColor}")
        print(f"    border.left={cell.border.left.style if cell.border.left else None}")
        print(f"    number_format={cell.number_format}")
        print(f"    comment={getattr(cell, 'comment', 'NOT AVAILABLE')}")
        break

# --- Check: merged_cells in read_only mode ---
print(f"\n=== 3. Merged cells in read_only mode ===")
try:
    print(f"  ws_ro.merged_cells: {list(ws_ro.merged_cells.ranges)}")
except AttributeError as e:
    print(f"  ws_ro.merged_cells: NOT AVAILABLE ({e})")

ws_normal = wb_normal["level2_sheet0"]
print(f"  ws_normal.merged_cells: {list(ws_normal.merged_cells.ranges)}")

# --- Benchmark: iterate single sheet (normal vs read_only) ---
print("\n=== 4. Single sheet cell iteration speed ===")
from dsm.xlsx_parser import extract_style

# Normal
t0 = time.perf_counter()
ws_n = wb_normal["level2_sheet5"]
count = 0
for row in ws_n.iter_rows(min_row=1, max_row=ws_n.max_row, max_col=ws_n.max_column):
    for cell in row:
        val = str(cell.value) if cell.value is not None else None
        s = extract_style(cell)
        count += 1
t_normal_iter = time.perf_counter() - t0
print(f"  Normal iter ({count} cells):    {t_normal_iter:.3f}s")

# Read-only
t0 = time.perf_counter()
ws_r = wb_ro["level2_sheet5"]
count2 = 0
for row in ws_r.iter_rows(min_row=1, max_row=701, max_col=15):
    for cell in row:
        val = str(cell.value) if cell.value is not None else None
        s = extract_style(cell)
        count2 += 1
t_ro_iter = time.perf_counter() - t0
print(f"  Read-only iter ({count2} cells): {t_ro_iter:.3f}s")

# --- Benchmark: parallel workers with read_only ---
print("\n=== 5. Simulate parallel import: per-worker read_only load + 1 sheet ===")
import multiprocessing as mp

def worker_readonly(args):
    fpath, sheet_name = args
    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=False)
    t_load = time.perf_counter() - t0

    ws = wb[sheet_name]
    count = 0
    t1 = time.perf_counter()
    for row in ws.iter_rows(min_row=1, max_row=701, max_col=15):
        for cell in row:
            val = str(cell.value) if cell.value is not None else None
            # extract_style equivalent (guard None)
            if cell.font:
                _ = cell.font.bold
            if cell.fill:
                _ = cell.fill.fgColor
            if cell.border:
                _ = cell.border
            count += 1
    t_iter = time.perf_counter() - t1
    wb.close()
    return {"sheet": sheet_name, "load": t_load, "iter": t_iter, "cells": count}

def worker_normal(args):
    fpath, sheet_name = args
    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(fpath, data_only=False)
    t_load = time.perf_counter() - t0

    ws = wb[sheet_name]
    count = 0
    t1 = time.perf_counter()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            val = str(cell.value) if cell.value is not None else None
            _ = cell.font.bold
            _ = cell.fill.fgColor
            _ = cell.border
            count += 1
    t_iter = time.perf_counter() - t1
    wb.close()
    return {"sheet": sheet_name, "load": t_load, "iter": t_iter, "cells": count}

sheet_names = [f"level2_sheet{i}" for i in range(20)]
args_list = [(str(xlsx_path), name) for name in sheet_names]

# read_only parallel
t0 = time.perf_counter()
with mp.Pool(4) as pool:
    results_ro = pool.map(worker_readonly, args_list)
t_ro_parallel = time.perf_counter() - t0

avg_load = sum(r["load"] for r in results_ro) / len(results_ro)
avg_iter = sum(r["iter"] for r in results_ro) / len(results_ro)
print(f"  read_only parallel (4 workers, 20 sheets): {t_ro_parallel:.3f}s")
print(f"    avg worker load: {avg_load:.3f}s, avg iter: {avg_iter:.3f}s")

# normal parallel
t0 = time.perf_counter()
with mp.Pool(4) as pool:
    results_n = pool.map(worker_normal, args_list)
t_normal_parallel = time.perf_counter() - t0

avg_load = sum(r["load"] for r in results_n) / len(results_n)
avg_iter = sum(r["iter"] for r in results_n) / len(results_n)
print(f"  normal parallel (4 workers, 20 sheets):    {t_normal_parallel:.3f}s")
print(f"    avg worker load: {avg_load:.3f}s, avg iter: {avg_iter:.3f}s")

print(f"\n  Speedup: {t_normal_parallel / t_ro_parallel:.1f}x")

# --- Bonus: merge info from XML ---
print("\n=== 6. Extract merge info from ZIP XML directly ===")
import zipfile
import xml.etree.ElementTree as ET

t0 = time.perf_counter()
merge_info = {}
with zipfile.ZipFile(xlsx_path) as zf:
    for entry in zf.namelist():
        if entry.startswith("xl/worksheets/sheet") and entry.endswith(".xml"):
            tree = ET.parse(zf.open(entry))
            root = tree.getroot()
            ns = root.tag.split('}')[0] + '}' if '}' in root.tag else ''
            merges = root.find(f'{ns}mergeCells')
            if merges is not None:
                sheet_merges = []
                for mc in merges.findall(f'{ns}mergeCell'):
                    sheet_merges.append(mc.get('ref'))
                merge_info[entry] = sheet_merges
t_xml_merge = time.perf_counter() - t0
print(f"  XML merge extraction: {t_xml_merge:.4f}s")
for k, v in list(merge_info.items())[:3]:
    print(f"    {k}: {v}")

# Cleanup
wb_normal.close()
wb_ro.close()
xlsx_path.unlink()
print("\nDone!")
