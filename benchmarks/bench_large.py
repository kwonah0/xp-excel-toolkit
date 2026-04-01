"""Large-scale benchmark: 20 sheets × ~10K cells each = ~200K cells total."""
import time
import tempfile
import shutil
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment

# ── Generate large xlsx ──────────────────────────────────────────────
# REGMAP headers: TYPE INDX PAGE PARA NAME D7 D6 D5 D4 D3 D2 D1 D0 INIT (+ extra col)
HEADERS = ["TYPE", "INDX", "PAGE", "PARA", "NAME",
           "D7", "D6", "D5", "D4", "D3", "D2", "D1", "D0", "INIT", "Extra"]
NUM_SHEETS = 20
NUM_DATA_ROWS = 667  # per sheet → 667×15 = 10,005 cells per sheet
NUM_IPS_PER_SHEET = 5

print(f"=== Generating large xlsx ({NUM_SHEETS} sheets × {NUM_DATA_ROWS} rows × {len(HEADERS)} cols ≈ {NUM_SHEETS * (NUM_DATA_ROWS + 1) * len(HEADERS):,} cells) ===")
t0 = time.perf_counter()
wb = openpyxl.Workbook()
wb.remove(wb.active)

for si in range(NUM_SHEETS):
    ws = wb.create_sheet(title=f"level2_sheet{si}")
    # header row
    for c, hdr in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=hdr)
        cell.font = openpyxl.styles.Font(bold=True, color="FF0000")
        cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", fill_type="solid")
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin"),
            right=openpyxl.styles.Side(style="thin"),
            top=openpyxl.styles.Side(style="medium"),
            bottom=openpyxl.styles.Side(style="medium"),
        )
    # data rows
    for r in range(2, NUM_DATA_ROWS + 2):
        ip_idx = (r - 2) % NUM_IPS_PER_SHEET
        ip_name = f"IP_{si}_{ip_idx}"
        row_data = [
            "REG",                          # TYPE
            f"0x{r:04X}",                   # INDX
            f"P{r % 4}",                    # PAGE
            f"para_{r}",                    # PARA
            ip_name,                        # NAME
            f"d7_{r}", f"d6_{r}", f"d5_{r}", f"d4_{r}",  # D7-D4
            f"d3_{r}", f"d2_{r}", f"d1_{r}", f"d0_{r}",  # D3-D0
            f"0x{r & 0xFF:02X}",            # INIT
            f"extra_{r}",                   # Extra
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = openpyxl.styles.Font(bold=(c == 1))
            if r % 10 == 0:
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", fill_type="solid")
            if r % 100 == 0 and c == 1:
                cell.comment = Comment(f"Note for row {r} in sheet {si}", "Author")

    # merges (on NAME column)
    for r in range(10, NUM_DATA_ROWS, 50):
        ws.merge_cells(start_row=r, start_column=5, end_row=r + 2, end_column=5)
    ws.merge_cells("B20:C20")

xlsx_path = Path(tempfile.mktemp(suffix=".xlsx"))
wb.save(xlsx_path)
wb.close()
t_gen = time.perf_counter() - t0
size_mb = xlsx_path.stat().st_size / 1024 / 1024
print(f"  Generated in {t_gen:.1f}s — {size_mb:.1f} MB")
print(f"  {NUM_SHEETS} sheets × {NUM_DATA_ROWS + 1} rows × {len(HEADERS)} cols = {NUM_SHEETS * (NUM_DATA_ROWS + 1) * len(HEADERS):,} cells")

# ── Prepare paths ────────────────────────────────────────────────────
db_path = xlsx_path.with_suffix(".db")
output_dir = Path(tempfile.mkdtemp(prefix="dsm_bench_"))

# ── 1. Import (sequential) ──────────────────────────────────────────
print("\n=== 1. dsm import (sequential) ===")
if db_path.exists():
    db_path.unlink()

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from dsm.models import Base, init_db
from dsm.xlsx_parser import import_xlsx

engine = create_engine(f"sqlite:///{db_path}", echo=False)
Base.metadata.create_all(engine)
Session = sessionmaker(bind=engine)

t0 = time.perf_counter()
with Session() as session:
    sheets = import_xlsx(session, xlsx_path)
    session.commit()
t_import_seq = time.perf_counter() - t0
print(f"  Time: {t_import_seq:.2f}s")
print(f"  Sheets imported: {len(sheets)}")
print(f"  DB size: {db_path.stat().st_size / 1024 / 1024:.1f} MB")

# ── 2. Split (sequential) ───────────────────────────────────────────
print("\n=== 2. dsm split (sequential) ===")
from dsm.splitter import split_regmap_from_db

out_seq = output_dir / "seq"
out_seq.mkdir()

t0 = time.perf_counter()
with Session() as session:
    result_seq = split_regmap_from_db(session, xlsx_path.name, out_seq)
t_split_seq = time.perf_counter() - t0
print(f"  Time: {t_split_seq:.2f}s")
print(f"  Files: {len(result_seq)}")

# ── 3. Split --parallel ─────────────────────────────────────────────
print("\n=== 3. dsm split --parallel ===")
from dsm.parallel import parallel_split_regmap

out_par = output_dir / "par"
out_par.mkdir()

t0 = time.perf_counter()
with Session() as session:
    result_par = parallel_split_regmap(session, xlsx_path, out_par)
t_split_par = time.perf_counter() - t0
print(f"  Time: {t_split_par:.2f}s")
print(f"  Files: {len(result_par)}")

# ── 4. Query ─────────────────────────────────────────────────────────
print("\n=== 4. Query (registers) ===")
from dsm.domain_models import Register

t0 = time.perf_counter()
with Session() as session:
    reg_count = session.query(Register).count()
t_query = time.perf_counter() - t0
print(f"  Register count: {reg_count:,}")
print(f"  Query time: {t_query:.3f}s")

# ── Summary ──────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print(f"=== SUMMARY ({NUM_SHEETS} sheets × ~10K cells = ~{NUM_SHEETS * (NUM_DATA_ROWS + 1) * len(HEADERS) // 1000}K cells) ===")
print("=" * 60)
print(f"""
  1. Import:              {t_import_seq:.2f}s
  2. Split (sequential):  {t_split_seq:.2f}s
  3. Split (parallel):    {t_split_par:.2f}s  ({t_split_seq/t_split_par:.1f}x faster)
  4. Query:               {t_query:.3f}s
""")

# Cleanup
xlsx_path.unlink()
db_path.unlink()
shutil.rmtree(output_dir)
print("Done!")
