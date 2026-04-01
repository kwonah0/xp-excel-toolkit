"""Profile patch_merge() to find where time is spent."""
from __future__ import annotations

import io
import tempfile
import time
from pathlib import Path

import openpyxl
from dsm.domain_models import REGMAP_FIELD_MAP, Register, _default_sheet_configs
from dsm.models import ExcelSheet, ExcelWorkbook, init_db
from dsm.patcher import _parse_split_registers, _parse_split_worker, _norm, patch_merge
from dsm.xlsx_parser import import_xlsx
from dsm.splitter import split_regmap_from_db


def create_test_data(tmp: Path, n_sheets=5, n_rows=500, n_cols=15):
    """Create a test xlsx with register data."""
    headers = list(REGMAP_FIELD_MAP.keys())[:n_cols]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for si in range(n_sheets):
        ws = wb.create_sheet(title=f"level2_sheet{si}")
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        for ri in range(2, n_rows + 2):
            for ci, h in enumerate(headers, 1):
                if h == "NAME":
                    ws.cell(row=ri, column=ci, value=f"IP_{ri % 5}")
                else:
                    ws.cell(row=ri, column=ci, value=f"val_{si}_{ri}_{ci}")

    xlsx_path = tmp / "test_large.xlsx"
    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


def main():
    with tempfile.TemporaryDirectory(prefix="dsm_profile_patch_") as tmp_str:
        tmp = Path(tmp_str)
        print("=== Creating test data ===")
        xlsx_path = create_test_data(tmp, n_sheets=20, n_rows=2000)
        db_path = tmp / "test_large.db"
        split_dir = tmp / "split_output"
        split_dir.mkdir()
        patched_path = tmp / "patched.xlsx"

        # 1. Import
        t0 = time.perf_counter()
        Session = init_db(f"sqlite:///{db_path}")
        with Session() as session:
            import_xlsx(session, xlsx_path)
            session.commit()
        t_import = time.perf_counter() - t0
        print(f"Import: {t_import:.2f}s")

        # 2. Split
        t0 = time.perf_counter()
        Session2 = init_db(f"sqlite:///{db_path}")
        with Session2() as session:
            wb_obj = session.query(ExcelWorkbook).first()
            split_regmap_from_db(session, wb_obj.filename, str(split_dir))
        t_split = time.perf_counter() - t0
        print(f"Split: {t_split:.2f}s")

        # 3. Profile patch_merge in detail
        print("\n=== Profiling patch_merge() ===")

        # Phase A: init_db + load blob
        t0 = time.perf_counter()
        SessionP = init_db(f"sqlite:///{db_path}")
        with SessionP() as session:
            wb_obj = session.query(ExcelWorkbook).first()
            blob = wb_obj.blob
        t_db_load = time.perf_counter() - t0
        print(f"  A. DB open + load blob:     {t_db_load:.3f}s  (blob size: {len(blob)/1024/1024:.1f} MB)")

        # Phase B: openpyxl.load_workbook from blob
        t0 = time.perf_counter()
        wb = openpyxl.load_workbook(io.BytesIO(blob))
        t_wb_load = time.perf_counter() - t0
        print(f"  B. openpyxl load from blob: {t_wb_load:.3f}s  (sheets: {len(wb.sheetnames)})")

        # Phase C: Build original register index
        t0 = time.perf_counter()
        SessionP2 = init_db(f"sqlite:///{db_path}")
        with SessionP2() as session:
            wb_obj = session.query(ExcelWorkbook).first()
            original_regs = {}
            level2_sheets = (
                session.query(ExcelSheet)
                .filter(ExcelSheet.workbook_id == wb_obj.id, ExcelSheet.name.like("level2_%"))
                .all()
            )
            for sheet_obj in level2_sheets:
                regs = session.query(Register).filter_by(sheet_id=sheet_obj.id).all()
                for reg in regs:
                    key = (sheet_obj.name, reg.name, reg.indx, reg.page, reg.para)
                    original_regs[key] = reg
        t_reg_index = time.perf_counter() - t0
        print(f"  C. Build register index:    {t_reg_index:.3f}s  (regs: {len(original_regs)})")

        # Phase D: Parse split files (sequential)
        split_files = sorted(split_dir.glob("*.xlsx"))
        t0 = time.perf_counter()
        all_split_regs = {}
        for sf in split_files:
            all_split_regs[sf.stem] = _parse_split_registers(sf)
        t_parse_seq = time.perf_counter() - t0
        total_split_regs = sum(len(v) for v in all_split_regs.values())
        print(f"  D1. Parse split (seq):      {t_parse_seq:.3f}s  (files: {len(split_files)}, regs: {total_split_regs})")

        # Phase D2: Parse split files (parallel)
        import multiprocessing as _mp
        t0 = time.perf_counter()
        num_w = min(_mp.cpu_count(), len(split_files))
        with _mp.Pool(processes=num_w) as pool:
            par_results = pool.map(_parse_split_worker, [str(f) for f in split_files])
        all_split_regs_par = {r["stem"]: r["regs"] for r in par_results if r["success"]}
        t_parse_par = time.perf_counter() - t0
        print(f"  D2. Parse split (parallel): {t_parse_par:.3f}s  ({t_parse_seq/t_parse_par:.1f}x speedup)")

        # Phase E: Compare + patch cells (without save)
        t0 = time.perf_counter()
        # Just simulate the comparison loop
        n_comparisons = 0
        for sheet_name, regs in all_split_regs.items():
            for reg_data in regs:
                n_comparisons += 1
        t_compare = time.perf_counter() - t0
        print(f"  E. Compare loop (no save):  {t_compare:.3f}s  (comparisons: {n_comparisons})")

        # Phase F: wb.save()
        t0 = time.perf_counter()
        wb.save(patched_path)
        t_save = time.perf_counter() - t0
        file_size = patched_path.stat().st_size / 1024 / 1024
        print(f"  F. wb.save():               {t_save:.3f}s  (file: {file_size:.1f} MB)")

        wb.close()

        # Phase G: Full patch_merge end-to-end
        patched_path2 = tmp / "patched2.xlsx"
        t0 = time.perf_counter()
        result = patch_merge(db_path, split_dir, patched_path2)
        t_full = time.perf_counter() - t0
        print(f"\n  TOTAL patch_merge():        {t_full:.3f}s  (changes: {len(result.changes)})")

        print("\n=== Summary ===")
        print(f"  Import:       {t_import:.2f}s")
        print(f"  Split:        {t_split:.2f}s")
        print(f"  Patch merge:  {t_full:.2f}s")


if __name__ == "__main__":
    main()
