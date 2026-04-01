"""Profile all DSM scenarios with synthetic large data.

Simulates: 20 sheets, 10,000+ cells per sheet (700+ rows × 15 cols).
Measures wall-clock time for each phase.
"""

import time
import shutil
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment


# ── Synthetic data generator ────────────────────────────────────

def generate_large_xlsx(path: Path, n_sheets: int = 20, n_rows: int = 700):
    """Create an xlsx with n_sheets level2-like sheets + 1 memorymap sheet.

    Each level2 sheet has n_rows data rows × 15 columns ≈ 10,500 cells.
    Total cells ≈ n_sheets × n_rows × 15 = 210,000 cells.
    """
    wb = openpyxl.Workbook()
    ws_default = wb.active

    # Header for level2 sheets
    headers = ["TYPE", "INDX", "PAGE", "PARA", "NAME",
               "D7", "D6", "D5", "D4", "D3", "D2", "D1", "D0", "INIT", "EXTRA"]

    ips_per_sheet = 3  # 3 distinct IPs per sheet

    for si in range(n_sheets):
        ws = wb.create_sheet(title=f"level2_block{si}")
        # Header row at row 1
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)

        # Data rows
        for ri in range(2, n_rows + 2):
            ip_idx = (ri - 2) % ips_per_sheet
            ip_name = f"IP_{si}_{ip_idx}"
            ws.cell(row=ri, column=1, value="RW2")
            ws.cell(row=ri, column=2, value=f"{si:02X}")
            ws.cell(row=ri, column=3, value=str((ri - 2) // 256))
            ws.cell(row=ri, column=4, value=str((ri - 2) % 256))
            ws.cell(row=ri, column=5, value=ip_name)
            for di in range(6, 14):
                ws.cell(row=ri, column=di, value=f"BIT{di - 6}")
            ws.cell(row=ri, column=14, value="0x00")
            ws.cell(row=ri, column=15, value=f"extra_{ri}")
            # Add a comment to every 100th cell
            if ri % 100 == 0:
                ws.cell(row=ri, column=5).comment = Comment(f"Note for row {ri}", "Author")

        # Add some merged cells (every 50 rows merge col 15 across 2 rows)
        for ri in range(2, n_rows, 50):
            if ri + 1 <= n_rows + 1:
                ws.merge_cells(start_row=ri, start_column=15, end_row=ri + 1, end_column=15)

    # Memorymap sheet
    ws_mm = wb.create_sheet(title="memorymap")
    mm_headers = ["BASEADDR", "Group", "midgroup", "Comment", "special"]
    for ci, h in enumerate(mm_headers, 1):
        ws_mm.cell(row=1, column=ci, value=h)
    for ri in range(2, 102):  # 100 memmap entries
        ws_mm.cell(row=ri, column=1, value=f"0x{ri * 2:02X}")
        ws_mm.cell(row=ri, column=2, value=f"GROUP_{ri}")
        ws_mm.cell(row=ri, column=3, value=f"MID_{ri}")
        ws_mm.cell(row=ri, column=4, value=f"Comment for entry {ri}")
        ws_mm.cell(row=ri, column=5, value="")

    # Remove default empty sheet
    wb.remove(ws_default)
    wb.save(path)
    wb.close()

    total_cells = n_sheets * n_rows * 15 + 100 * 5
    print(f"Generated: {path}")
    print(f"  {n_sheets + 1} sheets, ~{n_rows} rows/sheet, ~{total_cells:,} total cells")
    return path


# ── Profiling helpers ──────────────────────────────────────────

class Timer:
    def __init__(self, label: str):
        self.label = label
        self.elapsed = 0.0

    def __enter__(self):
        self._t0 = time.perf_counter()
        return self

    def __exit__(self, *args):
        self.elapsed = time.perf_counter() - self._t0

    def __str__(self):
        return f"{self.label:.<45} {self.elapsed:7.2f}s"


def profile_import(xlsx_path: Path, db_path: Path):
    """Profile: dsm import (sequential)."""
    from dsm.models import init_db
    from dsm.xlsx_parser import import_xlsx

    print("\n=== Scenario 1: import (sequential) ===")

    with Timer("Total import") as t_total:
        Session = init_db(f"sqlite:///{db_path}")
        with Session() as session:
            with Timer("  import_xlsx()") as t_import:
                sheets = import_xlsx(session, xlsx_path)
            with Timer("  session.commit()") as t_commit:
                session.commit()
            sheet_info = [(s.name, s.header_row) for s in sheets]

    print(t_import)
    print(t_commit)
    print(t_total)
    print(f"  Imported {len(sheet_info)} sheets")
    return db_path


def profile_split(xlsx_path: Path, db_path: Path, output_dir: Path):
    """Profile: dsm split (from existing DB)."""
    from dsm.models import init_db
    from dsm.splitter import split_regmap_from_db

    print("\n=== Scenario 2: split (sequential, from DB) ===")

    with Timer("Total split") as t_total:
        Session = init_db(f"sqlite:///{db_path}")
        with Session() as session:
            with Timer("  split_regmap_from_db()") as t_split:
                results = split_regmap_from_db(session, xlsx_path.name, output_dir)
            with Timer("  session.commit()") as t_commit:
                session.commit()

    print(t_split)
    print(t_commit)
    print(t_total)
    print(f"  Split into {len(results)} files")


def profile_split_parallel(xlsx_path: Path, db_path: Path, output_dir: Path, workers: int = 4):
    """Profile: dsm split --parallel."""
    from dsm.models import init_db
    from dsm.parallel import parallel_split_regmap

    print(f"\n=== Scenario 3: split --parallel (workers={workers}) ===")

    with Timer("Total parallel split") as t_total:
        Session = init_db(f"sqlite:///{db_path}")
        with Session() as session:
            with Timer("  parallel_split_regmap()") as t_split:
                results = parallel_split_regmap(session, xlsx_path, output_dir, workers=workers)
            with Timer("  session.commit()") as t_commit:
                session.commit()

    print(t_split)
    print(t_commit)
    print(t_total)
    print(f"  Split into {len(results)} files")


def profile_diff(db_path_a: Path, db_path_b: Path, diff_db_path: Path):
    """Profile: dsm diff."""
    from dsm.diff import diff_databases, save_diff_to_db, format_diff

    print("\n=== Scenario 4: diff ===")

    with Timer("Total diff") as t_total:
        with Timer("  diff_databases()") as t_diff:
            result = diff_databases(db_path_a, db_path_b)
        with Timer("  save_diff_to_db()") as t_save:
            save_diff_to_db(result, diff_db_path, db_path_a, db_path_b)
        with Timer("  format_diff()") as t_fmt:
            text = format_diff(result)

    print(t_diff)
    print(t_save)
    print(t_fmt)
    print(t_total)
    n = (len(result.added_regs) + len(result.removed_regs) +
         len(result.changed_regs) + len(result.added_memmap) +
         len(result.removed_memmap) + len(result.changed_memmap))
    print(f"  {n} total differences")


def profile_query(db_path: Path):
    """Profile: dsm query operations."""
    from dsm.models import init_db, ExcelSheet
    from dsm.domain_models import Register, MemoryMapEntry
    from sqlalchemy import func, distinct

    print("\n=== Scenario 5: query ===")

    Session = init_db(f"sqlite:///{db_path}")

    with Session() as session:
        with Timer("  query all sheets") as t1:
            sheets = session.query(ExcelSheet).all()
            _ = [(s.name, s.header_row) for s in sheets]

        with Timer("  query all registers") as t2:
            regs = session.query(Register).all()

        with Timer("  query distinct IPs") as t3:
            ips = session.query(distinct(Register.name)).all()

        with Timer("  query registers with filter") as t4:
            filtered = (session.query(Register)
                        .filter(Register.name.like("IP_0_%"))
                        .order_by(Register.excel_row)
                        .all())

        with Timer("  query memorymap") as t5:
            mm = session.query(MemoryMapEntry).all()

    print(t1)
    print(t2)
    print(t3)
    print(t4)
    print(t5)
    print(f"  {len(sheets)} sheets, {len(regs)} registers, "
          f"{len(ips)} IPs, {len(filtered)} filtered, {len(mm)} memmap")


# ── Main ───────────────────────────────────────────────────────

def main():
    tmpdir = Path(tempfile.mkdtemp(prefix="dsm_profile_"))
    print(f"Working directory: {tmpdir}")

    try:
        # Generate test data
        xlsx_path = tmpdir / "large_regmap.xlsx"
        with Timer("Generate xlsx") as t_gen:
            generate_large_xlsx(xlsx_path, n_sheets=20, n_rows=700)
        print(t_gen)

        # 1. Import
        db_path = tmpdir / "main.db"
        profile_import(xlsx_path, db_path)

        # 2. Split (sequential from DB)
        split_dir_seq = tmpdir / "split_seq"
        profile_split(xlsx_path, db_path, split_dir_seq)

        # 3. Split (parallel)
        split_dir_par = tmpdir / "split_par"
        profile_split_parallel(xlsx_path, db_path, split_dir_par, workers=4)

        # 4. Query
        profile_query(db_path)

        # File sizes
        print("\n=== File Sizes ===")
        for f in sorted(tmpdir.glob("*.db")):
            size_mb = f.stat().st_size / 1024 / 1024
            print(f"  {f.name:.<40} {size_mb:.1f} MB")
        print(f"  {xlsx_path.name:.<40} {xlsx_path.stat().st_size / 1024 / 1024:.1f} MB")

    finally:
        print(f"\nCleanup: {tmpdir}")
        shutil.rmtree(tmpdir, ignore_errors=True)


if __name__ == "__main__":
    main()
