"""openpyxl-based parser for .xlsx files."""

from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models import (
    COLUMN_MAP, Employee, ExcelCell, ExcelMerge, ExcelSheet, ExcelWorkbook,
)


def extract_style(cell: Cell) -> dict | None:
    """Extract visual style info from an openpyxl cell."""
    style: dict = {}

    # Background color
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
        rgb = fill.fgColor.rgb
        if isinstance(rgb, str) and len(rgb) == 8:
            style["bg_color"] = f"#{rgb[2:]}"  # strip alpha

    # Font
    font = cell.font
    if font:
        if font.bold:
            style["font_bold"] = True
        if font.color and font.color.rgb and font.color.rgb != "00000000":
            rgb = font.color.rgb
            if isinstance(rgb, str) and len(rgb) == 8:
                style["font_color"] = f"#{rgb[2:]}"

    # Number format
    if cell.number_format and cell.number_format != "General":
        style["number_format"] = cell.number_format

    # Borders
    border = cell.border
    if border:
        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(border, side_name)
            if side and side.style:
                style[f"border_{side_name}"] = side.style

    return style if style else None


def _build_merge_map(ws) -> dict[tuple[int, int], tuple[int, int]]:
    """Build mapping: (row, col) → (origin_row, origin_col) for merged cells."""
    merge_map: dict[tuple[int, int], tuple[int, int]] = {}
    for mr in ws.merged_cells.ranges:
        origin = (mr.min_row, mr.min_col)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_map[(r, c)] = origin
    return merge_map


def find_header_row(ws, min_cols: int = 3) -> int | None:
    """Auto-detect header row: first row where ≥ min_cols cells have string values."""
    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row or 30)):
        str_count = sum(
            1 for cell in row
            if cell.value is not None and isinstance(cell.value, str) and cell.value.strip()
        )
        if str_count >= min_cols:
            return row[0].row
    return None


def import_xlsx(
    session: Session,
    path: str | Path,
    sheet_name: str | None = None,
    header_row: int | None = None,
    field_map: dict[str, str] | None = None,
) -> ExcelSheet:
    """
    Import a .xlsx file into the DB.

    Args:
        session: SQLAlchemy session
        path: Path to .xlsx file
        sheet_name: Target sheet name (default: first sheet)
        header_row: 1-based header row index (auto-detected if None)
        field_map: Mapping of Excel column header → Employee field name
                   e.g. {"이름": "name", "부서": "department", ...}

    Returns:
        The created ExcelSheet ORM object.
    """
    path = Path(path)

    # Store original binary for round-trip
    blob = path.read_bytes()

    wb_xl = openpyxl.load_workbook(path, data_only=False)
    ws = wb_xl[sheet_name] if sheet_name else wb_xl.active

    # Create workbook record
    wb_obj = ExcelWorkbook(filename=path.name, blob=blob)
    session.add(wb_obj)
    session.flush()

    # Create sheet record
    sheet_obj = ExcelSheet(workbook_id=wb_obj.id, name=ws.title)
    session.add(sheet_obj)
    session.flush()

    # ── Merge ranges ────────────────────────────────────────────
    merge_map = _build_merge_map(ws)
    merge_db: dict[str, ExcelMerge] = {}  # "min_row:min_col" → ExcelMerge

    for mr in ws.merged_cells.ranges:
        m = ExcelMerge(
            sheet_id=sheet_obj.id,
            min_row=mr.min_row,
            min_col=mr.min_col,
            max_row=mr.max_row,
            max_col=mr.max_col,
        )
        session.add(m)
        session.flush()
        merge_db[f"{mr.min_row}:{mr.min_col}"] = m

    # ── Cells ───────────────────────────────────────────────────
    # Pre-read origin values for merged cells
    origin_values: dict[tuple[int, int], str] = {}
    for (r, c), (or_, oc) in merge_map.items():
        if (or_, oc) not in origin_values:
            origin_cell = ws.cell(row=or_, column=oc)
            origin_values[(or_, oc)] = (
                str(origin_cell.value) if origin_cell.value is not None else None
            )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            r, c = cell.row, cell.column

            # Determine value (fill merged cells with origin value)
            if (r, c) in merge_map:
                origin = merge_map[(r, c)]
                raw_val = origin_values.get(origin)
            else:
                raw_val = str(cell.value) if cell.value is not None else None

            # Merge linkage
            merge_key = None
            is_origin = False
            if (r, c) in merge_map:
                or_, oc = merge_map[(r, c)]
                key = f"{or_}:{oc}"
                merge_key = merge_db[key].id
                is_origin = (r == or_ and c == oc)

            # Style (only from the actual cell object, not merged placeholders)
            style = extract_style(cell) if cell.value is not None or is_origin else None

            cell_obj = ExcelCell(
                sheet_id=sheet_obj.id,
                row=r,
                col=c,
                raw_value=raw_val,
                style=style,
                merge_id=merge_key,
                is_merge_origin=is_origin,
            )
            session.add(cell_obj)

    session.flush()

    # ── Header detection + domain object creation ───────────────
    if header_row is None:
        header_row = find_header_row(ws)
    sheet_obj.header_row = header_row

    if header_row and field_map:
        # Build column index mapping from header row
        headers: dict[int, str] = {}
        for cell in ws[header_row]:
            if cell.value and isinstance(cell.value, str):
                headers[cell.column] = cell.value.strip()

        # Map Excel col index → domain field
        col_to_field: dict[int, str] = {}
        for col_idx, header_text in headers.items():
            if header_text in field_map:
                domain_field = field_map[header_text]
                col_to_field[col_idx] = domain_field
                COLUMN_MAP[domain_field] = col_idx

        # Create Employee records for data rows
        for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
            row_data: dict[str, object] = {}
            has_data = False
            for col_idx, field_name in col_to_field.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                # Fill from merge origin if needed
                if val is None and (row_idx, col_idx) in merge_map:
                    or_, oc = merge_map[(row_idx, col_idx)]
                    val = ws.cell(row=or_, column=oc).value
                if val is not None:
                    has_data = True
                row_data[field_name] = val

            if has_data:
                emp = Employee(
                    sheet_id=sheet_obj.id,
                    excel_row=row_idx,
                    **row_data,
                )
                session.add(emp)

    session.flush()
    wb_xl.close()
    return sheet_obj
