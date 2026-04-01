"""Benchmark: Full import pipeline — ORM per-row vs Core bulk insert.

Compares the ENTIRE _import_ws() pipeline (ExcelMerge + ExcelCell + Domain objects)
between the current ORM approach and Core bulk insert.

Target: ~100K cells (10K rows × 14 cols) to measure real-world impact.

Usage:
    uv run python tests/bench_full_pipeline.py
    uv run python tests/bench_full_pipeline.py --rows 10000
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path
from typing import Any

# Project setup
PROJECT_ROOT = Path(__file__).resolve().parent.parent
os.chdir(PROJECT_ROOT)
sys.path.insert(0, str(PROJECT_ROOT / "src"))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, insert, text
from sqlalchemy.orm import Session, sessionmaker

from dsm.models import Base, ExcelCell, ExcelMerge, ExcelSheet, ExcelWorkbook
from dsm.domain_models import REGMAP_FIELD_MAP, Register
from dsm.merge import MergeResolver
from dsm.xlsx_parser import extract_style, find_header_row


# ── 1. Generate large sample xlsx ─────────────────────────────────────

HEADERS = ["TYPE", "INDX", "PAGE", "PARA", "NAME",
           "D7", "D6", "D5", "D4", "D3", "D2", "D1", "D0", "INIT"]

REGISTER_TEMPLATES = [
    ("RW2", "EN", "MODE[1:0]", "RST", "RSVD", "RSVD", "RSVD", "RSVD", "RSVD", "0x00"),
    ("RO",  "BUSY", "DONE", "ERR", "CNT[4:0]", "CNT[4:0]", "CNT[4:0]", "CNT[4:0]", "CNT[4:0]", "0x00"),
    ("RW1", "GAIN[1:0]", "GAIN[1:0]", "DIV[2:0]", "DIV[2:0]", "DIV[2:0]", "POL", "PHASE", "LOCK", "0x24"),
    ("RO",  "DATA[7:0]", "DATA[7:0]", "DATA[7:0]", "DATA[7:0]", "DATA[7:0]", "DATA[7:0]", "DATA[7:0]", "DATA[7:0]", "0xFF"),
]

IP_NAMES = [
    "SENSOR_A", "AMPLIFIER", "DAC_CTRL", "ADC_CONV", "PLL_CFG",
    "GPIO_PORT", "TIMER_A", "SPI_MASTER", "I2C_SLAVE", "PWR_MGMT",
    "SENSOR_B", "SENSOR_C", "DMA_CH0", "DMA_CH1", "UART_0",
    "UART_1", "WATCHDOG", "RTC_TIMER", "CRC_GEN", "FLASH_CTRL",
]


def generate_large_xlsx(path: Path, num_rows: int) -> Path:
    """Generate a large register map xlsx for benchmarking."""
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "level2_bench"

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    header_font = Font(name="Consolas", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    data_font = Font(name="Consolas", size=10)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = center

    indx_counter = 10
    row = 2
    ip_idx = 0
    merge_ranges = []

    while row <= num_rows + 1:
        ip_name = IP_NAMES[ip_idx % len(IP_NAMES)]
        ip_idx += 1
        indx = str(indx_counter)
        indx_counter += 1

        group_start = row
        for para, tmpl in enumerate(REGISTER_TEMPLATES):
            if row > num_rows + 1:
                break
            type_, d7, d6, d5, d4, d3, d2, d1, d0, init = tmpl

            ws.cell(row=row, column=1, value=type_).font = data_font
            ws.cell(row=row, column=2, value=indx).font = data_font
            ws.cell(row=row, column=3, value="0").font = data_font
            ws.cell(row=row, column=4, value=str(para)).font = data_font
            ws.cell(row=row, column=5, value=ip_name).font = data_font
            ws.cell(row=row, column=6, value=d7).font = data_font
            ws.cell(row=row, column=7, value=d6).font = data_font
            ws.cell(row=row, column=8, value=d5).font = data_font
            ws.cell(row=row, column=9, value=d4).font = data_font
            ws.cell(row=row, column=10, value=d3).font = data_font
            ws.cell(row=row, column=11, value=d2).font = data_font
            ws.cell(row=row, column=12, value=d1).font = data_font
            ws.cell(row=row, column=13, value=d0).font = data_font
            ws.cell(row=row, column=14, value=init).font = data_font

            for col in range(1, 15):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = center
            if para == 0:
                ws.cell(row=row, column=6).fill = green_fill

            row += 1

        group_end = row - 1
        if group_end > group_start:
            merge_ranges.append((group_start, group_end, 2))
            merge_ranges.append((group_start, group_end, 5))

    for min_row, max_row, col in merge_ranges:
        col_letter = get_column_letter(col)
        ws.merge_cells(f"{col_letter}{min_row}:{col_letter}{max_row}")

    ws.freeze_panes = "A2"
    wb.save(path)
    wb.close()
    return path


# ── 2. Strategy A: Current ORM approach (exact copy of _import_ws) ────

def import_ws_orm(
    session: Session, ws, wb_id: int,
    field_map: dict[str, str], domain_cls: type,
) -> ExcelSheet:
    """Current _import_ws() — ORM objects + per-row flush."""
    sheet_obj = ExcelSheet(workbook_id=wb_id, name=ws.title)
    session.add(sheet_obj)
    session.flush()

    # -- Merge ranges (per-merge flush)
    merger = MergeResolver(ws)
    merge_db: dict[str, ExcelMerge] = {}

    for mr in merger.ranges:
        m = ExcelMerge(
            sheet_id=sheet_obj.id,
            min_row=mr.min_row, min_col=mr.min_col,
            max_row=mr.max_row, max_col=mr.max_col,
        )
        session.add(m)
        session.flush()
        merge_db[f"{mr.min_row}:{mr.min_col}"] = m

    # -- Cells (per-cell add, single flush)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            r, c = cell.row, cell.column

            if merger.is_merged(r, c):
                raw_val = merger.get_value(r, c)
            else:
                raw_val = str(cell.value) if cell.value is not None else None

            merge_key = None
            is_origin = merger.is_origin(r, c)
            origin = merger.get_origin(r, c)
            if origin:
                key = f"{origin[0]}:{origin[1]}"
                merge_key = merge_db[key].id

            style = extract_style(cell) if cell.value is not None or is_origin else None

            session.add(ExcelCell(
                sheet_id=sheet_obj.id, row=r, col=c,
                raw_value=raw_val, style=style,
                merge_id=merge_key, is_merge_origin=is_origin,
            ))

    session.flush()

    # -- Header detection + domain object creation (per-row flush)
    header_row = find_header_row(ws, field_map=field_map)
    sheet_obj.header_row = header_row

    if header_row and field_map and domain_cls:
        headers: dict[int, str] = {}
        for cell in ws[header_row]:
            if cell.value and isinstance(cell.value, str):
                headers[cell.column] = cell.value.strip()

        col_to_field: dict[int, str] = {}
        for col_idx, header_text in headers.items():
            if header_text in field_map:
                col_to_field[col_idx] = field_map[header_text]

        for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
            row_data: dict[str, Any] = {}
            has_data = False
            for col_idx, field_name in col_to_field.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is None and merger.is_merged(row_idx, col_idx):
                    val = merger.get_value(row_idx, col_idx)
                if val is not None:
                    has_data = True
                    row_data[field_name] = str(val)
                else:
                    row_data[field_name] = None

            if has_data:
                obj = domain_cls(
                    sheet_id=sheet_obj.id,
                    excel_row=row_idx,
                    **row_data,
                )
                session.add(obj)
                session.flush()

    session.flush()
    return sheet_obj


# ── 3. Strategy B: Core bulk insert for everything ────────────────────

def import_ws_bulk(
    session: Session, ws, wb_id: int,
    field_map: dict[str, str], domain_cls: type,
) -> ExcelSheet:
    """Optimized _import_ws() — Core bulk insert for merges, cells, and domain objects."""
    sheet_obj = ExcelSheet(workbook_id=wb_id, name=ws.title)
    session.add(sheet_obj)
    session.flush()

    sid = sheet_obj.id

    # -- Merge ranges: bulk insert + query back for IDs ─────────────────
    merger = MergeResolver(ws)

    merge_dicts = [
        {
            "sheet_id": sid,
            "min_row": mr.min_row,
            "min_col": mr.min_col,
            "max_row": mr.max_row,
            "max_col": mr.max_col,
        }
        for mr in merger.ranges
    ]

    # Build merge lookup: "min_row:min_col" -> merge_id
    merge_id_map: dict[str, int] = {}
    if merge_dicts:
        session.execute(insert(ExcelMerge), merge_dicts)
        session.flush()

        # Query back IDs
        rows = (
            session.query(ExcelMerge.id, ExcelMerge.min_row, ExcelMerge.min_col)
            .filter(ExcelMerge.sheet_id == sid)
            .all()
        )
        for mid, mrow, mcol in rows:
            merge_id_map[f"{mrow}:{mcol}"] = mid

    # -- Cells: collect dicts, bulk insert ──────────────────────────────
    cell_dicts: list[dict] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            r, c = cell.row, cell.column

            if merger.is_merged(r, c):
                raw_val = merger.get_value(r, c)
            else:
                raw_val = str(cell.value) if cell.value is not None else None

            merge_key = None
            is_origin = merger.is_origin(r, c)
            origin = merger.get_origin(r, c)
            if origin:
                key = f"{origin[0]}:{origin[1]}"
                merge_key = merge_id_map.get(key)

            style = extract_style(cell) if cell.value is not None or is_origin else None

            cell_dicts.append({
                "sheet_id": sid,
                "row": r,
                "col": c,
                "raw_value": raw_val,
                "style": style,
                "merge_id": merge_key,
                "is_merge_origin": is_origin,
            })

    if cell_dicts:
        # Chunk to avoid SQLite variable limit (~999 vars)
        CHUNK = 500
        for i in range(0, len(cell_dicts), CHUNK):
            session.execute(insert(ExcelCell), cell_dicts[i:i + CHUNK])
        session.flush()

    # -- Header detection + domain objects: bulk insert ─────────────────
    header_row = find_header_row(ws, field_map=field_map)
    sheet_obj.header_row = header_row

    if header_row and field_map and domain_cls:
        headers: dict[int, str] = {}
        for cell in ws[header_row]:
            if cell.value and isinstance(cell.value, str):
                headers[cell.column] = cell.value.strip()

        col_to_field: dict[int, str] = {}
        for col_idx, header_text in headers.items():
            if header_text in field_map:
                col_to_field[col_idx] = field_map[header_text]

        bulk_rows: list[dict[str, Any]] = []
        for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
            row_data: dict[str, Any] = {}
            has_data = False
            for col_idx, field_name in col_to_field.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is None and merger.is_merged(row_idx, col_idx):
                    val = merger.get_value(row_idx, col_idx)
                if val is not None:
                    has_data = True
                    row_data[field_name] = str(val)
                else:
                    row_data[field_name] = None

            if has_data:
                row_data["sheet_id"] = sid
                row_data["excel_row"] = row_idx
                bulk_rows.append(row_data)

        if bulk_rows:
            CHUNK = 500
            for i in range(0, len(bulk_rows), CHUNK):
                session.execute(insert(domain_cls), bulk_rows[i:i + CHUNK])
            session.flush()

    session.flush()
    return sheet_obj


# ── 4. Benchmark runner ───────────────────────────────────────────────

def run_benchmark(num_rows: int):
    bench_xlsx = Path("samples/bench_large.xlsx")
    print(f"=== Full Pipeline Benchmark ({num_rows} rows × 14 cols = {num_rows * 14:,} cells) ===\n")

    # Generate test data
    print(f"Generating {num_rows}-row xlsx...")
    t0 = time.perf_counter()
    generate_large_xlsx(bench_xlsx, num_rows)
    gen_time = time.perf_counter() - t0
    print(f"  Generated in {gen_time:.2f}s ({bench_xlsx.stat().st_size / 1024:.0f} KB)\n")

    # Load workbook once
    print(f"Loading workbook with openpyxl...")
    t0 = time.perf_counter()
    wb_xl = openpyxl.load_workbook(bench_xlsx, data_only=False)
    ws = wb_xl["level2_bench"]
    load_time = time.perf_counter() - t0
    print(f"  Loaded in {load_time:.2f}s")
    print(f"  Dimensions: {ws.max_row} rows × {ws.max_column} cols")
    print(f"  Merge ranges: {len(list(ws.merged_cells.ranges))}")

    # ── Strategy A: ORM ──────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  Strategy A: ORM (current _import_ws)")
    print(f"{'='*60}")

    db_a = Path("samples/bench_pipeline_orm.db")
    db_a.unlink(missing_ok=True)
    engine_a = create_engine(f"sqlite:///{db_a}", echo=False)
    Base.metadata.create_all(engine_a)
    SessionA = sessionmaker(bind=engine_a)

    with SessionA() as session:
        wb_obj = ExcelWorkbook(filename=bench_xlsx.name, blob=None)
        session.add(wb_obj)
        session.flush()

        t0 = time.perf_counter()
        sheet_a = import_ws_orm(session, ws, wb_obj.id, REGMAP_FIELD_MAP, Register)
        session.commit()
        elapsed_a = time.perf_counter() - t0

    print(f"  Total: {elapsed_a:.3f}s")

    # ── Strategy B: Core bulk ────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  Strategy B: Core bulk insert (all three: merge + cell + domain)")
    print(f"{'='*60}")

    db_b = Path("samples/bench_pipeline_bulk.db")
    db_b.unlink(missing_ok=True)
    engine_b = create_engine(f"sqlite:///{db_b}", echo=False)
    Base.metadata.create_all(engine_b)
    SessionB = sessionmaker(bind=engine_b)

    with SessionB() as session:
        wb_obj = ExcelWorkbook(filename=bench_xlsx.name, blob=None)
        session.add(wb_obj)
        session.flush()

        t0 = time.perf_counter()
        sheet_b = import_ws_bulk(session, ws, wb_obj.id, REGMAP_FIELD_MAP, Register)
        session.commit()
        elapsed_b = time.perf_counter() - t0

    print(f"  Total: {elapsed_b:.3f}s")

    # ── Verify correctness ─────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  Verification")
    print(f"{'='*60}")

    with SessionA() as sa, SessionB() as sb:
        # Check merge counts
        merges_a = sa.query(ExcelMerge).count()
        merges_b = sb.query(ExcelMerge).count()
        print(f"  ExcelMerge  : {merges_a} vs {merges_b} {'OK' if merges_a == merges_b else 'MISMATCH!'}")

        # Check cell counts
        cells_a = sa.query(ExcelCell).count()
        cells_b = sb.query(ExcelCell).count()
        print(f"  ExcelCell   : {cells_a:,} vs {cells_b:,} {'OK' if cells_a == cells_b else 'MISMATCH!'}")

        # Check register counts
        regs_a_count = sa.query(Register).count()
        regs_b_count = sb.query(Register).count()
        print(f"  Register    : {regs_a_count:,} vs {regs_b_count:,} {'OK' if regs_a_count == regs_b_count else 'MISMATCH!'}")

        # Spot-check register values (first 100 + last 100)
        regs_a = sa.query(Register).order_by(Register.excel_row).all()
        regs_b = sb.query(Register).order_by(Register.excel_row).all()

        check_indices = list(range(min(100, len(regs_a)))) + list(range(max(0, len(regs_a) - 100), len(regs_a)))
        check_indices = sorted(set(check_indices))

        mismatches = 0
        fields = ["type", "indx", "page", "para", "name",
                  "d7", "d6", "d5", "d4", "d3", "d2", "d1", "d0",
                  "init", "excel_row"]
        for i in check_indices:
            ra, rb = regs_a[i], regs_b[i]
            for f in fields:
                va, vb = getattr(ra, f), getattr(rb, f)
                if va != vb:
                    mismatches += 1
                    if mismatches <= 5:
                        print(f"  MISMATCH row={ra.excel_row} field={f}: {va!r} vs {vb!r}")

        if mismatches == 0:
            print(f"  Field values: spot-checked {len(check_indices)} registers — all match OK")
        else:
            print(f"  {mismatches} field mismatches!")

        # Spot-check cell values (merge linkage)
        cells_a_sample = (
            sa.query(ExcelCell)
            .filter(ExcelCell.merge_id.isnot(None))
            .order_by(ExcelCell.row, ExcelCell.col)
            .limit(50)
            .all()
        )
        cells_b_sample = (
            sb.query(ExcelCell)
            .filter(ExcelCell.merge_id.isnot(None))
            .order_by(ExcelCell.row, ExcelCell.col)
            .limit(50)
            .all()
        )

        cell_mismatches = 0
        for ca, cb in zip(cells_a_sample, cells_b_sample):
            if (ca.row, ca.col, ca.raw_value, ca.is_merge_origin) != \
               (cb.row, cb.col, cb.raw_value, cb.is_merge_origin):
                cell_mismatches += 1
                if cell_mismatches <= 3:
                    print(f"  CELL MISMATCH ({ca.row},{ca.col}): val={ca.raw_value!r} vs {cb.raw_value!r}")

        if cell_mismatches == 0:
            print(f"  Merged cells: spot-checked {len(cells_a_sample)} cells — all match OK")

    # ── Summary ────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  Summary ({num_rows:,} rows, {num_rows * 14:,} cells)")
    print(f"{'='*60}")
    print(f"  ORM (current)    : {elapsed_a:.3f}s")
    print(f"  Core bulk insert : {elapsed_b:.3f}s")
    speedup = elapsed_a / elapsed_b if elapsed_b > 0 else float("inf")
    print(f"  Speedup          : {speedup:.1f}x")
    print(f"  Time saved       : {elapsed_a - elapsed_b:.3f}s")

    # Cleanup
    wb_xl.close()
    for f in [db_a, db_b, bench_xlsx]:
        f.unlink(missing_ok=True)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", type=int, default=10000, help="Number of data rows")
    args = parser.parse_args()
    run_benchmark(args.rows)
