"""Benchmark: ORM per-row flush vs Core bulk insert for domain objects.

Generates a large xlsx, then imports it using both strategies and compares timing.

Usage:
    uv run python tests/bench_bulk_insert.py
    uv run python tests/bench_bulk_insert.py --rows 10000
"""

from __future__ import annotations

import argparse
import os
import sys
import time
from pathlib import Path
from typing import Any

# Project setup
PROJECT_ROOT = Path(__file__).resolve().parent.parent
os.chdir(PROJECT_ROOT)
sys.path.insert(0, str(PROJECT_ROOT / "src"))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, insert
from sqlalchemy.orm import Session, sessionmaker

from dsm.models import Base, ExcelCell, ExcelMerge, ExcelSheet, ExcelWorkbook
from dsm.domain_models import REGMAP_FIELD_MAP, Register
from dsm.merge import MergeResolver
from dsm.xlsx_parser import extract_style, find_header_row


# ── 1. Generate large sample xlsx ─────────────────────────────────────

HEADERS = ["TYPE", "INDX", "PAGE", "PARA", "NAME",
           "D7", "D6", "D5", "D4", "D3", "D2", "D1", "D0", "INIT"]

REGISTER_TEMPLATES = [
    ("RW2", "EN", "MODE[1:0]", "RST", "RSVD", "RSVD", "RSVD", "RSVD", "RSVD", "0x00"),
    ("RO",  "BUSY", "DONE", "ERR", "CNT[4:0]", "CNT[4:0]", "CNT[4:0]", "CNT[4:0]", "CNT[4:0]", "0x00"),
    ("RW1", "GAIN[1:0]", "GAIN[1:0]", "DIV[2:0]", "DIV[2:0]", "DIV[2:0]", "POL", "PHASE", "LOCK", "0x24"),
    ("RO",  "DATA[7:0]", "DATA[7:0]", "DATA[7:0]", "DATA[7:0]", "DATA[7:0]", "DATA[7:0]", "DATA[7:0]", "DATA[7:0]", "0xFF"),
]

IP_NAMES = [
    "SENSOR_A", "AMPLIFIER", "DAC_CTRL", "ADC_CONV", "PLL_CFG",
    "GPIO_PORT", "TIMER_A", "SPI_MASTER", "I2C_SLAVE", "PWR_MGMT",
    "SENSOR_B", "SENSOR_C", "DMA_CH0", "DMA_CH1", "UART_0",
    "UART_1", "WATCHDOG", "RTC_TIMER", "CRC_GEN", "FLASH_CTRL",
]


def generate_large_xlsx(path: Path, num_rows: int) -> Path:
    """Generate a large register map xlsx for benchmarking."""
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "level2_bench"

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    header_font = Font(name="Consolas", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    data_font = Font(name="Consolas", size=10)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Headers
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = center

    # Data rows
    indx_counter = 10
    row = 2
    ip_idx = 0
    merge_ranges = []  # (min_row, max_row, col) for vertical merges

    while row <= num_rows + 1:
        ip_name = IP_NAMES[ip_idx % len(IP_NAMES)]
        ip_idx += 1
        indx = str(indx_counter)
        indx_counter += 1

        # Each IP gets 4 rows (one template cycle)
        group_start = row
        for para, tmpl in enumerate(REGISTER_TEMPLATES):
            if row > num_rows + 1:
                break
            type_, d7, d6, d5, d4, d3, d2, d1, d0, init = tmpl

            ws.cell(row=row, column=1, value=type_).font = data_font
            ws.cell(row=row, column=2, value=indx).font = data_font
            ws.cell(row=row, column=3, value="0").font = data_font
            ws.cell(row=row, column=4, value=str(para)).font = data_font
            ws.cell(row=row, column=5, value=ip_name).font = data_font
            ws.cell(row=row, column=6, value=d7).font = data_font
            ws.cell(row=row, column=7, value=d6).font = data_font
            ws.cell(row=row, column=8, value=d5).font = data_font
            ws.cell(row=row, column=9, value=d4).font = data_font
            ws.cell(row=row, column=10, value=d3).font = data_font
            ws.cell(row=row, column=11, value=d2).font = data_font
            ws.cell(row=row, column=12, value=d1).font = data_font
            ws.cell(row=row, column=13, value=d0).font = data_font
            ws.cell(row=row, column=14, value=init).font = data_font

            # Apply borders and fill
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = center
            # Green fill for D7 (1-bit field) on first template
            if para == 0:
                ws.cell(row=row, column=6).fill = green_fill

            row += 1

        group_end = row - 1
        if group_end > group_start:
            # Vertical merge INDX (col 2) and NAME (col 5)
            merge_ranges.append((group_start, group_end, 2))
            merge_ranges.append((group_start, group_end, 5))

    # Apply vertical merges
    for min_row, max_row, col in merge_ranges:
        col_letter = get_column_letter(col)
        ws.merge_cells(f"{col_letter}{min_row}:{col_letter}{max_row}")

    ws.freeze_panes = "A2"
    wb.save(path)
    wb.close()
    return path


# ── 2. Import strategies ──────────────────────────────────────────────

def _import_cells_and_merges(session: Session, ws, sheet_obj: ExcelSheet, merger: MergeResolver):
    """Import cells and merges (shared setup for both strategies)."""
    merge_db: dict[str, ExcelMerge] = {}
    for mr in merger.ranges:
        m = ExcelMerge(
            sheet_id=sheet_obj.id,
            min_row=mr.min_row, min_col=mr.min_col,
            max_row=mr.max_row, max_col=mr.max_col,
        )
        session.add(m)
        session.flush()
        merge_db[f"{mr.min_row}:{mr.min_col}"] = m

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            r, c = cell.row, cell.column
            if merger.is_merged(r, c):
                raw_val = merger.get_value(r, c)
            else:
                raw_val = str(cell.value) if cell.value is not None else None

            merge_key = None
            is_origin = merger.is_origin(r, c)
            origin = merger.get_origin(r, c)
            if origin:
                key = f"{origin[0]}:{origin[1]}"
                merge_key = merge_db[key].id

            style = extract_style(cell) if cell.value is not None or is_origin else None

            session.add(ExcelCell(
                sheet_id=sheet_obj.id, row=r, col=c,
                raw_value=raw_val, style=style,
                merge_id=merge_key, is_merge_origin=is_origin,
            ))
    session.flush()
    return merger


def _detect_header_and_colmap(ws, field_map):
    """Detect header row and build col_to_field mapping."""
    header_row = find_header_row(ws, field_map=field_map)
    headers = {}
    for cell in ws[header_row]:
        if cell.value and isinstance(cell.value, str):
            headers[cell.column] = cell.value.strip()

    col_to_field = {}
    for col_idx, header_text in headers.items():
        if header_text in field_map:
            col_to_field[col_idx] = field_map[header_text]

    return header_row, col_to_field


def import_orm_per_row(session: Session, ws, sheet_obj: ExcelSheet, merger: MergeResolver,
                       header_row: int, col_to_field: dict[int, str]) -> int:
    """Current approach: ORM object per row + session.flush() per row."""
    count = 0
    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        row_data: dict[str, Any] = {}
        has_data = False
        for col_idx, field_name in col_to_field.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is None and merger.is_merged(row_idx, col_idx):
                val = merger.get_value(row_idx, col_idx)
            if val is not None:
                has_data = True
                row_data[field_name] = str(val)
            else:
                row_data[field_name] = None

        if has_data:
            obj = Register(
                sheet_id=sheet_obj.id,
                excel_row=row_idx,
                **row_data,
            )
            session.add(obj)
            session.flush()
            count += 1

    session.flush()
    return count


def import_core_bulk(session: Session, ws, sheet_obj: ExcelSheet, merger: MergeResolver,
                     header_row: int, col_to_field: dict[int, str]) -> int:
    """Optimized: collect dicts, Core-level bulk insert."""
    bulk_rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        row_data: dict[str, Any] = {}
        has_data = False
        for col_idx, field_name in col_to_field.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is None and merger.is_merged(row_idx, col_idx):
                val = merger.get_value(row_idx, col_idx)
            if val is not None:
                has_data = True
                row_data[field_name] = str(val)
            else:
                row_data[field_name] = None

        if has_data:
            row_data["sheet_id"] = sheet_obj.id
            row_data["excel_row"] = row_idx
            bulk_rows.append(row_data)

    if bulk_rows:
        session.execute(insert(Register), bulk_rows)
        session.flush()

    return len(bulk_rows)


# ── 3. Benchmark runner ───────────────────────────────────────────────

def run_benchmark(num_rows: int):
    """Run the benchmark with the specified number of data rows."""
    # Generate test data
    bench_xlsx = Path("samples/bench_large.xlsx")
    print(f"Generating {num_rows}-row xlsx...")
    t0 = time.perf_counter()
    generate_large_xlsx(bench_xlsx, num_rows)
    gen_time = time.perf_counter() - t0
    print(f"  Generated in {gen_time:.2f}s ({bench_xlsx.stat().st_size / 1024:.0f} KB)")

    # Load workbook once (shared across both runs)
    print(f"\nLoading workbook with openpyxl...")
    t0 = time.perf_counter()
    wb_xl = openpyxl.load_workbook(bench_xlsx, data_only=False)
    ws = wb_xl["level2_bench"]
    load_time = time.perf_counter() - t0
    print(f"  Loaded in {load_time:.2f}s")

    # Detect headers (shared)
    header_row, col_to_field = _detect_header_and_colmap(ws, REGMAP_FIELD_MAP)
    print(f"  Header row: {header_row}, fields: {len(col_to_field)}")

    results = {}

    # ── Strategy A: ORM per-row flush ──────────────────────────────
    print(f"\n{'='*60}")
    print(f"  Strategy A: ORM per-row add + flush")
    print(f"{'='*60}")

    db_a = Path("samples/bench_orm.db")
    db_a.unlink(missing_ok=True)
    engine_a = create_engine(f"sqlite:///{db_a}", echo=False)
    Base.metadata.create_all(engine_a)
    SessionA = sessionmaker(bind=engine_a)

    with SessionA() as session:
        blob = bench_xlsx.read_bytes()
        wb_obj = ExcelWorkbook(filename=bench_xlsx.name, blob=blob)
        session.add(wb_obj)
        session.flush()

        sheet_obj = ExcelSheet(workbook_id=wb_obj.id, name=ws.title)
        session.add(sheet_obj)
        session.flush()

        merger = MergeResolver(ws)
        _import_cells_and_merges(session, ws, sheet_obj, merger)
        sheet_obj.header_row = header_row

        # Time only the domain object creation part
        t0 = time.perf_counter()
        count = import_orm_per_row(session, ws, sheet_obj, merger, header_row, col_to_field)
        session.commit()
        elapsed_a = time.perf_counter() - t0

    results["orm"] = elapsed_a
    print(f"  Inserted {count} registers in {elapsed_a:.3f}s")
    print(f"  Rate: {count / elapsed_a:.0f} rows/sec")

    # ── Strategy B: Core bulk insert ───────────────────────────────
    print(f"\n{'='*60}")
    print(f"  Strategy B: Core bulk insert (dict list)")
    print(f"{'='*60}")

    db_b = Path("samples/bench_core.db")
    db_b.unlink(missing_ok=True)
    engine_b = create_engine(f"sqlite:///{db_b}", echo=False)
    Base.metadata.create_all(engine_b)
    SessionB = sessionmaker(bind=engine_b)

    with SessionB() as session:
        blob = bench_xlsx.read_bytes()
        wb_obj = ExcelWorkbook(filename=bench_xlsx.name, blob=blob)
        session.add(wb_obj)
        session.flush()

        sheet_obj = ExcelSheet(workbook_id=wb_obj.id, name=ws.title)
        session.add(sheet_obj)
        session.flush()

        merger = MergeResolver(ws)
        _import_cells_and_merges(session, ws, sheet_obj, merger)
        sheet_obj.header_row = header_row

        # Time only the domain object creation part
        t0 = time.perf_counter()
        count = import_core_bulk(session, ws, sheet_obj, merger, header_row, col_to_field)
        session.commit()
        elapsed_b = time.perf_counter() - t0

    results["core"] = elapsed_b
    print(f"  Inserted {count} registers in {elapsed_b:.3f}s")
    print(f"  Rate: {count / elapsed_b:.0f} rows/sec")

    # ── Verify correctness ─────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  Verification")
    print(f"{'='*60}")

    with SessionA() as sa, SessionB() as sb:
        regs_a = sa.query(Register).order_by(Register.excel_row).all()
        regs_b = sb.query(Register).order_by(Register.excel_row).all()

        assert len(regs_a) == len(regs_b), f"Count mismatch: {len(regs_a)} vs {len(regs_b)}"
        print(f"  Row count: {len(regs_a)} == {len(regs_b)} OK")

        mismatches = 0
        for ra, rb in zip(regs_a, regs_b):
            for field in ["type", "indx", "page", "para", "name",
                          "d7", "d6", "d5", "d4", "d3", "d2", "d1", "d0",
                          "init", "excel_row"]:
                va = getattr(ra, field)
                vb = getattr(rb, field)
                if va != vb:
                    mismatches += 1
                    if mismatches <= 5:
                        print(f"  MISMATCH row={ra.excel_row} field={field}: {va!r} vs {vb!r}")

        if mismatches == 0:
            print(f"  All field values match OK")
        else:
            print(f"  {mismatches} mismatches found!")

    # ── Summary ────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  Summary ({num_rows} data rows)")
    print(f"{'='*60}")
    print(f"  ORM per-row flush : {results['orm']:.3f}s")
    print(f"  Core bulk insert  : {results['core']:.3f}s")
    speedup = results["orm"] / results["core"] if results["core"] > 0 else float("inf")
    print(f"  Speedup           : {speedup:.1f}x")

    # Cleanup
    wb_xl.close()
    for f in [db_a, db_b, bench_xlsx]:
        f.unlink(missing_ok=True)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", type=int, default=5000, help="Number of data rows")
    args = parser.parse_args()
    run_benchmark(args.rows)
