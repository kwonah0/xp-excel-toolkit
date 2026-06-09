from __future__ import annotations

from pathlib import Path

import openpyxl
import xlrd
import xlwt
from openpyxl.styles import Font

from xp_excel_toolkit.exporter import apply_style
from xp_excel_toolkit.xls_parser import _extract_style_xls
from xp_excel_toolkit.xlsx_parser import extract_style


def test_extract_style_reads_strikethrough_from_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="struck")
    cell.font = Font(strike=True)

    style = extract_style(cell)

    assert style is not None
    assert style.get("font_strike") is True


def test_extract_style_omits_strikethrough_when_absent():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="plain")

    style = extract_style(cell)

    assert not (style or {}).get("font_strike")


def test_apply_style_writes_strikethrough_to_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="x")

    apply_style(cell, {"font_strike": True})

    assert cell.font.strike is True


def test_xlsx_strikethrough_round_trips(tmp_path: Path):
    src = tmp_path / "src.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="struck")
    cell.font = Font(strike=True)
    wb.save(src)

    reopened = openpyxl.load_workbook(src)
    style = extract_style(reopened.active.cell(row=1, column=1))

    out = tmp_path / "out.xlsx"
    wb2 = openpyxl.Workbook()
    apply_style(wb2.active.cell(row=1, column=1, value="struck"), style)
    wb2.save(out)

    final = openpyxl.load_workbook(out)
    assert final.active.cell(row=1, column=1).font.strike is True


def test_extract_style_xls_reads_strikethrough(tmp_path: Path):
    path = tmp_path / "s.xls"
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    font = xlwt.Font()
    font.struck_out = True
    xf = xlwt.XFStyle()
    xf.font = font
    ws.write(0, 0, "struck", xf)
    wb.save(str(path))

    book = xlrd.open_workbook(str(path), formatting_info=True)
    sheet = book.sheet_by_index(0)
    xf_index = sheet.cell_xf_index(0, 0)
    style = _extract_style_xls(book, xf_index)

    assert style is not None
    assert style.get("font_strike") is True
