"""Profile each phase of _import_ws() to find remaining bottlenecks.

Breaks down time spent in:
  1. openpyxl load_workbook()
  2. MergeResolver construction
  3. ExcelMerge bulk insert
  4. Cell iteration + dict building
  5. ExcelCell bulk insert (DB write)
  6. Header detection + domain dict building
  7. Domain bulk insert (DB write)

Usage:
    uv run python tests/profile_phases.py
    uv run python tests/profile_phases.py --rows 10000
"""

from __future__ import annotations

import argparse
import os
import sys
import time
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent
os.chdir(PROJECT_ROOT)
sys.path.insert(0, str(PROJECT_ROOT / "src"))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, insert
from sqlalchemy.orm import Session, sessionmaker

from excel_toolkit.models import Base, ExcelCell, ExcelMerge, ExcelSheet, ExcelWorkbook
from excel_toolkit.domain_models import REGMAP_FIELD_MAP, Register
from excel_toolkit.merge import MergeResolver
from excel_toolkit.xlsx_parser import extract_style, find_header_row, _BULK_CHUNK


# Reuse generate from bench
from bench_full_pipeline import generate_large_xlsx


def profile_import(num_rows: int):
    bench_xlsx = Path("samples/bench_profile.xlsx")

    print(f"=== Phase Profiling ({num_rows:,} rows × 14 cols = {num_rows * 14:,} cells) ===\n")

    # Generate
    t0 = time.perf_counter()
    generate_large_xlsx(bench_xlsx, num_rows)
    print(f"[gen] xlsx generated in {time.perf_counter() - t0:.2f}s  ({bench_xlsx.stat().st_size / 1024:.0f} KB)\n")

    timings: dict[str, float] = {}

    # Phase 1: openpyxl load
    t0 = time.perf_counter()
    wb_xl = openpyxl.load_workbook(bench_xlsx, data_only=False)
    ws = wb_xl["level2_bench"]
    timings["1_openpyxl_load"] = time.perf_counter() - t0

    # Phase 1b: file read (blob)
    t0 = time.perf_counter()
    blob = bench_xlsx.read_bytes()
    timings["1b_file_read"] = time.perf_counter() - t0

    # Setup DB
    db_path = Path("samples/bench_profile.db")
    db_path.unlink(missing_ok=True)
    engine = create_engine(f"sqlite:///{db_path}", echo=False)
    Base.metadata.create_all(engine)
    S = sessionmaker(bind=engine)
    session = S()

    wb_obj = ExcelWorkbook(filename=bench_xlsx.name, blob=blob)
    session.add(wb_obj)
    session.flush()
    sheet_obj = ExcelSheet(workbook_id=wb_obj.id, name=ws.title)
    session.add(sheet_obj)
    session.flush()
    sid = sheet_obj.id

    # Phase 2: MergeResolver construction
    t0 = time.perf_counter()
    merger = MergeResolver(ws)
    timings["2_merge_resolver"] = time.perf_counter() - t0

    # Phase 3: ExcelMerge bulk insert
    t0 = time.perf_counter()
    merge_dicts = [
        {"sheet_id": sid, "min_row": mr.min_row, "min_col": mr.min_col,
         "max_row": mr.max_row, "max_col": mr.max_col}
        for mr in merger.ranges
    ]
    merge_id_map: dict[str, int] = {}
    if merge_dicts:
        for i in range(0, len(merge_dicts), _BULK_CHUNK):
            session.execute(insert(ExcelMerge), merge_dicts[i:i + _BULK_CHUNK])
        session.flush()
        for mid, mrow, mcol in (
            session.query(ExcelMerge.id, ExcelMerge.min_row, ExcelMerge.min_col)
            .filter(ExcelMerge.sheet_id == sid).all()
        ):
            merge_id_map[f"{mrow}:{mcol}"] = mid
    timings["3_merge_insert"] = time.perf_counter() - t0

    # Phase 4: Cell iteration + dict building (NO DB write)
    t0 = time.perf_counter()
    cell_dicts: list[dict] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            r, c = cell.row, cell.column
            if merger.is_merged(r, c):
                raw_val = merger.get_value(r, c)
            else:
                raw_val = str(cell.value) if cell.value is not None else None
            merge_key = None
            is_origin = merger.is_origin(r, c)
            origin = merger.get_origin(r, c)
            if origin:
                merge_key = merge_id_map.get(f"{origin[0]}:{origin[1]}")
            style = extract_style(cell) if cell.value is not None or is_origin else None
            cell_dicts.append({
                "sheet_id": sid, "row": r, "col": c,
                "raw_value": raw_val, "style": style,
                "merge_id": merge_key, "is_merge_origin": is_origin,
            })
    timings["4_cell_iter_dict"] = time.perf_counter() - t0

    # Phase 4b: measure extract_style alone
    t0 = time.perf_counter()
    style_count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None or merger.is_origin(cell.row, cell.column):
                extract_style(cell)
                style_count += 1
    timings["4b_extract_style_only"] = time.perf_counter() - t0

    # Phase 5: ExcelCell bulk insert (DB write only)
    t0 = time.perf_counter()
    if cell_dicts:
        for i in range(0, len(cell_dicts), _BULK_CHUNK):
            session.execute(insert(ExcelCell), cell_dicts[i:i + _BULK_CHUNK])
        session.flush()
    timings["5_cell_db_write"] = time.perf_counter() - t0

    # Phase 6: Header detection + domain dict building
    t0 = time.perf_counter()
    field_map = REGMAP_FIELD_MAP
    header_row = find_header_row(ws, field_map=field_map)
    sheet_obj.header_row = header_row

    headers: dict[int, str] = {}
    for cell in ws[header_row]:
        if cell.value and isinstance(cell.value, str):
            headers[cell.column] = cell.value.strip()
    col_to_field: dict[int, str] = {}
    for col_idx, header_text in headers.items():
        if header_text in field_map:
            col_to_field[col_idx] = field_map[header_text]

    bulk_rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        row_data: dict[str, Any] = {}
        has_data = False
        for col_idx, field_name in col_to_field.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is None and merger.is_merged(row_idx, col_idx):
                val = merger.get_value(row_idx, col_idx)
            if val is not None:
                has_data = True
                row_data[field_name] = str(val)
            else:
                row_data[field_name] = None
        if has_data:
            row_data["sheet_id"] = sid
            row_data["excel_row"] = row_idx
            bulk_rows.append(row_data)
    timings["6_domain_dict_build"] = time.perf_counter() - t0

    # Phase 7: Domain bulk insert (DB write only)
    t0 = time.perf_counter()
    if bulk_rows:
        for i in range(0, len(bulk_rows), _BULK_CHUNK):
            session.execute(insert(Register), bulk_rows[i:i + _BULK_CHUNK])
        session.flush()
    timings["7_domain_db_write"] = time.perf_counter() - t0

    session.commit()
    session.close()

    # Report
    total = sum(timings.values())
    print(f"{'Phase':<30} {'Time':>8} {'%':>6}")
    print("-" * 48)
    for phase, elapsed in sorted(timings.items()):
        pct = elapsed / total * 100
        bar = "#" * int(pct / 2)
        print(f"  {phase:<28} {elapsed:>7.3f}s {pct:>5.1f}%  {bar}")
    print("-" * 48)
    print(f"  {'TOTAL':<28} {total:>7.3f}s")

    print(f"\n  Cells: {len(cell_dicts):,}")
    print(f"  Merges: {len(merge_dicts):,}")
    print(f"  Domain rows: {len(bulk_rows):,}")
    print(f"  Styles extracted: {style_count:,}")

    # Cleanup
    wb_xl.close()
    db_path.unlink(missing_ok=True)
    bench_xlsx.unlink(missing_ok=True)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", type=int, default=10000)
    args = parser.parse_args()
    profile_import(args.rows)
