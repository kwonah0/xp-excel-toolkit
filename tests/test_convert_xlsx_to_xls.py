"""convert_xlsx_to_xls — LibreOffice .xlsx → legacy .xls (round-trip of xls→xlsx)."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
import pytest

from xp_excel_toolkit.ingest.convert import convert_xls_to_xlsx, convert_xlsx_to_xls


def _libreoffice_available() -> bool:
    return any(shutil.which(b) for b in ("soffice", "libreoffice"))


pytestmark = pytest.mark.skipif(
    not _libreoffice_available(), reason="LibreOffice not available")


def test_xlsx_to_xls_and_back(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "hi"
    ws["B2"] = 42
    src = tmp_path / "t.xlsx"
    wb.save(src)

    xls = convert_xlsx_to_xls(src)
    assert xls.exists() and xls.suffix == ".xls"
    assert xls.parent == src.parent          # default output_dir = sibling

    back = convert_xls_to_xlsx(xls, output_dir=tmp_path)
    rt = openpyxl.load_workbook(back).active
    assert rt["A1"].value == "hi" and rt["B2"].value == 42


def test_missing_input_raises(tmp_path: Path):
    with pytest.raises(FileNotFoundError):
        convert_xlsx_to_xls(tmp_path / "nope.xlsx")
