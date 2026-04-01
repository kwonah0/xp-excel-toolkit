"""Case 2: Split regmap by IP — level2 시트별로 IP 단위 xlsx 분리.

사용법:
    uv run python tests/manual/case_2_split.py

이 스크립트는:
  1. split_regmap()으로 regmap_sample.xlsx를 IP 단위로 분리
  2. level2_common.xlsx → SENSOR_A, AMPLIFIER, DAC_CTRL, ADC_CONV, PLL_CFG 시트
  3. level2_buscon.xlsx → GPIO_PORT, TIMER_A, SPI_MASTER, I2C_SLAVE, PWR_MGMT 시트
  4. 생성된 각 파일의 시트 구조 확인
  5. 분리된 시트 내 merge/데이터 보존 확인
"""

import os
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
os.chdir(PROJECT_ROOT)

from dsm import split_regmap, init_db

SAMPLE = Path("samples/regmap_sample.xlsx")
OUTPUT_DIR = Path("samples/split_output")
DB_PATH = Path("samples/case_2.db")


def main():
    DB_PATH.unlink(missing_ok=True)
    if OUTPUT_DIR.exists():
        for f in OUTPUT_DIR.glob("*.xlsx"):
            f.unlink()

    Session = init_db(f"sqlite:///{DB_PATH}")

    # ── 1. Split ───────────────────────────────────────────────
    print("=" * 60)
    print("  Split regmap by IP")
    print("=" * 60)

    with Session() as session:
        results = split_regmap(session, SAMPLE, OUTPUT_DIR)
        session.commit()

    print(f"\n  생성된 파일:")
    for src_name, out_path in results.items():
        print(f"    {src_name} → {out_path}")

    # ── 2. 각 파일 구조 확인 ───────────────────────────────────
    for src_name, out_path in results.items():
        print(f"\n{'=' * 60}")
        print(f"  {out_path.name}")
        print("=" * 60)

        wb = openpyxl.load_workbook(out_path, read_only=False)
        print(f"  Sheets: {wb.sheetnames}")

        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f"\n  [{sn}]")
            print(f"    Size: {ws.max_row} rows x {ws.max_column} cols")

            # Header 확인
            headers = [ws.cell(1, c).value for c in range(1, (ws.max_column or 0) + 1)]
            print(f"    Headers: {headers}")

            # Merge 확인
            merge_count = len(ws.merged_cells.ranges)
            h_merges = [mr for mr in ws.merged_cells.ranges if mr.min_row == mr.max_row]
            v_merges = [mr for mr in ws.merged_cells.ranges if mr.min_col == mr.max_col]
            print(f"    Merges: {merge_count} (H={len(h_merges)}, V={len(v_merges)})")

            # 데이터 행 출력
            print(f"    Data:")
            for r in range(2, (ws.max_row or 1) + 1):
                row_vals = []
                for c in range(1, min(15, (ws.max_column or 0) + 1)):
                    v = ws.cell(r, c).value
                    row_vals.append(str(v) if v else "-")
                type_ = row_vals[0] if len(row_vals) > 0 else ""
                indx = row_vals[1] if len(row_vals) > 1 else ""
                name = row_vals[4] if len(row_vals) > 4 else ""
                bits = " | ".join(row_vals[5:13]) if len(row_vals) > 12 else ""
                init_ = row_vals[13] if len(row_vals) > 13 else ""
                print(f"      row {r}: {type_:>3} idx={indx:>3} {name:>12}  [{bits}]  {init_}")

            # 색상 확인 (첫 데이터 행의 bit field 영역)
            if ws.max_row and ws.max_row >= 2:
                print(f"    Colors (row 2, D7..D0):")
                for c in range(6, 14):
                    cell = ws.cell(2, c)
                    bg = None
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                        rgb = cell.fill.fgColor.rgb
                        if isinstance(rgb, str) and rgb != "00000000":
                            bg = rgb
                    val = cell.value or "-"
                    col_name = f"D{13 - c}"
                    print(f"      {col_name} (col {c}): value={val}, bg={bg}")

        wb.close()

    print(f"\n출력 디렉토리: {OUTPUT_DIR.resolve()}")
    print(f"DB: {DB_PATH.resolve()}")
    print("Done!")


if __name__ == "__main__":
    main()
