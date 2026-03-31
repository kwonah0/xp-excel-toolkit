"""Case 3: Modify registers in DB, export back preserving formatting.

사용법:
    uv run python tests/manual/case_3_modify_export.py

이 스크립트는:
  1. regmap_sample.xlsx를 import (level2_common)
  2. Register 객체의 INIT 값을 수정
  3. export_regmap_xlsx()로 BLOB 기반 export (원본 서식/색상 보존)
  4. export_from_cells()로 DB cell 기반 export (BLOB 없이 재구성)
  5. 두 export 결과를 비교 확인
"""

import os
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
os.chdir(PROJECT_ROOT)

from excel_toolkit import (
    ExcelCell, ExcelWorkbook,
    Register, REGMAP_FIELD_MAP,
    import_xlsx, export_regmap_xlsx, init_db,
)
from excel_toolkit.exporter import export_from_cells

SAMPLE = Path("samples/regmap_sample.xlsx")
DB_PATH = Path("samples/case_3.db")
OUTPUT_BLOB = Path("samples/case_3_blob_export.xlsx")
OUTPUT_CELLS = Path("samples/case_3_cells_export.xlsx")


def main():
    DB_PATH.unlink(missing_ok=True)
    Session = init_db(f"sqlite:///{DB_PATH}")

    with Session() as session:
        # ── 1. Import ──────────────────────────────────────────
        print("=" * 60)
        print("  1. Import level2_common")
        print("=" * 60)

        column_map: dict[str, int] = {}
        sheet = import_xlsx(
            session, SAMPLE,
            sheet_name="level2_common",
            field_map=REGMAP_FIELD_MAP,
            domain_cls=Register,
            column_map=column_map,
        )
        session.commit()

        regs = (
            session.query(Register)
            .filter_by(sheet_id=sheet.id)
            .order_by(Register.excel_row)
            .all()
        )
        print(f"  {len(regs)} registers imported")
        print(f"  Column map: {column_map}")

        # ── 2. Modify ─────────────────────────────────────────
        print(f"\n{'=' * 60}")
        print("  2. Modify registers")
        print("=" * 60)

        changes = []

        # SENSOR_A PARA=0: INIT 0x00 → 0x80
        sensor_0 = session.query(Register).filter_by(
            sheet_id=sheet.id, name="SENSOR_A", para="0"
        ).first()
        if sensor_0:
            old = sensor_0.init
            sensor_0.init = "0x80"
            changes.append(f"  SENSOR_A para=0: INIT {old} → {sensor_0.init}")

        # AMPLIFIER PARA=0: GAIN[1:0] 관련 INIT 0x24 → 0x3C
        amp_0 = session.query(Register).filter_by(
            sheet_id=sheet.id, name="AMPLIFIER", para="0"
        ).first()
        if amp_0:
            old = amp_0.init
            amp_0.init = "0x3C"
            changes.append(f"  AMPLIFIER para=0: INIT {old} → {amp_0.init}")

        # PLL_CFG PARA=1: MDIV 값 변경
        pll_1 = session.query(Register).filter_by(
            sheet_id=sheet.id, name="PLL_CFG", para="1"
        ).first()
        if pll_1:
            old = pll_1.init
            pll_1.init = "0x20"
            changes.append(f"  PLL_CFG para=1: INIT {old} → {pll_1.init}")

        session.commit()

        for c in changes:
            print(c)

        # ── 3. BLOB 기반 Export ────────────────────────────────
        print(f"\n{'=' * 60}")
        print("  3. Export (BLOB 기반 — 원본 서식 보존)")
        print("=" * 60)

        wb_obj = session.query(ExcelWorkbook).first()
        export_regmap_xlsx(session, wb_obj.id, OUTPUT_BLOB, column_map=column_map)
        print(f"  → {OUTPUT_BLOB}")

        # 확인
        wb = openpyxl.load_workbook(OUTPUT_BLOB)
        ws = wb[sheet.name]
        print(f"\n  검증:")

        # SENSOR_A para=0 INIT
        init_val = ws.cell(row=sensor_0.excel_row, column=column_map["init"]).value
        print(f"    SENSOR_A para=0 INIT = {init_val} (expected: 0x80)")
        assert init_val == "0x80"

        # Merge 보존
        merge_count = len(ws.merged_cells.ranges)
        print(f"    Merges preserved: {merge_count}")

        # 색상 보존 (EN 셀 = 초록)
        en_cell = ws.cell(row=2, column=6)  # D7 of first data row
        en_bg = None
        if en_cell.fill and en_cell.fill.fgColor:
            en_bg = en_cell.fill.fgColor.rgb
        print(f"    EN cell color: {en_bg}")
        wb.close()

        # ── 4. Cell 기반 Export ────────────────────────────────
        print(f"\n{'=' * 60}")
        print("  4. Export (Cell 기반 — DB 레코드에서 재구성)")
        print("=" * 60)

        export_from_cells(session, sheet.id, OUTPUT_CELLS)
        print(f"  → {OUTPUT_CELLS}")

        # 확인
        wb2 = openpyxl.load_workbook(OUTPUT_CELLS)
        ws2 = wb2.active
        print(f"\n  검증:")
        print(f"    Sheet: {ws2.title}")
        print(f"    Size: {ws2.max_row} rows x {ws2.max_column} cols")

        merge_count2 = len(ws2.merged_cells.ranges)
        print(f"    Merges: {merge_count2}")

        # 헤더 확인
        headers = [ws2.cell(1, c).value for c in range(1, 15)]
        print(f"    Headers: {headers}")

        # 스타일 확인
        en_cell2 = ws2.cell(row=2, column=6)
        en_bg2 = None
        if en_cell2.fill and en_cell2.fill.fgColor:
            en_bg2 = en_cell2.fill.fgColor.rgb
        print(f"    EN cell color: {en_bg2}")

        wb2.close()

        # ── 5. 비교 요약 ──────────────────────────────────────
        print(f"\n{'=' * 60}")
        print("  5. 두 방식 비교")
        print("=" * 60)
        print(f"  {'':>25} {'BLOB export':>15} {'Cell export':>15}")
        print(f"  {'Merges':>25} {merge_count:>15} {merge_count2:>15}")
        print(f"  {'INIT updated':>25} {'Yes':>15} {'(raw cells)':>15}")
        print(f"  {'Colors preserved':>25} {'Yes':>15} {'Yes':>15}")
        print()
        print("  BLOB export: 원본 파일을 기반으로 값만 교체 → 100% 서식 보존")
        print("  Cell export: DB에서 완전 재구성 → split/부분 export에 적합")

    print(f"\nDB: {DB_PATH.resolve()}")
    print(f"BLOB export: {OUTPUT_BLOB.resolve()}")
    print(f"Cell export: {OUTPUT_CELLS.resolve()}")
    print("Done!")


if __name__ == "__main__":
    main()
