"""import_xlsx sheet_names 필터 — 지정 시트만 cell 적재, blob 은 전체 보존."""
from __future__ import annotations

import openpyxl
from sqlalchemy import select

from xp_excel_toolkit import ExcelSheet, ExcelWorkbook, import_xlsx, init_db


def _make_multi(path):
    wb = openpyxl.Workbook()
    wb.active.title = "Spec"
    wb.active["A1"] = "spec_data"
    wb.create_sheet("Other")["A1"] = "other_data"
    wb.create_sheet("Notes")["A1"] = "notes_data"
    wb.save(path)


def test_sheet_names_filter_imports_only_selected(tmp_path):
    xlsx = tmp_path / "multi.xlsx"
    _make_multi(xlsx)
    sf = init_db(f"sqlite:///{tmp_path / 't.db'}")
    with sf() as s:
        import_xlsx(s, xlsx, sheet_names=["Spec"])
        s.commit()
        names = {sh.name for sh in s.scalars(select(ExcelSheet)).all()}
        assert names == {"Spec"}                      # 지정 시트만 적재
        wb = s.scalar(select(ExcelWorkbook))
        assert wb is not None and wb.blob             # 원본 blob 전체 보존(round-trip export 용)


def test_default_imports_all_sheets(tmp_path):
    xlsx = tmp_path / "multi.xlsx"
    _make_multi(xlsx)
    sf = init_db(f"sqlite:///{tmp_path / 't.db'}")
    with sf() as s:
        import_xlsx(s, xlsx)
        s.commit()
        names = {sh.name for sh in s.scalars(select(ExcelSheet)).all()}
        assert names == {"Spec", "Other", "Notes"}    # 기본 = 전부
