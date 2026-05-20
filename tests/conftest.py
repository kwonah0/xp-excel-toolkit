"""Test fixtures: a small synthetic xlsx with merges, styles, comments, and
two domain-shaped sheets, used by the round-trip tests.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side
from sqlalchemy import ForeignKey, Text
from sqlalchemy.orm import Mapped, mapped_column, relationship

from excel_toolkit import Base, ExcelSheet


# ── Toy domain models (mimic a host package like pinmap) ────────────

class PinEntry(Base):
    """A flat-table-style domain row."""

    __tablename__ = "test_pin_entry"

    id: Mapped[int] = mapped_column(primary_key=True)
    pin_no: Mapped[str | None] = mapped_column(Text)
    name: Mapped[str | None] = mapped_column(Text)
    direction: Mapped[str | None] = mapped_column(Text)

    sheet_id: Mapped[int | None] = mapped_column(ForeignKey("excel_sheet.id"))
    excel_row: Mapped[int | None]

    sheet: Mapped[ExcelSheet | None] = relationship()


PIN_FIELD_MAP = {
    "Pin": "pin_no",
    "Name": "name",
    "Dir": "direction",
}


@pytest.fixture
def pin_field_map() -> dict[str, str]:
    return PIN_FIELD_MAP


@pytest.fixture
def pin_domain_cls():
    return PinEntry


# ── Synthetic xlsx fixture ──────────────────────────────────────────

@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """Build a small xlsx with:
      - a Pinmap sheet (flat table — header + 4 data rows)
      - a Notes sheet (plain cells, no domain) — exercises non-domain sheet passthrough
      - one horizontal merge (title row) + one vertical merge (grouped pins)
      - cell background colors, bold font, a thin border, a comment
    """
    path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()

    # ── Pinmap sheet ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Pinmap_A"

    # Title row (will be horizontally merged across 3 cols)
    ws.cell(row=1, column=1, value="Pinmap for Chip A")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    # Header row at row 2
    headers = ["Pin", "Name", "Dir"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Data rows
    data = [
        ("A1", "VDD",  "PWR"),
        ("A2", "GND",  "GND"),
        ("A3", "SDA",  "I/O"),
        ("A4", "SCL",  "I/O"),
    ]
    for i, (pn, nm, di) in enumerate(data, start=3):
        ws.cell(row=i, column=1, value=pn)
        ws.cell(row=i, column=2, value=nm)
        ws.cell(row=i, column=3, value=di)

    # Vertical merge on Dir column for I/O group (rows 5-6)
    ws.merge_cells(start_row=5, start_column=3, end_row=6, end_column=3)
    # (the origin already holds "I/O" from row 5)

    # Style on a data cell + comment
    vdd_cell = ws.cell(row=3, column=2)  # VDD name
    vdd_cell.border = Border(top=Side(style="thin"))
    vdd_cell.comment = Comment("Supply pin", "tester")

    # ── Notes sheet (non-domain — should pass through untouched) ─
    notes = wb.create_sheet("Notes")
    notes.cell(row=1, column=1, value="Free-form notes")
    notes.cell(row=2, column=1, value="line 1")
    notes.cell(row=3, column=1, value="line 2")

    wb.save(path)
    wb.close()
    return path
