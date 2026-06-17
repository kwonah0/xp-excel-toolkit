"""구조 편집 — insert/delete + 병합 경계 조정 + 서식 클론."""
from __future__ import annotations

import copy

import openpyxl
from openpyxl.styles import Font

from xp_excel_toolkit import delete_cols, delete_rows, insert_cols, insert_rows


def _wb():
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, 6):
        for c in range(1, 6):
            ws.cell(r, c, f"r{r}c{c}")
    ws.cell(2, 2).font = Font(bold=True)        # 스타일 표식
    ws.merge_cells("B1:D1")                       # 열 그룹 병합
    return ws


def test_insert_rows_shifts_values():
    ws = _wb()
    insert_rows(ws, 3)
    assert ws.cell(4, 2).value == "r3c2"          # 3행→4행
    assert ws.cell(3, 2).value is None            # 새 행 빈칸


def test_insert_cols_extends_spanning_merge():
    ws = _wb()
    insert_cols(ws, 3)                            # B1:D1(2~4) 한가운데
    assert "B1:E1" in {str(m) for m in ws.merged_cells.ranges}   # 2~5 로 확장


def test_insert_cols_shifts_right_merge():
    ws = _wb()
    ws.merge_cells("D1:E1")                       # insert 우측 병합
    insert_cols(ws, 3)
    assert "E1:F1" in {str(m) for m in ws.merged_cells.ranges}   # 통째 이동


def test_insert_clones_style():
    ws = _wb()
    insert_rows(ws, 3, copy_style_from=2)         # bold 인 2행 서식 클론
    assert ws.cell(3, 2).font.bold is True


def test_delete_cols_shrinks_merge():
    ws = _wb()
    insert_cols(ws, 3)                            # B1:E1
    delete_cols(ws, 3)                            # 다시 B1:D1
    assert "B1:D1" in {str(m) for m in ws.merged_cells.ranges}


def test_delete_rows_pulls_up():
    ws = _wb()
    delete_rows(ws, 3)
    assert ws.cell(3, 2).value == "r4c2"          # 4행→3행


def test_delete_start_of_merge_clamps_like_excel():
    ws = _wb()                                    # merge B1:D1 (cols 2~4)
    delete_cols(ws, 2)                            # 병합 시작 열(B) 삭제
    # 엑셀: min 은 idx 로 clamp → B1:C1 (왼쪽으로 확장 X)
    assert "B1:C1" in {str(m) for m in ws.merged_cells.ranges}


def test_delete_end_of_merge_shrinks():
    ws = _wb()
    delete_cols(ws, 4)                            # 병합 끝 열(D) 삭제
    assert "B1:C1" in {str(m) for m in ws.merged_cells.ranges}
