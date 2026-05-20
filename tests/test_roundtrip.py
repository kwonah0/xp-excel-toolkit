"""End-to-end smoke tests for excel_toolkit.

Covers the round-trip: synthetic xlsx → import (cells + merges + styles +
comments + domain ORM rows) → modify a domain row → export with original
formatting preserved → re-load and assert.
"""

from __future__ import annotations

import openpyxl

from excel_toolkit import (
    ExcelCell,
    ExcelMerge,
    ExcelWorkbook,
    ExportHandler,
    SheetConfig,
    export_domain_xlsx,
    import_xlsx,
    init_db,
)


def test_import_cells_and_merges(sample_xlsx):
    Session = init_db("sqlite:///:memory:")
    with Session() as session:
        sheets = import_xlsx(session, sample_xlsx, sheet_configs={})
        session.commit()

        names = {s.name for s in sheets}
        assert names == {"Pinmap_A", "Notes"}

        pinmap = next(s for s in sheets if s.name == "Pinmap_A")

        # Header row written: title at row 1, headers at row 2, 4 data rows
        # Total cells with content >= 3 (title) + 3 (headers) + 4*3 (data) = 18
        cells = session.query(ExcelCell).filter_by(sheet_id=pinmap.id).all()
        non_empty = [c for c in cells if c.raw_value]
        assert len(non_empty) >= 18

        # Two merges: title (1 horizontal) + Dir group (1 vertical)
        merges = session.query(ExcelMerge).filter_by(sheet_id=pinmap.id).all()
        assert len(merges) == 2

        horizontal = [m for m in merges if m.max_col > m.min_col]
        vertical = [m for m in merges if m.max_row > m.min_row]
        assert len(horizontal) == 1
        assert len(vertical) == 1


def test_import_styles_and_comments(sample_xlsx):
    Session = init_db("sqlite:///:memory:")
    with Session() as session:
        sheets = import_xlsx(session, sample_xlsx, sheet_configs={})
        session.commit()

        pinmap = next(s for s in sheets if s.name == "Pinmap_A")

        # Header row 2 has yellow fill + bold
        header = (
            session.query(ExcelCell)
            .filter_by(sheet_id=pinmap.id, row=2, col=1)
            .one()
        )
        assert header.style is not None
        assert header.style.get("bg_color") == "#FFFF00"
        assert header.style.get("font_bold") is True

        # VDD cell has comment + top border
        vdd = (
            session.query(ExcelCell)
            .filter_by(sheet_id=pinmap.id, row=3, col=2)
            .one()
        )
        assert vdd.comment is not None
        assert "Supply" in vdd.comment
        assert vdd.style and vdd.style.get("border_top") == "thin"


def test_domain_rows_via_sheet_config(sample_xlsx, pin_field_map, pin_domain_cls):
    """SheetConfig with field_map + domain_cls produces ORM rows."""
    sheet_configs = {
        "Pinmap_*": SheetConfig(
            field_map=pin_field_map,
            domain_cls=pin_domain_cls,
        ),
    }

    Session = init_db("sqlite:///:memory:")
    with Session() as session:
        import_xlsx(session, sample_xlsx, sheet_configs=sheet_configs)
        session.commit()

        rows = session.query(pin_domain_cls).order_by(pin_domain_cls.excel_row).all()
        assert len(rows) == 4

        # The header is at row 2; data starts at row 3
        first = rows[0]
        assert first.pin_no == "A1"
        assert first.name == "VDD"
        assert first.direction == "PWR"
        assert first.excel_row == 3

        # Vertical merge fills the merged Dir cell on row 6
        scl = next(r for r in rows if r.pin_no == "A4")
        assert scl.direction == "I/O"   # filled from merge origin (row 5)


def test_roundtrip_modify_and_export(tmp_path, sample_xlsx, pin_field_map, pin_domain_cls):
    """Modify a domain row, export, reload, and confirm both the new value AND
    the original formatting (merge ranges, header fill) survive."""
    sheet_configs = {
        "Pinmap_*": SheetConfig(
            field_map=pin_field_map,
            domain_cls=pin_domain_cls,
        ),
    }

    Session = init_db("sqlite:///:memory:")
    with Session() as session:
        import_xlsx(session, sample_xlsx, sheet_configs=sheet_configs)
        session.commit()

        # Mutate the SDA row
        sda = session.query(pin_domain_cls).filter_by(pin_no="A3").one()
        sda.name = "SDA_NEW"
        session.commit()

        out = tmp_path / "out.xlsx"
        handlers = [
            ExportHandler(
                pattern="Pinmap_*",
                field_map=pin_field_map,
                domain_cls=pin_domain_cls,
            ),
        ]
        export_domain_xlsx(session, out, handlers)

        # Sanity: workbook exists in DB
        wb_obj = session.query(ExcelWorkbook).first()
        assert wb_obj is not None

    # Re-read exported xlsx
    wb_check = openpyxl.load_workbook(out)
    assert set(wb_check.sheetnames) == {"Pinmap_A", "Notes"}

    ws = wb_check["Pinmap_A"]

    # Mutated value made it through
    assert ws.cell(row=5, column=2).value == "SDA_NEW"

    # Other domain values unchanged
    assert ws.cell(row=3, column=2).value == "VDD"

    # Non-domain sheet untouched
    notes = wb_check["Notes"]
    assert notes.cell(row=1, column=1).value == "Free-form notes"

    # Merges preserved (1 horizontal title + 1 vertical Dir = 2)
    assert len(ws.merged_cells.ranges) == 2

    # Header fill preserved
    hdr_fill = ws.cell(row=2, column=1).fill
    assert hdr_fill.fgColor.rgb.endswith("FFFF00")

    wb_check.close()


def test_export_from_cells_no_blob(tmp_path, sample_xlsx):
    """export_from_cells builds an xlsx from cell/merge records alone."""
    from excel_toolkit import export_from_cells

    Session = init_db("sqlite:///:memory:")
    with Session() as session:
        sheets = import_xlsx(session, sample_xlsx, sheet_configs={})
        session.commit()

        pinmap = next(s for s in sheets if s.name == "Pinmap_A")
        out = tmp_path / "from_cells.xlsx"
        export_from_cells(session, pinmap.id, out)

    wb_check = openpyxl.load_workbook(out)
    ws = wb_check.active
    assert ws.title == "Pinmap_A"
    assert ws.cell(row=2, column=1).value == "Pin"          # header survived
    assert ws.cell(row=3, column=1).value == "A1"           # data survived
    assert len(ws.merged_cells.ranges) == 2                 # merges survived
    wb_check.close()
