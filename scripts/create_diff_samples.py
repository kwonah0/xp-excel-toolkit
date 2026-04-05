"""Generate two sample xlsx files (old / new) with known differences for manual diff testing.

Differences between old and new:
  1. Cell value change  — SENSOR_A INIT 0x00 → 0xAA (level2_common)
  2. Row insertion       — NEW_REG added after SENSOR_A registers (level2_common)
  3. Row deletion        — PLL_CFG last register removed (level2_common)
  4. Comment change      — cell (2,1) TYPE "RW2" gets a note added (level2_common)
  5. Register field      — TIMER_A PERIOD[7:0] INIT 0xFF → 0x80 (level2_buscon)
  6. MemoryMap change    — PLL_CFG comment text changed
  7. MemoryMap addition  — new entry WATCHDOG added
  8. MemoryMap removal   — PWR_MGMT entry removed
"""

from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "samples"

HEADERS = ["TYPE", "INDX", "PAGE", "PARA", "NAME",
           "D7", "D6", "D5", "D4", "D3", "D2", "D1", "D0", "INIT"]
D7_COL, D0_COL = 6, 13

STYLES = {
    "header_font": Font(name="Consolas", size=10, bold=True, color="FFFFFF"),
    "header_fill": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
    "data_font": Font(name="Consolas", size=10),
    "name_font": Font(name="Consolas", size=10, bold=True),
    "green_fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "yellow_fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "rsvd_fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "thin_border": Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    ),
    "center": Alignment(horizontal="center", vertical="center"),
}


# ── Shared register data ──────────────────────────────────────────

COMMON_REGS = [
    # SENSOR_A (INDX 57)
    ("RW2", "57", "0", "0", "SENSOR_A", [("EN",7,7),("MODE[1:0]",6,5),("RST",4,4),("RSVD",3,0)], "0x00"),
    ("RO",  "57", "0", "1", "SENSOR_A", [("BUSY",7,7),("DONE",6,6),("ERR",5,5),("CNT[4:0]",4,0)], "0x00"),
    ("RO",  "57", "0", "2", "SENSOR_A", [("DATA[15:8]",7,0)], "0xFF"),
    ("RO",  "57", "0", "3", "SENSOR_A", [("DATA[7:0]",7,0)], "0x00"),
    # AMPLIFIER (INDX 46)
    ("RW1", "46", "0", "0", "AMPLIFIER", [("GAIN[1:0]",7,6),("DIV[2:0]",5,3),("POL",2,2),("PHASE",1,1),("LOCK",0,0)], "0x24"),
    ("RW1", "46", "0", "1", "AMPLIFIER", [("TH[7:0]",7,0)], "0x80"),
    # PLL_CFG (INDX 88)
    ("RW2", "88", "0", "0", "PLL_CFG", [("PLL_EN",7,7),("BYPASS",6,6),("RSVD",5,4),("NDIV[3:0]",3,0)], "0x01"),
    ("RW2", "88", "0", "1", "PLL_CFG", [("MDIV[7:0]",7,0)], "0x10"),
    ("RO",  "88", "0", "2", "PLL_CFG", [("LOCKED",7,7),("CAL_DONE",6,6),("RSVD",5,0)], "0x00"),
]

BUSCON_REGS = [
    # TIMER_A (INDX 34)
    ("RW2", "34", "0", "0", "TIMER_A", [("TIM_EN",7,7),("ONESHOT",6,6),("PRE[1:0]",5,4),("RSVD",3,0)], "0x00"),
    ("RW2", "34", "0", "1", "TIMER_A", [("PERIOD[7:0]",7,0)], "0xFF"),
    ("RO",  "34", "1", "2", "TIMER_A", [("CNT[7:0]",7,0)], "0x00"),
    # SPI_MASTER (INDX 65)
    ("RW1", "65", "0", "0", "SPI_MASTER", [("SPI_EN",7,7),("CPOL",6,6),("CPHA",5,5),("BR[2:0]",4,2),("LSB",1,1),("RSVD",0,0)], "0x00"),
    ("RW2", "65", "0", "1", "SPI_MASTER", [("TX[7:0]",7,0)], "0x00"),
    ("RO",  "65", "0", "2", "SPI_MASTER", [("RX[7:0]",7,0)], "0x00"),
]

MEMMAP_HEADERS = ["BASEADDR", "Group", "midgroup", "Comment", "special"]

MEMMAP_BASE = [
    ("0x57", "SENSOR_A",   "analog",  "Temperature sensor A",      ""),
    ("0x46", "AMPLIFIER",  "analog",  "Programmable gain amp",      ""),
    ("0x88", "PLL_CFG",    "clock",   "PLL frequency synthesizer",  ""),
    ("0x34", "TIMER_A",    "digital", "16-bit timer/counter",       ""),
    ("0x65", "SPI_MASTER", "comm",    "SPI master controller",      ""),
    ("0x99", "PWR_MGMT",   "power",   "Power management unit",      "always_on"),
]


# ── Helper functions ──────────────────────────────────────────────

def _dx_to_col(d_bit: int) -> int:
    return D7_COL + (7 - d_bit)


def _write_header(ws):
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = STYLES["header_font"]
        c.fill = STYLES["header_fill"]
        c.border = STYLES["thin_border"]
        c.alignment = STYLES["center"]


def _write_register_row(ws, row, reg):
    type_, indx, page, para, name, bits, init = reg
    for col, val in [(1, type_), (2, indx), (3, page), (4, para)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = STYLES["data_font"]; c.border = STYLES["thin_border"]; c.alignment = STYLES["center"]
    c = ws.cell(row=row, column=5, value=name)
    c.font = STYLES["name_font"]; c.border = STYLES["thin_border"]; c.alignment = STYLES["center"]

    for col in range(D7_COL, D0_COL + 1):
        c = ws.cell(row=row, column=col)
        c.font = STYLES["data_font"]; c.border = STYLES["thin_border"]; c.alignment = STYLES["center"]

    for label, d_hi, d_lo in bits:
        col_start = _dx_to_col(d_hi)
        col_end = _dx_to_col(d_lo)
        bit_width = d_hi - d_lo + 1
        ws.cell(row=row, column=col_start).value = label
        fill = STYLES["rsvd_fill"] if label == "RSVD" else (STYLES["green_fill"] if bit_width == 1 else STYLES["yellow_fill"])
        for col in range(col_start, col_end + 1):
            ws.cell(row=row, column=col).fill = fill
        if col_end > col_start:
            ws.merge_cells(f"{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}")

    c = ws.cell(row=row, column=14, value=init)
    c.font = STYLES["data_font"]; c.border = STYLES["thin_border"]; c.alignment = STYLES["center"]


def _write_registers(ws, registers):
    _write_header(ws)
    for idx, reg in enumerate(registers):
        _write_register_row(ws, idx + 2, reg)
    ws.freeze_panes = "A2"


def _write_memmap(ws, entries):
    for i, h in enumerate(MEMMAP_HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = STYLES["header_font"]; c.fill = STYLES["header_fill"]
        c.border = STYLES["thin_border"]; c.alignment = STYLES["center"]
    for idx, entry in enumerate(entries):
        row = idx + 2
        for col, val in enumerate(entry, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = STYLES["data_font"]; c.border = STYLES["thin_border"]; c.alignment = STYLES["center"]
    ws.freeze_panes = "A2"


# ── Build OLD xlsx ────────────────────────────────────────────────

def create_old() -> Path:
    path = OUTPUT_DIR / "diff_old.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("level2_common")
    _write_registers(ws, COMMON_REGS)

    ws = wb.create_sheet("level2_buscon")
    _write_registers(ws, BUSCON_REGS)

    ws = wb.create_sheet("memorymap")
    _write_memmap(ws, MEMMAP_BASE)

    wb.save(path)
    wb.close()
    return path


# ── Build NEW xlsx (with known differences) ───────────────────────

def create_new() -> Path:
    path = OUTPUT_DIR / "diff_new.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- level2_common: modifications ---
    new_common = list(COMMON_REGS)

    # Diff 1: SENSOR_A para=0 INIT changed 0x00 → 0xAA
    r = list(new_common[0])
    r[6] = "0xAA"
    new_common[0] = tuple(r)

    # Diff 2: insert NEW_REG after SENSOR_A (after index 3)
    new_reg = ("RW1", "57", "0", "4", "SENSOR_A",
               [("NEW_BIT", 7, 7), ("RSVD", 6, 0)], "0x00")
    new_common.insert(4, new_reg)

    # Diff 3: remove PLL_CFG last register (LOCKED/CAL_DONE, was index 8, now 9 after insert)
    new_common.pop()  # remove last = PLL_CFG RO para=2

    ws = wb.create_sheet("level2_common")
    _write_registers(ws, new_common)

    # Diff 4: add comment on cell (2,1) — first data row TYPE cell
    ws.cell(row=2, column=1).comment = Comment("Changed access type for rev2", "Engineer")

    # --- level2_buscon: modifications ---
    new_buscon = list(BUSCON_REGS)

    # Diff 5: TIMER_A PERIOD[7:0] INIT 0xFF → 0x80
    r = list(new_buscon[1])
    r[6] = "0x80"
    new_buscon[1] = tuple(r)

    ws = wb.create_sheet("level2_buscon")
    _write_registers(ws, new_buscon)

    # --- memorymap: modifications ---
    new_mm = list(MEMMAP_BASE)

    # Diff 6: PLL_CFG comment changed
    idx_pll = next(i for i, e in enumerate(new_mm) if e[1] == "PLL_CFG")
    e = list(new_mm[idx_pll])
    e[3] = "PLL v2 — updated lock detector"
    new_mm[idx_pll] = tuple(e)

    # Diff 7: add WATCHDOG entry
    new_mm.append(("0xA0", "WATCHDOG", "digital", "Watchdog timer", ""))

    # Diff 8: remove PWR_MGMT
    new_mm = [e for e in new_mm if e[1] != "PWR_MGMT"]

    ws = wb.create_sheet("memorymap")
    _write_memmap(ws, new_mm)

    wb.save(path)
    wb.close()
    return path


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    old = create_old()
    new = create_new()
    print(f"Created: {old}")
    print(f"Created: {new}")
    print()
    print("Known differences:")
    print("  1. [level2_common] SENSOR_A para=0 INIT: 0x00 → 0xAA")
    print("  2. [level2_common] NEW_REG inserted (SENSOR_A para=4)")
    print("  3. [level2_common] PLL_CFG para=2 (LOCKED/CAL_DONE) removed")
    print("  4. [level2_common] Cell (2,1) comment added")
    print("  5. [level2_buscon] TIMER_A PERIOD INIT: 0xFF → 0x80")
    print("  6. [memorymap] PLL_CFG comment text changed")
    print("  7. [memorymap] WATCHDOG entry added")
    print("  8. [memorymap] PWR_MGMT entry removed")


if __name__ == "__main__":
    main()
