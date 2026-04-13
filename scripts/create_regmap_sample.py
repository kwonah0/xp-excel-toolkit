"""Generate register map specification sample (.xlsx) with diverse merge patterns.

Color scheme for bit fields (D7..D0):
  - Green:  single-bit fields (1-bit, e.g. EN, RST, BUSY)
  - Yellow: multi-bit merged fields (e.g. MODE[1:0], CNT[4:0], DATA[7:0])
  - Gray:   reserved fields (RSVD)

Vertical merges:
  - INDX: same index values merged vertically
  - PAGE: same page values merged within INDX boundary
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "samples"

# Headers: columns A(1) through N(14)
HEADERS = ["TYPE", "INDX", "PAGE", "PARA", "NAME",
           "D7", "D6", "D5", "D4", "D3", "D2", "D1", "D0", "INIT"]
NUM_COLS = len(HEADERS)  # 14

# Column indices (1-based) for bit fields D7..D0
D7_COL = 6
D0_COL = 13

# -- Sheet data: sheet_name -> list of registers ----------------------------

SHEETS = {
    "level2_common": [
        # INDX 57 / SENSOR_A (page 0: 4 regs)
        ("RW2", "57", "0", "0", "SENSOR_A", [
            ("EN", 7, 7), ("MODE[1:0]", 6, 5), ("RST", 4, 4), ("RSVD", 3, 0),
        ], "0x00"),
        ("RO", "57", "0", "1", "SENSOR_A", [
            ("BUSY", 7, 7), ("DONE", 6, 6), ("ERR", 5, 5), ("CNT[4:0]", 4, 0),
        ], "0x00"),
        ("RO", "57", "0", "2", "SENSOR_A", [
            ("DATA[15:8]", 7, 0),
        ], "0xFF"),
        ("RO", "57", "0", "3", "SENSOR_A", [
            ("DATA[7:0]", 7, 0),
        ], "0x00"),

        # INDX 46 / AMPLIFIER (page 0: 2 regs, page 1: 2 regs)
        ("RW1", "46", "0", "0", "AMPLIFIER", [
            ("GAIN[1:0]", 7, 6), ("DIV[2:0]", 5, 3), ("POL", 2, 2), ("PHASE", 1, 1), ("LOCK", 0, 0),
        ], "0x24"),
        ("RW1", "46", "0", "1", "AMPLIFIER", [
            ("TH[7:0]", 7, 0),
        ], "0x80"),
        ("RW2", "46", "1", "2", "AMPLIFIER", [
            ("OVF_EN", 7, 7), ("UDF_EN", 6, 6), ("RSVD", 5, 4),
            ("DONE_EN", 3, 3), ("ERR_EN", 2, 2), ("TRIG[1:0]", 1, 0),
        ], "0x00"),
        ("WO", "46", "1", "3", "AMPLIFIER", [
            ("OVF_CLR", 7, 7), ("UDF_CLR", 6, 6), ("RSVD", 5, 4),
            ("DONE_CLR", 3, 3), ("ERR_CLR", 2, 2), ("RSVD", 1, 0),
        ], "0x00"),

        # INDX 23 / DAC_CTRL (page 0: 3 regs)
        ("RW1", "23", "0", "0", "DAC_CTRL", [
            ("DAC_EN", 7, 7), ("RSVD", 6, 4), ("CH_SEL[2:0]", 3, 1), ("PDWN", 0, 0),
        ], "0x00"),
        ("RW2", "23", "0", "1", "DAC_CTRL", [
            ("DAC_H[7:0]", 7, 0),
        ], "0x00"),
        ("RW2", "23", "0", "2", "DAC_CTRL", [
            ("DAC_L[7:4]", 7, 4), ("RSVD", 3, 0),
        ], "0x00"),

        # INDX 12 / ADC_CONV (page 0: 2 regs, page 1: 1 reg)
        ("RW1", "12", "0", "0", "ADC_CONV", [
            ("START", 7, 7), ("CONT", 6, 6), ("AVG[1:0]", 5, 4), ("RSVD", 3, 0),
        ], "0x00"),
        ("RO", "12", "0", "1", "ADC_CONV", [
            ("ADC[7:0]", 7, 0),
        ], "0x00"),
        ("RO", "12", "1", "2", "ADC_CONV", [
            ("ADC[11:8]", 7, 4), ("RSVD", 3, 0),
        ], "0x00"),

        # INDX 88 / PLL_CFG (page 0: 3 regs)
        ("RW2", "88", "0", "0", "PLL_CFG", [
            ("PLL_EN", 7, 7), ("BYPASS", 6, 6), ("RSVD", 5, 4), ("NDIV[3:0]", 3, 0),
        ], "0x01"),
        ("RW2", "88", "0", "1", "PLL_CFG", [
            ("MDIV[7:0]", 7, 0),
        ], "0x10"),
        ("RO", "88", "0", "2", "PLL_CFG", [
            ("LOCKED", 7, 7), ("CAL_DONE", 6, 6), ("RSVD", 5, 0),
        ], "0x00"),
    ],

    "level2_buscon": [
        # INDX 91 / GPIO_PORT (page 0: 3 regs)
        ("RW1", "91", "0", "0", "GPIO_PORT", [
            ("DIR[7:0]", 7, 0),
        ], "0xFF"),
        ("RW1", "91", "0", "1", "GPIO_PORT", [
            ("OUT[7:0]", 7, 0),
        ], "0x00"),
        ("RO", "91", "0", "2", "GPIO_PORT", [
            ("IN[7:0]", 7, 0),
        ], "0x00"),

        # INDX 34 / TIMER_A (page 0: 2 regs, page 1: 2 regs)
        ("RW2", "34", "0", "0", "TIMER_A", [
            ("TIM_EN", 7, 7), ("ONESHOT", 6, 6), ("PRE[1:0]", 5, 4), ("RSVD", 3, 0),
        ], "0x00"),
        ("RW2", "34", "0", "1", "TIMER_A", [
            ("PERIOD[7:0]", 7, 0),
        ], "0xFF"),
        ("RO", "34", "1", "2", "TIMER_A", [
            ("CNT[7:0]", 7, 0),
        ], "0x00"),
        ("RW1", "34", "1", "3", "TIMER_A", [
            ("CMP_EN", 7, 7), ("CMP_IE", 6, 6), ("RSVD", 5, 0),
        ], "0x00"),

        # INDX 65 / SPI_MASTER (page 0: 3 regs)
        ("RW1", "65", "0", "0", "SPI_MASTER", [
            ("SPI_EN", 7, 7), ("CPOL", 6, 6), ("CPHA", 5, 5),
            ("BR[2:0]", 4, 2), ("LSB", 1, 1), ("RSVD", 0, 0),
        ], "0x00"),
        ("RW2", "65", "0", "1", "SPI_MASTER", [
            ("TX[7:0]", 7, 0),
        ], "0x00"),
        ("RO", "65", "0", "2", "SPI_MASTER", [
            ("RX[7:0]", 7, 0),
        ], "0x00"),

        # INDX 78 / I2C_SLAVE (page 0: 2 regs, page 1: 2 regs)
        ("RW1", "78", "0", "0", "I2C_SLAVE", [
            ("I2C_EN", 7, 7), ("GC_EN", 6, 6), ("ADDR[6:0]", 5, 0),
        ], "0x00"),
        ("RO", "78", "0", "1", "I2C_SLAVE", [
            ("RXF", 7, 7), ("TXE", 6, 6), ("BUSY", 5, 5), ("NACK", 4, 4), ("RSVD", 3, 0),
        ], "0x40"),
        ("RW2", "78", "1", "2", "I2C_SLAVE", [
            ("TX[7:0]", 7, 0),
        ], "0x00"),
        ("RO", "78", "1", "3", "I2C_SLAVE", [
            ("RX[7:0]", 7, 0),
        ], "0x00"),

        # INDX 99 / PWR_MGMT (page 0: 2 regs)
        ("RW1", "99", "0", "0", "PWR_MGMT", [
            ("SLEEP", 7, 7), ("STDBY", 6, 6), ("LDO_EN", 5, 5),
            ("RSVD", 4, 3), ("VSEL[2:0]", 2, 0),
        ], "0x20"),
        ("RO", "99", "0", "1", "PWR_MGMT", [
            ("POR", 7, 7), ("BOD", 6, 6), ("RSVD", 5, 0),
        ], "0x80"),
    ],
}


# -- Memory map data: (BASEADDR, Group, midgroup, Comment, special) ---------
MEMMAP_HEADERS = ["BASEADDR", "Group", "midgroup", "Comment", "special"]

MEMMAP_ENTRIES = [
    ("0x57", "SENSOR_A",   "analog",  "Temperature sensor A",      ""),
    ("0x46", "AMPLIFIER",  "analog",  "Programmable gain amp",      ""),
    ("0x23", "DAC_CTRL",   "analog",  "12-bit DAC output",         ""),
    ("0x12", "ADC_CONV",   "analog",  "12-bit ADC converter",      ""),
    ("0x88", "PLL_CFG",    "clock",   "PLL frequency synthesizer", ""),
    ("0x91", "GPIO_PORT",  "digital", "General purpose I/O",       ""),
    ("0x34", "TIMER_A",    "digital", "16-bit timer/counter",      ""),
    ("0x65", "SPI_MASTER", "comm",    "SPI master controller",     ""),
    ("0x78", "I2C_SLAVE",  "comm",    "I2C slave interface",       "shared"),
    ("0x99", "PWR_MGMT",   "power",   "Power management unit",     "always_on"),
]


def _dx_to_col(d_bit: int) -> int:
    """Convert D-bit number (7..0) to 1-based column index."""
    # D7 = col 6, D6 = col 7, ..., D0 = col 13
    return D7_COL + (7 - d_bit)


def _populate_sheet(ws, registers: list, styles: dict):
    """Populate a worksheet with header, data rows, merges, and formatting."""
    header_font = styles["header_font"]
    header_fill = styles["header_fill"]
    data_font = styles["data_font"]
    name_font = styles["name_font"]
    green_fill = styles["green_fill"]
    yellow_fill = styles["yellow_fill"]
    rsvd_fill = styles["rsvd_fill"]
    thin_border = styles["thin_border"]
    center = styles["center"]

    # -- Row 1: Headers
    for i, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    # -- Data rows (row 2 onwards)
    for idx, (type_, indx, page, para, name, bits, init) in enumerate(registers):
        row = idx + 2

        # TYPE
        c = ws.cell(row=row, column=1, value=type_)
        c.font = data_font; c.border = thin_border; c.alignment = center

        # INDX
        c = ws.cell(row=row, column=2, value=indx)
        c.font = data_font; c.border = thin_border; c.alignment = center

        # PAGE
        c = ws.cell(row=row, column=3, value=page)
        c.font = data_font; c.border = thin_border; c.alignment = center

        # PARA
        c = ws.cell(row=row, column=4, value=para)
        c.font = data_font; c.border = thin_border; c.alignment = center

        # NAME
        c = ws.cell(row=row, column=5, value=name)
        c.font = name_font; c.border = thin_border; c.alignment = center

        # D7..D0 bit fields — apply base border
        for col in range(D7_COL, D0_COL + 1):
            c = ws.cell(row=row, column=col)
            c.font = data_font; c.border = thin_border; c.alignment = center

        # Write bit field labels, merges, and colors
        for label, d_hi, d_lo in bits:
            col_start = _dx_to_col(d_hi)
            col_end = _dx_to_col(d_lo)
            bit_width = d_hi - d_lo + 1

            ws.cell(row=row, column=col_start).value = label

            if label == "RSVD":
                fill = rsvd_fill
            elif bit_width == 1:
                fill = green_fill
            else:
                fill = yellow_fill

            for col in range(col_start, col_end + 1):
                ws.cell(row=row, column=col).fill = fill

            if col_end > col_start:
                start_letter = get_column_letter(col_start)
                end_letter = get_column_letter(col_end)
                ws.merge_cells(f"{start_letter}{row}:{end_letter}{row}")

        # INIT
        c = ws.cell(row=row, column=14, value=init)
        c.font = data_font; c.border = thin_border; c.alignment = center

    # -- Vertical merges (auto-computed from data)
    indx_groups: dict[str, list[int]] = {}
    for idx, (_, indx, *_rest) in enumerate(registers):
        indx_groups.setdefault(indx, []).append(idx + 2)

    for rows in indx_groups.values():
        if len(rows) > 1:
            ws.merge_cells(f"B{rows[0]}:B{rows[-1]}")
            ws[f"B{rows[0]}"].alignment = center
            ws.merge_cells(f"E{rows[0]}:E{rows[-1]}")
            ws[f"E{rows[0]}"].alignment = center

    page_groups: dict[tuple[str, str], list[int]] = {}
    for idx, (_, indx, page, *_rest) in enumerate(registers):
        page_groups.setdefault((indx, page), []).append(idx + 2)

    for rows in page_groups.values():
        if len(rows) > 1:
            ws.merge_cells(f"C{rows[0]}:C{rows[-1]}")
            ws[f"C{rows[0]}"].alignment = center

    # -- Column widths
    widths = {"A": 7, "B": 7, "C": 7, "D": 7, "E": 12}
    for col in range(D7_COL, D0_COL + 1):
        widths[get_column_letter(col)] = 12
    widths["N"] = 8

    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # -- Freeze panes: header row
    ws.freeze_panes = "A2"


def _populate_memorymap(ws, styles: dict):
    """Populate the memorymap worksheet."""
    header_font = styles["header_font"]
    header_fill = styles["header_fill"]
    data_font = styles["data_font"]
    thin_border = styles["thin_border"]
    center = styles["center"]

    # Headers
    for i, h in enumerate(MEMMAP_HEADERS, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    # Data rows
    for idx, (baseaddr, group, midgroup, comment, special) in enumerate(MEMMAP_ENTRIES):
        row = idx + 2
        for col, val in enumerate([baseaddr, group, midgroup, comment, special], start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = data_font
            c.border = thin_border
            c.alignment = center

    # Column widths
    for col, w in enumerate([12, 14, 10, 30, 12], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"


# -- Overview data ---------------------------------------------------------
OVERVIEW_DATA = [
    # (col_A, col_B, col_C)
    ("#General Option", None, "Comment"),
    ("chip_name", "SENSOR_IC_V2", "Top-level chip identifier"),
    ("version", "2.1.0", "Specification version"),
    ("author", "design_team", None),
    ("date", "2025-01-15", "Last modified date"),
    ("#clock_gating", "enabled", "This option is commented out"),
    (None, None, None),  # empty row
    ("#Multi-resolution", None, None),
    ("base_resolution", "1024", "Base address resolution in bytes"),
    ("page_size", "256", None),
    ("max_pages", "4", "Maximum number of pages per index"),
    ("#extended_mode", "true", "Experimental feature, disabled"),
    (None, None, None),  # empty row
    ("#Power Management", None, "Power domain settings"),
    ("default_voltage", "1.8", "Default supply voltage (V)"),
    ("sleep_mode", "deep", "deep | light | standby"),
    ("wakeup_source", "timer", "timer | gpio | i2c"),
    ("retention", "true", "Enable register retention in sleep"),
]


def _populate_overview(ws, styles: dict):
    """Populate the overview worksheet with dict-like key-value data."""
    data_font = styles["data_font"]
    thin_border = styles["thin_border"]
    header_font = styles["header_font"]
    header_fill = styles["header_fill"]

    for idx, (a, b, c) in enumerate(OVERVIEW_DATA):
        row = idx + 1
        cell_a = ws.cell(row=row, column=1, value=a)
        cell_b = ws.cell(row=row, column=2, value=b)
        cell_c = ws.cell(row=row, column=3, value=c)

        for cell in (cell_a, cell_b, cell_c):
            cell.font = data_font
            cell.border = thin_border

        # Highlight category rows
        if a and a.startswith("#") and b is None:
            cell_a.font = header_font
            cell_a.fill = header_fill

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40


def create_regmap_xlsx() -> Path:
    """Create register map sample .xlsx with color-coded bit fields."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    path = OUTPUT_DIR / "regmap_sample.xlsx"

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Shared styles
    styles = {
        "header_font": Font(name="Consolas", size=10, bold=True, color="FFFFFF"),
        "header_fill": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
        "data_font": Font(name="Consolas", size=10),
        "name_font": Font(name="Consolas", size=10, bold=True),
        "green_fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "yellow_fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "rsvd_fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "thin_border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
        "center": Alignment(horizontal="center", vertical="center"),
    }

    for sheet_name, registers in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        _populate_sheet(ws, registers, styles)

    # -- memorymap sheet ----------------------------------------------------
    ws_mm = wb.create_sheet(title="memorymap")
    _populate_memorymap(ws_mm, styles)

    # -- overview sheet -----------------------------------------------------
    ws_ov = wb.create_sheet(title="overview")
    _populate_overview(ws_ov, styles)

    wb.save(path)
    wb.close()
    print(f"Created: {path}")
    return path


if __name__ == "__main__":
    create_regmap_xlsx()
