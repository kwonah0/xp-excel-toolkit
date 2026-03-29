"""Generate example Excel files (.xlsx and .xls) with merged cells, colors, borders."""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import xlwt


OUTPUT_DIR = Path(__file__).parent / "samples"


def create_sample_xlsx() -> Path:
    """Create sample .xlsx with merged cells, colors, and borders."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    path = OUTPUT_DIR / "sample.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "직원목록"

    # ── Styles ───────────────────────────────────────────────
    title_font = Font(name="맑은 고딕", size=14, bold=True)
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    salary_fmt = "#,##0"

    # ── Row 1: Title (merged A1:D1) ──────────────────────────
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "2024년 매출 보고서"
    c.font = title_font
    c.alignment = center

    # ── Row 2: Date ──────────────────────────────────────────
    ws["A2"] = "작성일: 2024-03-15"

    # ── Row 3: blank ─────────────────────────────────────────

    # ── Row 4: Headers ───────────────────────────────────────
    headers = ["부서", "이름", "직급", "급여"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    # ── Row 5-9: Data ────────────────────────────────────────
    data = [
        ("영업", "김철수", "과장", 3000),
        (None,   "이영희", "대리", 2500),  # merged 부서
        (None,   "박민수", "사원", 2000),  # merged 부서
        ("개발", "최지훈", "차장", 4000),
        ("개발", "한수진", "과장", 3500),
    ]

    for row_offset, (dept, name, pos, sal) in enumerate(data):
        r = 5 + row_offset
        if dept is not None:
            ws.cell(row=r, column=1, value=dept)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=pos)
        sal_cell = ws.cell(row=r, column=4, value=sal)
        sal_cell.number_format = salary_fmt

        # Apply borders
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = thin_border

    # ── Merge: 영업 department (A5:A7) ───────────────────────
    ws.merge_cells("A5:A7")
    ws["A5"].alignment = center

    # ── Yellow background for 개발 차장 row ──────────────────
    for c in range(1, 5):
        ws.cell(row=8, column=c).fill = yellow_fill

    # ── Column widths ────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12

    wb.save(path)
    wb.close()
    print(f"Created: {path}")
    return path


def create_sample_xls() -> Path:
    """Create sample .xls with merged cells, colors, and borders."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    path = OUTPUT_DIR / "sample.xls"

    wb = xlwt.Workbook()
    ws = wb.add_sheet("직원목록")

    # ── Styles ───────────────────────────────────────────────
    title_style = xlwt.easyxf(
        "font: name 맑은 고딕, height 280, bold on;"
        "alignment: horizontal center, vertical center;"
    )
    header_style = xlwt.easyxf(
        "font: name 맑은 고딕, bold on, colour white;"
        "pattern: pattern solid, fore_colour light_blue;"
        "borders: left thin, right thin, top thin, bottom thin;"
        "alignment: horizontal center, vertical center;"
    )
    normal_style = xlwt.easyxf(
        "borders: left thin, right thin, top thin, bottom thin;"
    )
    salary_style = xlwt.easyxf(
        "borders: left thin, right thin, top thin, bottom thin;",
        num_format_str="#,##0",
    )
    yellow_style = xlwt.easyxf(
        "pattern: pattern solid, fore_colour yellow;"
        "borders: left thin, right thin, top thin, bottom thin;"
    )
    yellow_salary_style = xlwt.easyxf(
        "pattern: pattern solid, fore_colour yellow;"
        "borders: left thin, right thin, top thin, bottom thin;",
        num_format_str="#,##0",
    )

    # ── Row 0: Title (merged A0:D0) ──────────────────────────
    ws.write_merge(0, 0, 0, 3, "2024년 매출 보고서", title_style)

    # ── Row 1: Date ──────────────────────────────────────────
    ws.write(1, 0, "작성일: 2024-03-15")

    # ── Row 2: blank ─────────────────────────────────────────

    # ── Row 3: Headers ───────────────────────────────────────
    headers = ["부서", "이름", "직급", "급여"]
    for i, h in enumerate(headers):
        ws.write(3, i, h, header_style)

    # ── Merge: 영업 department (row 4~6, col 0) — must be written BEFORE individual cells
    ws.write_merge(4, 6, 0, 0, "영업", normal_style)

    # ── Row 4-8: Data ────────────────────────────────────────
    data = [
        # (dept, name, pos, sal) — dept=None means already handled by merge
        (None,   "김철수", "과장", 3000),  # row 4 — dept merged
        (None,   "이영희", "대리", 2500),  # row 5 — dept merged
        (None,   "박민수", "사원", 2000),  # row 6 — dept merged
        ("개발", "최지훈", "차장", 4000),  # row 7
        ("개발", "한수진", "과장", 3500),  # row 8
    ]

    for row_offset, (dept, name, pos, sal) in enumerate(data):
        r = 4 + row_offset
        is_yellow = (row_offset == 3)  # 최지훈 row
        ns = yellow_style if is_yellow else normal_style
        ss = yellow_salary_style if is_yellow else salary_style

        if dept is not None:
            ws.write(r, 0, dept, ns)
        ws.write(r, 1, name, ns)
        ws.write(r, 2, pos, ns)
        ws.write(r, 3, sal, ss)

    # ── Column widths (xlwt uses 1/256 of character width) ───
    ws.col(0).width = 256 * 12
    ws.col(1).width = 256 * 12
    ws.col(2).width = 256 * 10
    ws.col(3).width = 256 * 12

    wb.save(str(path))
    print(f"Created: {path}")
    return path


if __name__ == "__main__":
    create_sample_xlsx()
    create_sample_xls()
